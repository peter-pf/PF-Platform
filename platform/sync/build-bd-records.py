#!/usr/bin/env python3
"""
BD Records ingest (build-bd-records.py)
=======================================
Reads the PF BD Master workbook "Organizations" + "Contacts" tabs and emits the
BD CRM base feed: companies, each with its linked contacts. The FIELD LIST for
each record is taken from ROW 4 of each tab (the spec'd field headers). We do NOT
invent fields. A contact is linked to its company by its "Organization" value
matching a company "Name".

Output: platform/data/bd-records.js
  window.PF_BD_RECORDS = {
    companies: [ { id, fields:{...row4 fields...}, contacts:[ {id, fields:{...}} ] } ],
    unlinkedContacts: [ {id, fields:{...}, orgName} ],   // contacts whose Organization
                                                          // matches no company (reported)
    companyFields: [ ...row4 headers in order... ],
    contactFields: [ ...row4 headers in order... ],
    generated, source, sourceTabs
  }

This is the READ-ONLY BASE. Manual additions + interactions live in KV (the
write-back overlay via /api/bd-record + /api/bd-interaction) and are merged on
top of this base in the UI. Re-syncing this file never erases KV data.

GATING: BD's own tool data. /data/bd-records.js -> business_dev area in
auth.js DATA_FILE_AREAS. admin + partner + business_dev only; field_ops BLOCKED.

Thread caps set BEFORE importing openpyxl (solved-box-gotchas).

Usage:
  python3 build-bd-records.py
  python3 build-bd-records.py --local sync/downloads/PF_BD_Master.xlsm
"""

# --- thread caps BEFORE any heavy import ---
import os
for _v in ("OPENBLAS_NUM_THREADS", "OMP_NUM_THREADS", "MKL_NUM_THREADS",
           "NUMEXPR_NUM_THREADS", "VECLIB_MAXIMUM_THREADS"):
    os.environ.setdefault(_v, "1")

import sys
import re
import json
import hashlib
import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
PLATFORM = os.path.dirname(HERE)
DATA_DIR = os.path.join(PLATFORM, "data")
OUT_JS = os.path.join(DATA_DIR, "bd-records.js")
DEFAULT_XLSM = os.path.join(HERE, "downloads", "PF_BD_Master.xlsm")

SOURCE_FILE = "PF_BD_Master.xlsm"
ORG_TAB = "Organizations"
CON_TAB = "Contacts"
HEADER_ROW = 4              # field list lives in row 4 (1-indexed)
NAME_COL = 2               # col C (0-indexed) holds the record Name on both tabs
CON_ORG_COL = 5            # col F (0-indexed) on Contacts holds "Organization"


def _token():
    sys.path.insert(0, "/home/aiciv/tools")
    from pf_email import _token as t
    return t()


def download_workbook():
    import urllib.request
    import urllib.parse
    tok = _token()
    drive = os.environ["SP_DRIVE_ID"]
    q = urllib.parse.quote("PF_BD_Master")
    url = "https://graph.microsoft.com/v1.0/drives/%s/root/search(q='%s')" % (drive, q)
    req = urllib.request.Request(url, headers={"Authorization": "Bearer " + tok})
    res = json.load(urllib.request.urlopen(req, timeout=120))
    fid = None
    for it in res.get("value", []):
        if str(it.get("name", "")).lower().startswith("pf_bd_master"):
            fid = it.get("id"); break
    if not fid:
        raise SystemExit("Could not find PF_BD_Master in SharePoint drive")
    curl = "https://graph.microsoft.com/v1.0/drives/%s/items/%s/content" % (drive, fid)
    creq = urllib.request.Request(curl, headers={"Authorization": "Bearer " + tok})
    return urllib.request.urlopen(creq, timeout=180).read()


def cellval(v):
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def headers_for(ws):
    """Row-4 field headers as an ordered list of (col_index, name)."""
    rows = list(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
    hdr = rows[0] if rows else ()
    out = []
    for i, c in enumerate(hdr):
        name = cellval(c)
        if name:
            out.append((i, name))
    return out


def read_records(ws, fields):
    """Each data row (from HEADER_ROW+1) with a non-empty Name -> {fields:{...}}."""
    out = []
    for r in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        if len(r) <= NAME_COL:
            continue
        name = cellval(r[NAME_COL])
        if not name:
            continue
        rec_fields = {}
        for (i, fname) in fields:
            rec_fields[fname] = cellval(r[i]) if i < len(r) else ""
        out.append(rec_fields)
    return out


def norm_key(s):
    return re.sub(r"\s+", " ", (s or "").strip()).lower()


# Strip common legal/suffix noise so "Lauth" can match "Lauth Group, Inc." etc.
# Used ONLY as a deterministic second pass AFTER an exact normalized match misses.
_SUFFIX_RE = re.compile(
    r"(,?\s+(inc|incorporated|llc|l\.l\.c|ltd|co|corp|corporation|company|"
    r"group|construction|builders|building|contractors|contracting|"
    r"development|developers|enterprises|services|holdings)\.?)+\s*$",
    re.IGNORECASE,
)


def strip_suffix(s):
    """Normalized name with trailing legal/industry suffixes + punctuation removed."""
    n = norm_key(s)
    prev = None
    # strip repeatedly so "Acme Construction Group Inc" collapses fully
    while prev != n:
        prev = n
        n = _SUFFIX_RE.sub("", n)
        n = n.strip().rstrip(",.-").strip()
    return n


def collision_id(prefix, seed, seen):
    """Stable, collision-ONLY id. First occurrence of a seed keeps the plain
    hash (so 99% of ids never move across re-syncs). The Nth duplicate of the
    SAME seed gets a deterministic '#N' suffix BEFORE hashing, so only the
    genuinely colliding records shift -- everything else stays stable."""
    n = seen.get(seed, 0) + 1
    seen[seed] = n
    eff = seed if n == 1 else (seed + "#" + str(n))
    h = hashlib.sha1(eff.encode("utf-8")).hexdigest()[:10]
    return prefix + "_" + h, (n > 1)


def build(xlsm_bytes):
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsm_bytes), data_only=True, read_only=True)

    omitted = []
    for tab in (ORG_TAB, CON_TAB):
        if tab not in wb.sheetnames:
            omitted.append("Tab missing: %r" % tab)

    org_ws = wb[ORG_TAB]
    con_ws = wb[CON_TAB]
    org_fields = headers_for(org_ws)
    con_fields = headers_for(con_ws)

    # Name field must exist on both (it is the record label + link key).
    org_field_names = [n for (_, n) in org_fields]
    con_field_names = [n for (_, n) in con_fields]
    if "Name" not in org_field_names:
        omitted.append("Organizations row 4 has no 'Name' field")
    if "Name" not in con_field_names:
        omitted.append("Contacts row 4 has no 'Name' field")
    if "Organization" not in con_field_names:
        omitted.append("Contacts row 4 has no 'Organization' field (contacts cannot be linked to companies)")

    orgs = read_records(org_ws, org_fields)
    cons = read_records(con_ws, con_fields)
    wb.close()

    # ---- companies (collision-only unique ids) + match indexes -------------
    companies = []
    by_name = {}        # exact normalized Name -> company
    by_stripped = {}    # suffix-stripped Name -> [companies]  (for fuzzy pass)
    co_seen = {}        # seed -> count, for collision-only disambiguation
    co_id_warnings = []
    for of in orgs:
        seed = norm_key(of.get("Name", ""))
        cid, collided = collision_id("co", seed, co_seen)
        if collided:
            co_id_warnings.append("Duplicate company Name (collision-suffixed id): %r" % of.get("Name", ""))
        comp = {"id": cid, "fields": of, "contacts": []}
        companies.append(comp)
        by_name[seed] = comp
        st = strip_suffix(of.get("Name", ""))
        if st:
            by_stripped.setdefault(st, []).append(comp)

    # ---- contacts: exact -> fuzzy -> unlinked ------------------------------
    ct_seen = {}        # seed -> count, for collision-only disambiguation
    unlinked = []
    fuzzy_links = []    # [{contact, matchedCompany, rule}] for Jonathan to audit
    for cf in cons:
        name = cf.get("Name", "")
        org_name = cf.get("Organization", "")
        seed = name + "||" + org_name
        cid, collided = collision_id("ct", seed, ct_seen)
        rec = {"id": cid, "fields": cf}

        comp = None
        rule = None
        org_norm = norm_key(org_name)

        if org_norm and org_norm in by_name:
            comp = by_name[org_norm]   # exact normalized match (primary)
        elif org_norm:
            # Deterministic second pass (only if Organization is non-blank).
            org_stripped = strip_suffix(org_name)
            cands = []
            seen_ids = set()

            # rule 1: exact match on suffix-stripped forms
            if org_stripped and org_stripped in by_stripped:
                for c in by_stripped[org_stripped]:
                    if c["id"] not in seen_ids:
                        cands.append((c, "stripped-exact")); seen_ids.add(c["id"])

            # rule 2: company normalized name STARTS WITH the contact org
            #         (catches regional suffixes: "Alston Construction" ->
            #          "Alston Construction - IL"). Require a word boundary so
            #          "Smith" does not match "Smithfield".
            if not cands and org_norm:
                for c in companies:
                    cn = norm_key(c["fields"].get("Name", ""))
                    if cn == org_norm:
                        continue
                    if cn.startswith(org_norm) and (
                        len(cn) == len(org_norm) or not cn[len(org_norm)].isalnum()
                    ):
                        if c["id"] not in seen_ids:
                            cands.append((c, "company-startswith-org")); seen_ids.add(c["id"])

            if len(cands) == 1:
                comp, rule = cands[0][0], cands[0][1]
                fuzzy_links.append({
                    "contact": name,
                    "contactOrg": org_name,
                    "matchedCompany": comp["fields"].get("Name", ""),
                    "rule": rule,
                })
            elif len(cands) > 1:
                # Ambiguous: do NOT guess. Leave unlinked, tag the candidates.
                rec2 = dict(rec)
                rec2["orgName"] = org_name
                rec2["ambiguousCandidates"] = [c["fields"].get("Name", "") for (c, _r) in cands]
                unlinked.append(rec2)
                continue

        if comp:
            comp["contacts"].append(rec)
        else:
            rec2 = dict(rec)
            rec2["orgName"] = org_name   # blank => genuine orphan
            unlinked.append(rec2)

    out = {
        "companies": companies,
        "unlinkedContacts": unlinked,
        "fuzzyLinks": fuzzy_links,
        "companyFields": org_field_names,
        "contactFields": con_field_names,
        "generated": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source": SOURCE_FILE,
        "sourceTabs": [ORG_TAB, CON_TAB],
    }
    if omitted:
        out["_omitted"] = omitted
    return out, omitted, co_id_warnings


def write_js(data):
    note = (
        "// AUTO-GENERATED by sync/build-bd-records.py — DO NOT EDIT BY HAND.\n"
        "// Source: %s :: '%s' + '%s' tabs  (field list = row 4)  generated %s\n"
        "// BD CRM BASE (read-only). Manual additions + interactions live in KV\n"
        "// (/api/bd-record + /api/bd-interaction) and merge on top in the UI.\n"
        "// ACCESS: business_dev — admin/partner/business_dev (field_ops BLOCKED).\n"
        % (SOURCE_FILE, ORG_TAB, CON_TAB, data["generated"])
    )
    body = "window.PF_BD_RECORDS = " + json.dumps(data, indent=2, ensure_ascii=False) + ";\n"
    with open(OUT_JS, "w") as f:
        f.write(note + body)
    linked = sum(len(c["contacts"]) for c in data["companies"])
    nunlinked = len(data["unlinkedContacts"])
    ncon = linked + nunlinked
    print("wrote", OUT_JS)
    print("  companies: %d" % len(data["companies"]))
    print("  contacts: %d (linked %d, unlinked %d)" % (ncon, linked, nunlinked))
    print("  fuzzy links (auditable): %d" % len(data.get("fuzzyLinks", [])))
    for fl in data.get("fuzzyLinks", []):
        print("    [%s] %r (org %r) -> %r"
              % (fl["rule"], fl["contact"], fl["contactOrg"], fl["matchedCompany"]))
    nambig = sum(1 for u in data["unlinkedContacts"] if u.get("ambiguousCandidates"))
    if nambig:
        print("  ambiguous (left unlinked, candidates tagged): %d" % nambig)
        for u in data["unlinkedContacts"]:
            if u.get("ambiguousCandidates"):
                print("    %r (org %r) -> candidates %s"
                      % (u["fields"].get("Name", ""), u.get("orgName", ""), u["ambiguousCandidates"]))


def main():
    if "--local" in sys.argv:
        path = sys.argv[sys.argv.index("--local") + 1]
    elif os.path.exists(DEFAULT_XLSM):
        path = DEFAULT_XLSM
    else:
        path = None
    if path:
        with open(path, "rb") as f:
            xlsm = f.read()
        print("source:", path)
    else:
        xlsm = download_workbook()
        print("source: SharePoint download")

    data, omitted, co_id_warnings = build(xlsm)
    write_js(data)
    if omitted:
        for o in omitted:
            print("  OMITTED:", o)
    else:
        print("  OMITTED: none (all spec'd tabs + fields present)")

    # ID-collision report (companies + contacts).
    if co_id_warnings:
        for w in co_id_warnings:
            print("  ID WARNING:", w)
    else:
        print("  ID WARNING: none (no duplicate company Name)")

    # Hard verify: every emitted id (companies + all contacts) is UNIQUE.
    all_ids = [c["id"] for c in data["companies"]]
    for c in data["companies"]:
        all_ids += [ct["id"] for ct in c["contacts"]]
    all_ids += [u["id"] for u in data["unlinkedContacts"]]
    dupes = set(x for x in all_ids if all_ids.count(x) > 1)
    if dupes:
        print("  ID UNIQUENESS: FAIL — %d colliding ids: %s" % (len(dupes), list(dupes)))
        sys.exit(1)
    print("  ID UNIQUENESS: OK — %d ids all unique" % len(all_ids))


if __name__ == "__main__":
    main()
