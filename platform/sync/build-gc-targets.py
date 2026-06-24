#!/usr/bin/env python3
"""
GC targets ingest (build-gc-targets.py)
=======================================
Reads the PF BD Master "GC's to Contact" tab and emits the target-account list
subdivided by sector. INSPECTED STRUCTURE (verified, NOT assumed):

  The tab is a MATRIX, not a columnar record table:
    - ROW 2 holds the SECTOR / list names as COLUMN HEADERS
      (e.g. "Best guess ENR top 100", "Top ENR by Indiana", "Industrial",
       "IN apartments", "OH apartments", "Wind Turbine GCs", "GC's", "Ray's List").
    - Each COLUMN below row 2 is a vertical list of GC names for that sector.
    - Col 0 is a 1..N rank index (layout scaffolding, NOT data).
    - One column (col 3) has a BLANK row-2 header but holds data (an OH regional
      continuation). We DO NOT drop it: it gets a fallback label and the blank
      header is reported in _omitted.
    - Some GC cells embed a " - location" note; we keep the full cell as Name and
      also split a Note where a dash separator is present.

Output: platform/data/gc-targets.js
  window.PF_GC_TARGETS = {
    sectors: [ { name, gcs:[ {id, fields:{Name, Note}} ] } ],
    fields: ['Name','Note'],
    generated, source, sourceTab
  }

Collision-safe ids (op-style: first occurrence keeps the plain hash, Nth dup of a
seed hashes seed#N). A GC can appear in several sectors; ids are unique PER
(sector,name) so the same GC in two sectors gets two distinct ids.

GATING: BD's own tool data. /data/gc-targets.js -> business_dev. field_ops BLOCKED.

Thread caps set BEFORE importing openpyxl (solved-box-gotchas).

Usage:
  python3 build-gc-targets.py
  python3 build-gc-targets.py --local sync/downloads/PF_BD_Master.xlsm
"""

# --- thread caps BEFORE any heavy import ---
import os
for _v in ("OPENBLAS_NUM_THREADS", "OMP_NUM_THREADS", "MKL_NUM_THREADS",
           "NUMEXPR_NUM_THREADS", "VECLIB_MAXIMUM_THREADS"):
    os.environ.setdefault(_v, "1")

import sys
import re
import json
import hashlib
import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
PLATFORM = os.path.dirname(HERE)
DATA_DIR = os.path.join(PLATFORM, "data")
OUT_JS = os.path.join(DATA_DIR, "gc-targets.js")
DEFAULT_XLSM = os.path.join(HERE, "downloads", "PF_BD_Master.xlsm")

SOURCE_FILE = "PF_BD_Master.xlsm"
GC_TAB = "GCs to Contact"
SECTOR_HEADER_ROW = 2      # row 2 holds the sector/list names (column headers)
RANK_COL = 0               # col 0 is the 1..N rank index (skip)
# A name/note separator: en dash, em dash, or hyphen surrounded by spaces.
NOTE_SPLIT_RE = re.compile(r"\s+[–—-]\s+")


def _token():
    sys.path.insert(0, "/home/aiciv/tools")
    from pf_email import _token as t
    return t()


def download_workbook():
    import urllib.request
    import urllib.parse
    tok = _token()
    drive = os.environ["SP_DRIVE_ID"]
    q = urllib.parse.quote("PF_BD_Master")
    url = "https://graph.microsoft.com/v1.0/drives/%s/root/search(q='%s')" % (drive, q)
    req = urllib.request.Request(url, headers={"Authorization": "Bearer " + tok})
    res = json.load(urllib.request.urlopen(req, timeout=120))
    fid = None
    for it in res.get("value", []):
        if str(it.get("name", "")).lower().startswith("pf_bd_master"):
            fid = it.get("id"); break
    if not fid:
        raise SystemExit("Could not find PF_BD_Master in SharePoint drive")
    curl = "https://graph.microsoft.com/v1.0/drives/%s/items/%s/content" % (drive, fid)
    creq = urllib.request.Request(curl, headers={"Authorization": "Bearer " + tok})
    return urllib.request.urlopen(creq, timeout=180).read()


def cellval(v):
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def collision_id(prefix, seed, seen):
    n = seen.get(seed, 0) + 1
    seen[seed] = n
    eff = seed if n == 1 else (seed + "#" + str(n))
    h = hashlib.sha1(eff.encode("utf-8")).hexdigest()[:10]
    return prefix + "_" + h, (n > 1)


def split_name_note(text):
    """A GC cell may embed a ' - location' note. Keep the full Name, split a Note."""
    parts = NOTE_SPLIT_RE.split(text, 1)
    if len(parts) == 2 and parts[1].strip():
        return parts[0].strip(), parts[1].strip()
    return text, ""


def build(xlsm_bytes):
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsm_bytes), data_only=True, read_only=True)

    omitted = []
    if GC_TAB not in wb.sheetnames:
        wb.close()
        return ({
            "sectors": [], "fields": ["Name", "Note"],
            "generated": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "source": SOURCE_FILE, "sourceTab": GC_TAB,
            "_omitted": ["Tab missing: %r" % GC_TAB],
        }, ["Tab missing: %r" % GC_TAB], [])

    ws = wb[GC_TAB]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    ncols = max((len(r) for r in rows), default=0)
    header = rows[SECTOR_HEADER_ROW - 1] if len(rows) >= SECTOR_HEADER_ROW else ()

    sectors = []
    seen = {}
    id_warnings = []
    total = 0
    for j in range(1, ncols):  # skip col 0 (rank index)
        raw_hdr = cellval(header[j]) if j < len(header) else ""
        if raw_hdr:
            sector_name = raw_hdr
        else:
            # data under a blank header -> do not drop; label + report.
            sector_name = "Column %d (no sector header)" % j
        gcs = []
        for i in range(SECTOR_HEADER_ROW, len(rows)):   # data starts BELOW row 2
            v = cellval(rows[i][j]) if j < len(rows[i]) else ""
            if not v:
                continue
            name, note = split_name_note(v)
            seed = (sector_name + "||" + v).lower()
            gid, collided = collision_id("gc", seed, seen)
            if collided:
                id_warnings.append("Duplicate GC entry within a sector: %r in %r" % (v, sector_name))
            gcs.append({"id": gid, "fields": {"Name": name, "Note": note}})
            total += 1
        if gcs:  # only emit sectors that actually have GCs
            if not raw_hdr:
                omitted.append("Sector column %d has a BLANK row-2 header (%d GCs kept under fallback label)"
                               % (j, len(gcs)))
            sectors.append({"name": sector_name, "gcs": gcs})

    out = {
        "sectors": sectors,
        "fields": ["Name", "Note"],
        "generated": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source": SOURCE_FILE,
        "sourceTab": GC_TAB,
        "totalGcs": total,
    }
    if omitted:
        out["_omitted"] = omitted
    return out, omitted, id_warnings


def write_js(data):
    note = (
        "// AUTO-GENERATED by sync/build-gc-targets.py — DO NOT EDIT BY HAND.\n"
        "// Source: %s :: '%s' tab  (MATRIX: row 2 = sector headers, columns = GC lists)\n"
        "// generated %s\n"
        "// GC target accounts by sector. ACCESS: business_dev — admin/partner/\n"
        "// business_dev (field_ops BLOCKED).\n"
        % (SOURCE_FILE, GC_TAB, data["generated"])
    )
    body = "window.PF_GC_TARGETS = " + json.dumps(data, indent=2, ensure_ascii=False) + ";\n"
    with open(OUT_JS, "w") as f:
        f.write(note + body)
    print("wrote", OUT_JS)
    print("  sectors: %d, total GC entries: %d" % (len(data["sectors"]), data.get("totalGcs", 0)))
    for sx in data["sectors"]:
        print("    - %s: %d" % (sx["name"], len(sx["gcs"])))


def main():
    if "--local" in sys.argv:
        path = sys.argv[sys.argv.index("--local") + 1]
    elif os.path.exists(DEFAULT_XLSM):
        path = DEFAULT_XLSM
    else:
        path = None
    if path:
        with open(path, "rb") as f:
            xlsm = f.read()
        print("source:", path)
    else:
        xlsm = download_workbook()
        print("source: SharePoint download")

    data, omitted, id_warnings = build(xlsm)
    write_js(data)
    if omitted:
        for o in omitted:
            print("  OMITTED/NOTE:", o)
    else:
        print("  OMITTED: none")
    if id_warnings:
        print("  ID WARNINGS: %d duplicate (sector,name) entries collision-suffixed" % len(id_warnings))
    # verify id uniqueness
    ids = [g["id"] for sx in data["sectors"] for g in sx["gcs"]]
    dupes = set(x for x in ids if ids.count(x) > 1)
    if dupes:
        print("  ID UNIQUENESS: FAIL — %s" % list(dupes)); sys.exit(1)
    print("  ID UNIQUENESS: OK — %d ids all unique" % len(ids))


if __name__ == "__main__":
    main()
