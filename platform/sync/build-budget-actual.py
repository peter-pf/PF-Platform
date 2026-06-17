#!/usr/bin/env python3
"""
Build Budget vs Actual — POET (26-002)   [Internal ops job-cost tracker]
=======================================================================
Parses the "Budget vs Actual" sheet of POET's Turnover Budget workbook and
writes platform/data/budget-actual-poet.js (window.PF_BUDGET_ACTUAL_POET = {...}),
which the per-project Budget vs Actual view (under Project Management) reads.

This is FINANCIAL data — it belongs in the project record's financial area under
Project Management, NEVER in the field-facing fo-projects view.

SOURCE (real, no fabrication):
  SharePoint:
    04 - Project Management/02 - Projects/26-002 - POET Projects - POET/
      26-0330 POET Turnover Budget.xlsm   sheet "Budget vs Actual"
  Structure (confirmed live 2026-06-17):
    Row 1: Name: | POET Bioprocessing | Job #: | 26-002
    Row 2: Location: | (Project Location)
    Row 3/4: header band -> A "Cost Code" B "Description" C "Budget" (Precon Budget)
             D "Actual Costs" E "Variance" F "Vendor or Supplier" G "Notes"
    Rows 5..~133: detail. Category header rows carry the category name (col A or B)
             with the category subtotal in C/D/E. Cost-code line rows carry a code
             in col A (e.g. 5051) and description in col B.
    Rows ~138..151: a clean summary band (per-category + grand total).
             Row "Total Construction Contract" = grand total (Budget/Actual/Variance).

Rules:
  - Variance is RECOMPUTED = budget - actual (do not trust the sheet's stored value).
  - Skip blank/zero-only filler rows sensibly, but KEEP real budgeted lines even when
    actual is zero (a $500 budgeted, $0 actual line is real and must show).
  - NEVER fabricate. Blank stays blank.

Invoice links (clickable supporting docs for actual-cost lines):
  Resolved from the POET project folder (Invoicing / Vendors). A line gets a SPECIFIC
  invoice link only when a document is reasonably matched (vendor name / amount).
  Where no specific doc is matched, the line gets a link to the relevant invoice/vendor
  FOLDER so the user can still pull one up. If nothing is found, NO link is emitted.
  Matched 2026-06-17:
    - 5710 Reprographics  $271.46  -> "Order Success _ FedEx Office.pdf" (exact $271.46 total)
  Folder fallbacks: Invoicing folder; Vendors folder.

Reuses Graph auth + drive patterns from build-project-record.py
(token via /home/aiciv/tools/pf_email.py _token(); SP_DRIVE_ID from /home/aiciv/.env).
Thread caps set BEFORE openpyxl import (this box ~300 pid limit; solved-box-gotchas).

Usage:
  python3 build-budget-actual.py            # writes data/budget-actual-poet.js
  python3 build-budget-actual.py --dump     # print assembled record, no write
"""

# --- thread caps BEFORE any heavy import (solved-box-gotchas) ---
import os
for _v in ("OPENBLAS_NUM_THREADS", "OMP_NUM_THREADS", "MKL_NUM_THREADS",
           "NUMEXPR_NUM_THREADS", "VECLIB_MAXIMUM_THREADS"):
    os.environ.setdefault(_v, "1")

import io
import re
import sys
import json
import urllib.request
import urllib.parse
import urllib.error
from datetime import datetime, timezone

# --- reuse the Graph token helper + env loader from pf_email.py ---
sys.path.insert(0, "/home/aiciv/tools")
from pf_email import _token, _env  # noqa: E402

GRAPH = "https://graph.microsoft.com/v1.0"

HERE = os.path.dirname(os.path.abspath(__file__))
PLATFORM = os.path.dirname(HERE)
DATA_DIR = os.path.join(PLATFORM, "data")
OUT_JS = os.path.join(DATA_DIR, "budget-actual-poet.js")

PROJECT_NUMBER = "26-002"
PROJECT_NAME = "POET Biosciences"
SP_PROJECT_FOLDER = "04 - Project Management/02 - Projects/26-002 - POET Projects - POET"
BUDGET_FILE = f"{SP_PROJECT_FOLDER}/26-0330 POET Turnover Budget.xlsm"
BUDGET_SHEET = "Budget vs Actual"

# Invoicing / Vendors folders (for clickable supporting docs)
INVOICING_FOLDER = f"{SP_PROJECT_FOLDER}/02 - Project Management/Invoicing"
VENDORS_FOLDER = f"{SP_PROJECT_FOLDER}/02 - Project Management/Vendors"

_env_cache = _env()
DRIVE_ID = _env_cache.get("SP_DRIVE_ID", "")


# ---------------- Graph helpers (pattern from build-project-record.py) ----------------
def gget(token, url):
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    return json.loads(urllib.request.urlopen(req).read())


def get_item_by_path(token, path):
    """Return driveItem metadata (incl webUrl) for a folder/file path, or None."""
    p = urllib.parse.quote(path)
    url = f"{GRAPH}/drives/{DRIVE_ID}/root:/{p}"
    try:
        return gget(token, url)
    except urllib.error.HTTPError:
        return None


def download_path(token, path):
    p = urllib.parse.quote(path)
    url = f"{GRAPH}/drives/{DRIVE_ID}/root:/{p}:/content"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    return urllib.request.urlopen(req).read()


def weburl(token, path):
    """webUrl for a folder/file, or '' if not found."""
    item = get_item_by_path(token, path)
    if item and item.get("webUrl"):
        return item["webUrl"]
    return ""


# ---------------- value helpers ----------------
def _txt(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s.upper() in ("N/A", "NA", "TBD"):
        return ""
    return s


def _num(v):
    """Return a float for a numeric/dollar cell, or None if not numeric.
    Blank cells -> None (rendered blank, never 0)."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return None
    s2 = s.replace("$", "").replace(",", "").strip()
    # parenthesised negatives e.g. (1,082.85)
    neg = False
    if s2.startswith("(") and s2.endswith(")"):
        neg = True
        s2 = s2[1:-1]
    try:
        f = float(s2)
        return -f if neg else f
    except ValueError:
        return None


def _round2(x):
    return None if x is None else round(x + 0.0, 2)


# ---------------- parse ----------------
# Cost-code line rows have a numeric code in col A (e.g. "5051", "5710").
CODE_RX = re.compile(r"^=?\d{3,4}$")  # some cells are typed "=5110" -> normalize

# A row is a same-sheet ROLL-UP when its budget formula sums/adds OTHER rows of
# THIS sheet (e.g. "=SUM(C56:C61)", "=C63+C68+C71+C78", "=C55"). A LEAF row pulls
# its value from another sheet or a literal (e.g. "='OH Calcs'!G31", "=0").
ROLLUP_RX = re.compile(r"^=(SUM\(C\d|C\d+[+\-]|C\d+$)", re.I)
# A SUB-rollup uses SUM of a contiguous leaf range, e.g. "=SUM(C56:C61)".
SUMRANGE_RX = re.compile(r"^=SUM\(C\d+:C\d+\)$", re.I)
# Extract the same-sheet C-row references a formula points at.
CREF_RX = re.compile(r"C(\d+)")


def _norm_code(a):
    a = (a or "").strip()
    if a.startswith("="):
        a = a[1:].strip()
    return a


def parse_budget_actual(token):
    """Parse the Budget vs Actual sheet faithfully, using the cell FORMULAS to
    distinguish category / sub-rollup / leaf rows so we never double-count.
    Two workbook loads: data_only (values) + formulas (structure)."""
    import openpyxl
    raw = download_path(token, BUDGET_FILE)
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, keep_vba=False)
    wbf = openpyxl.load_workbook(io.BytesIO(raw), data_only=False, keep_vba=False)
    if BUDGET_SHEET not in wb.sheetnames:
        raise SystemExit(f"ERROR: sheet '{BUDGET_SHEET}' not found in {BUDGET_FILE}")
    ws = wb[BUDGET_SHEET]
    wsf = wbf[BUDGET_SHEET]

    def cell(r, c):
        return ws.cell(row=r, column=c).value

    def fmla(r, c):
        v = wsf.cell(row=r, column=c).value
        return str(v) if v is not None else ""

    # ---- header (Name / Job # / Location) ----
    job = _txt(cell(1, 4)) or PROJECT_NUMBER
    name = _txt(cell(1, 2)) or PROJECT_NAME
    location = _txt(cell(2, 2))
    if location.lower() in ("project location", ""):
        location = ""  # placeholder text -> blank, never fabricated

    # ---- find the detail header row + the summary band ----
    header_row = None
    for r in range(1, min(ws.max_row, 12) + 1):
        if _txt(cell(r, 1)).lower() == "cost code":
            header_row = r
            break
    if header_row is None:
        header_row = 4

    summary_row = None
    for r in range(header_row + 1, ws.max_row + 1):
        if _txt(cell(r, 3)).lower().startswith("estimating budget"):
            summary_row = r
            break
    detail_end = (summary_row - 1) if summary_row else ws.max_row

    # ---- two-pass classification via the FORMULA GRAPH (deterministic) ----
    # Pass 1: for every detail row, record formula + whether it's a same-sheet rollup,
    #         and which same-sheet C-rows its formula references.
    rows_meta = {}
    for r in range(header_row + 1, detail_end + 1):
        cf = fmla(r, 3).replace(" ", "")
        is_rollup = bool(ROLLUP_RX.match(cf))
        refs = [int(x) for x in CREF_RX.findall(cf)] if cf.startswith("=") else []
        # only same-sheet refs that fall in the detail band
        refs = [x for x in refs if header_row < x <= detail_end]
        rows_meta[r] = {"cf": cf, "is_rollup": is_rollup, "refs": refs}

    rollup_rows = {r for r, m in rows_meta.items() if m["is_rollup"]}

    # A MAJOR CATEGORY = a rollup row that references OTHER ROLLUP rows (rollup-of-
    # rollups), i.e. its children are themselves subtotals. These are exactly the 7
    # top categories (rows 5,54,62,81,102,123,131) that tie to the summary band.
    # (Row 54 "=C55" and row 131 "=C132" reference a single sub-rollup -> still major.)
    major_rows = set()
    for r in rollup_rows:
        child_refs = rows_meta[r]["refs"]
        if child_refs and all(c in rollup_rows for c in child_refs):
            major_rows.add(r)

    # Pass 2: walk rows, opening a group on each MAJOR row; everything else is either
    # a SUB-ROLLUP (rollup, not major -> shown as sub-header, value NOT summed) or a
    # LEAF cost line (counted).
    groups = []
    cur = None

    def flush():
        nonlocal cur
        if cur is not None:
            groups.append(cur)
        cur = None

    for r in range(header_row + 1, detail_end + 1):
        a_raw = _txt(cell(r, 1))
        a = _norm_code(a_raw)
        b = _txt(cell(r, 2))
        budget = _num(cell(r, 3))
        actual = _num(cell(r, 4))
        vendor = _txt(cell(r, 6))
        notes = _txt(cell(r, 7))
        is_code = bool(CODE_RX.match(a_raw))
        m = rows_meta[r]

        # spacer / fully empty row
        if not a and not b and budget is None and actual is None:
            continue

        if r in major_rows:
            flush()
            title = b or a or "(Category)"
            cur = {
                "title": title,
                "rows": [],
                "subtotal": {
                    "budget": _round2(budget),
                    "actual": _round2(actual),
                    "variance": _round2((budget or 0.0) - (actual or 0.0)),
                },
            }
            continue

        if cur is None:
            cur = {"title": "(Uncategorized)", "rows": [],
                   "subtotal": {"budget": None, "actual": None, "variance": None}}

        # SUB-ROLLUP (rollup but not major): sub-header inside category. Not summed.
        if m["is_rollup"]:
            cur["rows"].append({
                "cost_code": a if is_code else "",
                "description": b or a,
                "budget": _round2(budget),
                "actual": _round2(actual),
                "variance": _round2((budget or 0.0) - (actual or 0.0)),
                "vendor": vendor,
                "notes": notes,
                "is_subtotal": True,
            })
            continue

        # LEAF cost line. Skip pure-noise (no money, no description, no code).
        if (budget in (None, 0)) and (actual in (None, 0)) and not b:
            continue
        variance = None
        if budget is not None or actual is not None:
            variance = _round2((budget or 0.0) - (actual or 0.0))
        cur["rows"].append({
            "cost_code": a if is_code else "",
            "description": b,
            "budget": _round2(budget),
            "actual": _round2(actual),
            "variance": variance,
            "vendor": vendor,
            "notes": notes,
            "is_subtotal": False,
        })
    flush()

    groups = [g for g in groups
              if g["rows"] or any(v not in (None, 0) for v in g["subtotal"].values())]

    # ---- grand total from the summary band ("Total Construction Contract") ----
    grand = {"budget": None, "actual": None, "variance": None, "label": ""}
    if summary_row:
        for r in range(summary_row, ws.max_row + 1):
            lbl = _txt(cell(r, 2))
            if lbl.lower().startswith("total construction contract"):
                gb = _num(cell(r, 3))
                ga = _num(cell(r, 4))
                grand = {
                    "label": lbl,
                    "budget": _round2(gb),
                    "actual": _round2(ga),
                    "variance": _round2((gb or 0.0) - (ga or 0.0)),
                }
                break

    return {
        "job": job,
        "name": name,
        "location": location,
        "groups": groups,
        "grand_total": grand,
        "source_file": "26-0330 POET Turnover Budget.xlsm",
        "source_sheet": BUDGET_SHEET,
    }


# ---------------- invoice links ----------------
def resolve_invoices(token, record):
    """Attach a supporting-document link to actual-cost lines where determinable.

    Strategy (never fabricate):
      1. SPECIFIC file match (vendor/amount) -> link the file's webUrl, match='file'.
      2. Otherwise, link the relevant FOLDER (Invoicing/Vendors) so the user can pull
         a doc -> match='folder'.
      3. If neither folder resolves -> no link.
    Only ACTUAL-cost lines (actual not None and != 0) get a link; budgeted-but-unspent
    lines get none (no cost incurred yet -> no invoice).
    """
    invoicing_url = weburl(token, INVOICING_FOLDER)
    vendors_url = weburl(token, VENDORS_FOLDER)

    # Verified specific-file matches (label normalized by cost_code).
    # FedEx Office order total == $271.46 == 5710 Reprographics actual. Exact match.
    fedex_path = f"{VENDORS_FOLDER}/Order Success _ FedEx Office.pdf"
    fedex_url = weburl(token, fedex_path)

    SPECIFIC = {}
    if fedex_url:
        SPECIFIC["5710"] = {
            "url": fedex_url,
            "label": "FedEx Office order ($271.46)",
            "how": "Exact amount match: FedEx Office order total $271.46 = 5710 Reprographics actual $271.46",
        }

    matched = []
    for grp in record["groups"]:
        for row in grp["rows"]:
            # roll-up sub-header rows are not invoiceable cost lines
            if row.get("is_subtotal"):
                row["invoice"] = None
                continue
            actual = row.get("actual")
            if actual is None or actual == 0:
                row["invoice"] = None
                continue
            code = row.get("cost_code", "")
            spec = SPECIFIC.get(code)
            if spec:
                row["invoice"] = {"url": spec["url"], "label": spec["label"], "match": "file"}
                matched.append({"cost_code": code, "description": row["description"],
                                "actual": actual, "match": "file",
                                "doc": spec["label"], "how": spec["how"]})
                continue
            # folder fallback: invoicing for PF-billing-adjacent, vendors otherwise.
            # We point cost lines at the Invoicing folder primarily (it holds payment
            # records); fall back to Vendors if Invoicing didn't resolve.
            url = invoicing_url or vendors_url
            if url:
                folder_label = "Invoicing folder" if invoicing_url else "Vendors folder"
                row["invoice"] = {"url": url, "label": folder_label, "match": "folder"}
                matched.append({"cost_code": code, "description": row["description"],
                                "actual": actual, "match": "folder",
                                "doc": folder_label,
                                "how": "No discrete vendor invoice file in folder; linked to folder so user can pull one"})
            else:
                row["invoice"] = None
                matched.append({"cost_code": code, "description": row["description"],
                                "actual": actual, "match": "none",
                                "doc": "", "how": "No invoice file or folder resolved -> no link"})
    record["_invoice_match_log"] = matched
    record["_invoice_folders"] = {"invoicing": invoicing_url, "vendors": vendors_url}
    return record


# ---------------- assemble + write ----------------
def assemble(token):
    record = parse_budget_actual(token)
    record = resolve_invoices(token, record)
    record["project_number"] = PROJECT_NUMBER
    record["generated"] = datetime.now(timezone.utc).isoformat() + "Z"
    record["data_note"] = (
        "Internal job-cost tracker. Parsed from the POET Turnover Budget workbook "
        "(sheet 'Budget vs Actual'). Variance recomputed = Budget - Actual. Blanks "
        "stay blank (never fabricated). Actual-cost lines link to a supporting vendor "
        "invoice where matched, otherwise to the Invoicing/Vendors folder. Once "
        "QuickBooks is connected, actuals will flow in automatically (QB export); until "
        "then this reflects the PM-maintained Turnover Budget."
    )
    return record


def write_js(record):
    # Strip the internal log keys from the shipped data file (keep it lean + clean).
    clean = {k: v for k, v in record.items() if not k.startswith("_")}
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(OUT_JS, "w") as f:
        f.write("// AUTO-GENERATED by sync/build-budget-actual.py — do not edit by hand.\n")
        f.write("// POET (26-002) Budget vs Actual — internal job-cost tracker.\n")
        f.write("// Source: 26-0330 POET Turnover Budget.xlsm, sheet 'Budget vs Actual'.\n")
        f.write("// Variance recomputed = Budget - Actual. Blanks stay blank.\n")
        f.write("window.PF_BUDGET_ACTUAL_POET = ")
        json.dump(clean, f, indent=2)
        f.write(";\n")
    return OUT_JS


def _fmt(x):
    return "" if x is None else f"{x:,.2f}"


def _summary(record):
    print(f"Budget vs Actual — {record['name']} (Job {record['job']})  "
          f"location={record['location'] or '(blank)'}")
    print(f"Source: {record['source_file']} / {record['source_sheet']}")
    sum_b = sum_a = sum_v = 0.0          # sum of CATEGORY SUBTOTALS
    leaf_b = leaf_a = 0.0                # sum of LEAF lines only (true reconcile)
    print("\nGROUPS:")
    for g in record["groups"]:
        st = g["subtotal"]
        print(f"\n  == {g['title']} ==  subtotal  Budget {_fmt(st['budget'])}  "
              f"Actual {_fmt(st['actual'])}  Variance {_fmt(st['variance'])}  "
              f"({len(g['rows'])} lines)")
        if st["budget"] is not None:
            sum_b += st["budget"]
        if st["actual"] is not None:
            sum_a += st["actual"]
        if st["variance"] is not None:
            sum_v += st["variance"]
        for row in g["rows"]:
            if not row.get("is_subtotal"):
                leaf_b += row["budget"] or 0.0
                leaf_a += row["actual"] or 0.0
            inv = row.get("invoice")
            tag = ""
            if inv:
                tag = f"   [INV {inv['match']}: {inv['label']}]"
            marker = "  +" if row.get("is_subtotal") else "   "
            print(f"   {marker}{row['cost_code']:>5}  {row['description'][:36]:<36} "
                  f"B {_fmt(row['budget']):>12}  A {_fmt(row['actual']):>12}  "
                  f"V {_fmt(row['variance']):>12}{tag}")
    gt = record["grand_total"]
    print("\nGRAND TOTAL (from summary band):")
    print(f"  {gt.get('label','Total')}:  Budget {_fmt(gt['budget'])}  "
          f"Actual {_fmt(gt['actual'])}  Variance {_fmt(gt['variance'])}")
    print("\nRECONCILE:")
    print(f"  sum of category subtotals:  Budget {_fmt(round(sum_b,2))}  "
          f"Actual {_fmt(round(sum_a,2))}  Variance {_fmt(round(sum_v,2))}")
    print(f"  sum of LEAF lines only:     Budget {_fmt(round(leaf_b,2))}  "
          f"Actual {_fmt(round(leaf_a,2))}")
    if gt["budget"] is not None:
        db = round(sum_b - gt["budget"], 2)
        da = round(sum_a - (gt["actual"] or 0), 2)
        lb = round(leaf_b - gt["budget"], 2)
        la = round(leaf_a - (gt["actual"] or 0), 2)
        ok = abs(db) < 0.05 and abs(da) < 0.05
        print(f"  category-subtotals vs grand total: Budget {db:+.2f}  Actual {da:+.2f}"
              f"  -> {'RECONCILES' if ok else 'CHECK'}")
        print(f"  leaf-sum vs grand total:           Budget {lb:+.2f}  Actual {la:+.2f}"
              f"  -> {'RECONCILES' if abs(lb) < 0.05 and abs(la) < 0.05 else 'CHECK'}")
    print("\nINVOICE MATCH LOG:")
    for m in record["_invoice_match_log"]:
        print(f"  [{m['match']:>6}] {m['cost_code']} {m['description'][:32]:<32} "
              f"actual {_fmt(m['actual'])}  -> {m['doc']}  ({m['how'][:60]})")
    print(f"\nInvoice folders: invoicing={'OK' if record['_invoice_folders']['invoicing'] else 'MISS'} "
          f"vendors={'OK' if record['_invoice_folders']['vendors'] else 'MISS'}")


def main():
    dump = "--dump" in sys.argv[1:]
    if not DRIVE_ID:
        print("ERROR: SP_DRIVE_ID not set in /home/aiciv/.env", file=sys.stderr)
        sys.exit(1)
    print(f"build-budget-actual (POET) — {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC",
          file=sys.stderr)
    token = _token()
    record = assemble(token)
    _summary(record)
    if dump:
        clean = {k: v for k, v in record.items() if not k.startswith("_")}
        print(json.dumps(clean, indent=2))
        return
    out = write_js(record)
    print(f"\nWrote: {out}", file=sys.stderr)


if __name__ == "__main__":
    main()
