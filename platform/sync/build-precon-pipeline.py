#!/usr/bin/env python3
"""
build-precon-pipeline.py — Preconstruction pipeline (all stage buckets) for the PF Platform.

Reads EVERY row from the Project Bid Log (BOTH the "Agg Pier Bid Log" and
"Helical Pier Bid Log" sheets) and buckets each job into a Preconstruction
stage by the COL-A SECTION it sits under — i.e. the stage block (sub-table)
in the spreadsheet, NOT the per-row Bid Status. Brad's directive: bucket by
section. Writes data/precon-pipeline.js:

  window.PF_PRECON = {
    ap: { actively_bidding:[], budget_pricing:[], feasibility_review:[],
          submitted_bids:[], awarded:[], not_awarded:[] },
    hp: { ...same six buckets... },
    generated, source, source_url,
    uncategorized: [ {discipline, status, number, name, section} ]
  }

Consumed by the 12 Preconstruction stage panels in index.html (precon-ap-*,
precon-hp-*).

Source of truth: SharePoint `01 - Admin / 13 - Master Spreadsheets / Project Bid Log.xlsx`.
  - The two sheets have DIFFERENT header-row positions and column indices, so we
    DETECT each sheet's header row (the row containing "Project Number"/"Bid Status")
    and map every column BY NAME. Nothing is hardcoded by index.

HOW SECTIONS WORK (verified):
  Each sheet stacks several sub-tables. A col-A SECTION HEADER row holds a stage
  label in column A while the Project Number / Project Name columns are empty.
  That section applies to every job beneath it until the next section header.
  The FIRST block (between the field-header row and the first explicit section
  header) is IMPLICITLY "Actively Bidding" (there is no literal header row — jobs
  just sit at the top). Section headers are detected ROBUSTLY (no hardcoded rows).

SECTION LABEL -> STAGE BUCKET (case-insensitive, contains-match):
  - "actively bidding"  OR implicit top block -> actively_bidding
  - "budget pricing"                           -> budget_pricing
  - "feasibility"                              -> feasibility_review (none today)
  - contains "submit" (e.g. "Submitted Bids",
      "Bid Submitted - woohoo LJ!")            -> submitted_bids
  - contains "award" but NOT "not award"       -> awarded
  - contains "not award"                       -> not_awarded
  - any other section label                    -> uncategorized (reported, never dropped)

A row counts as a job if it has a Project Number. The `completed` flag is set
when a row's Bid Status text starts with "Completed".

Never fabricate — blank cells stay blank.

Run with thread caps BEFORE importing openpyxl (Box pid/thread guard):
  OMP_NUM_THREADS=1 OPENBLAS_NUM_THREADS=1 MKL_NUM_THREADS=1 python3 build-precon-pipeline.py
"""

import os
# Cap BLAS / OpenMP threads before any heavy import (openpyxl/numpy) — Box thread guard.
for _v in ("OMP_NUM_THREADS", "OPENBLAS_NUM_THREADS", "MKL_NUM_THREADS",
           "NUMEXPR_NUM_THREADS", "VECLIB_MAXIMUM_THREADS"):
    os.environ.setdefault(_v, "1")

import sys
import json
import urllib.request
from datetime import datetime, timezone, date

sys.path.insert(0, "/home/aiciv/tools")
import pf_email  # noqa: E402  (token + .env reader)

import openpyxl  # noqa: E402

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.normpath(os.path.join(HERE, "..", "data"))
DL_DIR = os.path.join(HERE, "downloads")

BID_LOG_ITEM = "016ISVH6Y7M7KQIB5C5FDLNKI5H3IZFXRK"
BID_LOG_WEBURL = (
    "https://pierfoundations.sharepoint.com/sites/pf/_layouts/15/Doc.aspx?"
    "sourcedoc=%7B04D5671F-A207-46E9-B6A9-1D3ED192DE2A%7D&file=Project%20Bid%20Log.xlsx"
    "&action=default&mobileredirect=true"
)

# Sheets to scan -> discipline key for the output payload.
SHEETS = [
    ("Agg Pier Bid Log", "ap"),
    ("Helical Pier Bid Log", "hp"),
]

# The six stage buckets, in pipeline order.
BUCKETS = [
    "actively_bidding",
    "budget_pricing",
    "feasibility_review",
    "submitted_bids",
    "awarded",
    "not_awarded",
]

# project_number -> showModule id of a BUILT project record.
# Only POET (26-002) has a built record today. Add ids here as records ship.
RECORD_LINKS = {
    "26-002": "project-poet",
}


def _token():
    return pf_email._token()


def _drive():
    return pf_email._env()["SP_DRIVE_ID"]


def download_bid_log():
    os.makedirs(DL_DIR, exist_ok=True)
    fp = os.path.join(DL_DIR, "Project_Bid_Log.xlsx")
    url = "https://graph.microsoft.com/v1.0/drives/%s/items/%s/content" % (_drive(), BID_LOG_ITEM)
    req = urllib.request.Request(url, headers={"Authorization": "Bearer " + _token()})
    with urllib.request.urlopen(req, timeout=120) as r, open(fp, "wb") as f:
        f.write(r.read())
    print("  downloaded Project_Bid_Log.xlsx (%d bytes)" % os.path.getsize(fp))
    return fp


def find_header_row(ws, scan=14):
    """Return the 1-indexed row that contains both 'Project Number' and 'Bid Status'."""
    for r in range(1, scan + 1):
        joined = " ".join(
            str(ws.cell(row=r, column=c).value or "") for c in range(1, 40)
        )
        if "Bid Status" in joined and "Project Number" in joined:
            return r
    return None


def header_map(ws, hr):
    """{header_text: 1-indexed col} for the header row."""
    m = {}
    for c in range(1, 40):
        v = ws.cell(row=hr, column=c).value
        if v not in (None, ""):
            m[str(v).strip()] = c
    return m


def clean(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def clean_date(v):
    """Dates come back as datetime/date; format YYYY-MM-DD. Strings pass through trimmed."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def num_or_none(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# Structural / non-job text that appears in the Project Name column. Both sheets
# stack several sub-tables (one per stage), each prefixed by a "Project Information"
# title row and a repeated header row ("Project Number / Project Name / Bid Status").
# These are layout scaffolding, NOT jobs, and must never be emitted.
NON_JOB_NAMES = {"project name", "project information", "bid log",
                 "pier foundations bid log"}

def section_to_bucket(label):
    """Map a col-A SECTION label to a stage bucket via case-insensitive contains-match.
    Returns a bucket key, or None if the label is unrecognized (-> uncategorized)."""
    low = (label or "").strip().lower()
    if not low:
        return None
    if "not award" in low:
        return "not_awarded"
    if "award" in low:
        return "awarded"
    if "submit" in low:               # "Submitted Bids", "Bid Submitted - woohoo LJ!"
        return "submitted_bids"
    if "budget pricing" in low:
        return "budget_pricing"
    if "feasibility" in low:
        return "feasibility_review"
    if "actively bidding" in low:
        return "actively_bidding"
    return None


def is_section_header(col_a, number, name):
    """A section header row: col A holds a NON-NUMERIC stage label while the
    Project Name column is empty (or only carries the 'Project Information'
    title) AND there is no Project Number. Detected ROBUSTLY — no hardcoded rows.

    NOTE on this spreadsheet's reality: col A is the per-row sequence index
    ("Number"), so JOB rows carry a numeric col A (1, 2, 3 …). The dedicated
    "Project Number" column is sparsely filled (mostly only in Awarded), so we
    do NOT require it to identify a job — a section header is distinguished by a
    TEXT col-A label with no project name beneath it on that same row."""
    a = (col_a or "").strip()
    if not a:
        return False
    # Job rows carry a numeric per-row index in col A; section labels are text.
    if a.replace(".", "", 1).isdigit():
        return False
    # A row with a real Project Number is a job, not a header.
    if (number or "").strip():
        return False
    # A header row has no Project Name (or only the "Project Information" title).
    nm = (name or "").strip().lower()
    if nm and nm not in NON_JOB_NAMES:
        return False
    return True


def extract(ws, disc_key):
    """Return (buckets_dict, uncategorized_list) for one sheet."""
    buckets = {b: [] for b in BUCKETS}
    uncategorized = []

    hr = find_header_row(ws)
    if not hr:
        print("  WARN: no header row found in '%s' — skipping" % ws.title)
        return buckets, uncategorized
    h = header_map(ws, hr)

    c_num = h.get("Project Number")
    c_name = h.get("Project Name")
    c_city = h.get("City / State")
    c_gc = h.get("General Contractor")
    c_stat = h.get("Bid Status")
    c_val = h.get("Bid Total Value")
    c_due = h.get("Due Date")
    c_invite = h.get("Invite Date")

    if not (c_num and c_name):
        print("  WARN: '%s' missing key columns (number/name) — skipping" % ws.title)
        return buckets, uncategorized

    counted = 0
    data_rows = 0
    unknown_sections = {}  # raw label -> count, for reporting
    # The FIRST block (between the field-header row and the first explicit section
    # header) is IMPLICITLY "Actively Bidding". Seed the section accordingly.
    current_section_raw = "Actively Bidding"   # display label of the block we're in
    current_bucket = "actively_bidding"        # resolved bucket (None = unmapped section)

    for r in range(hr + 1, ws.max_row + 1):
        col_a = clean(ws.cell(row=r, column=1).value)
        number = clean(ws.cell(row=r, column=c_num).value) if c_num else ""
        name = clean(ws.cell(row=r, column=c_name).value) if c_name else ""
        status_raw = ws.cell(row=r, column=c_stat).value if c_stat else None

        # Section header? -> switch the current section/bucket and move on.
        if is_section_header(col_a, number, name):
            current_section_raw = col_a
            current_bucket = section_to_bucket(col_a)
            if current_bucket is None:
                unknown_sections[col_a] = unknown_sections.get(col_a, 0) + 1
            continue

        # A row counts as a job if it carries project content. The "Project Number"
        # column is sparse on these sheets, so we treat a row as a job when it has a
        # Project Name (the populated identifier in practice) OR an explicit Project
        # Number. Pure spacer rows (no number, no name) are skipped.
        if not name and not number:
            continue  # blank/spacer row
        # Defensive: skip any structural scaffolding that slipped through by name.
        if name.lower() in NON_JOB_NAMES:
            continue

        data_rows += 1
        # completed flag: Bid Status text starts with "Completed".
        completed = clean(status_raw).lower().startswith("completed")

        if current_bucket is None:
            # Job lives under an unmapped section -> uncategorized (never dropped).
            uncategorized.append({
                "discipline": disc_key,
                "status": clean(status_raw),
                "number": number,
                "name": name,
                "section": current_section_raw,
            })
            continue

        job = {
            "number": number,
            "name": name,
            "city_state": clean(ws.cell(row=r, column=c_city).value) if c_city else "",
            "gc": clean(ws.cell(row=r, column=c_gc).value) if c_gc else "",
            "value": num_or_none(ws.cell(row=r, column=c_val).value) if c_val else None,
            "due_date": clean_date(ws.cell(row=r, column=c_due).value) if c_due else "",
            "invite_date": clean_date(ws.cell(row=r, column=c_invite).value) if c_invite else "",
            "completed": completed,
            "record": RECORD_LINKS.get(number),  # showModule id or None
        }
        buckets[current_bucket].append(job)
        counted += 1

    print("  '%s' (hdr row %d): %d data rows -> %d bucketed, %d uncategorized" %
          (ws.title, hr, data_rows, counted, len(uncategorized)))
    for b in BUCKETS:
        if buckets[b]:
            print("      %-18s %d" % (b, len(buckets[b])))
    if unknown_sections:
        for lbl, _ in unknown_sections.items():
            print("      UNKNOWN SECTION (jobs -> uncategorized): %r" % lbl)
    return buckets, uncategorized


def main():
    print("PF Preconstruction pipeline build — %s UTC" %
          datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"))
    print("Authenticating + downloading bid log...")
    fp = download_bid_log()
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)

    out = {"ap": {b: [] for b in BUCKETS}, "hp": {b: [] for b in BUCKETS}}
    uncategorized = []

    for sheet_name, disc_key in SHEETS:
        if sheet_name not in wb.sheetnames:
            print("  WARN: sheet '%s' not present — skipping" % sheet_name)
            continue
        buckets, unc = extract(wb[sheet_name], disc_key)
        out[disc_key] = buckets
        uncategorized.extend(unc)

    payload = {
        "generated": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M") + " UTC",
        "source": "SharePoint/01 - Admin/13 - Master Spreadsheets/Project Bid Log.xlsx",
        "source_url": BID_LOG_WEBURL,
        "ap": out["ap"],
        "hp": out["hp"],
        "uncategorized": uncategorized,
    }

    os.makedirs(DATA_DIR, exist_ok=True)
    out_path = os.path.join(DATA_DIR, "precon-pipeline.js")
    with open(out_path, "w") as f:
        f.write("// AUTO-GENERATED by sync/build-precon-pipeline.py — do not edit by hand.\n")
        f.write("// Preconstruction pipeline: every Project Bid Log row bucketed by its col-A SECTION (stage block).\n")
        f.write("window.PF_PRECON = ")
        json.dump(payload, f, indent=2)
        f.write(";\n")

    # Summary
    print("\nWrote %s" % out_path)
    for disc in ("ap", "hp"):
        total = sum(len(out[disc][b]) for b in BUCKETS)
        print("  %s total: %d" % (disc, total))
        for b in BUCKETS:
            print("      %-18s %d" % (b, len(out[disc][b])))
    print("  uncategorized: %d" % len(uncategorized))
    if uncategorized:
        for u in uncategorized:
            print("      [%s] status=%-22r section=%-16s '%s'" %
                  (u["discipline"], u["status"], u.get("section", ""), u["name"]))


if __name__ == "__main__":
    main()
