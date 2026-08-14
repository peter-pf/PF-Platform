#!/usr/bin/env python3
"""
Build Turnover-Stone data  ->  platform/data/turnover-stone.js
==============================================================
Per-project "Estimated Stone (TN)" for the office General Info section, pulled
from each project's TURNOVER BUDGET workbook (NOT the Bid Log, NOT the Garbin
Prelim). Brad 2026-08-14:

    "General Info 'Estimated Stone (TN)' should pull from the project's turnover
     doc -- the '<project> Turnover Budget.xlsm' under
     04 - Project Management/02 - Projects/<project>/. Read the 'Detailed Budget'
     tab, find the 'Tons of Stone' line, copy that value into Estimated Stone (TN).
     Make it per-project generalizable."

WHAT IT READS (verified against 26-002 POET, 2026-08-14):
    - Tab  'Detailed Budget '   (NOTE the trailing space; matched case-insensitively
                                 via a `detail.*budget` regex so the space + any casing
                                 survives).
    - The row whose label cell == "Tons of Stone" (col G on POET). The VALUE is read
      from the FIRST numeric cell to the RIGHT of the label on that same row (col H on
      POET); the unit is read from the next cell to the right (col I = "TONS"), for
      provenance only. This is LABEL-ANCHORED + right-scan, not a hardcoded cell, so it
      survives column/row shifts between projects.

FORMULA / STALE-CACHE GOTCHA (solved-openpyxl-stales-formula-cache):
    The "Tons of Stone" cell is a FORMULA (POET base H15 = `=H9*0.21`, CO variant
    H15 = `=H9*0.21+41`). openpyxl `data_only=True` returns Excel's CACHED result from
    the workbook's LAST SAVE. That is a valid READ as long as the workbook was saved in
    Excel (POET's cache is present + sane: base 3570.63, CO 4195.01). BUT if the cached
    value is None (workbook edited by openpyxl / never Excel-saved -> cache dropped), we
    DO NOT write a blank/zero and DO NOT fabricate -- we FLAG the project with
    status 'formula-no-cache' and leave stone_tn null so the frontend renders the honest
    amber blank. Fail-closed, never a wrong number.

WHICH WORKBOOK (multi-file jobs):
    A project can carry the BASE Turnover Budget PLUS a change-order variant
    (e.g. "26-0709 POET Turnover Budget w add'l Bin CO.xlsm"). The CURRENT live budget is
    the one the invoice-coding workflow writes to = the MOST-RECENTLY-MODIFIED file. We
    reuse the EXACT resolver from build-budget-actuals.py (most-recent wins, manifest
    override honored) so Estimated Stone tracks the same workbook the Actuals do.

OUTPUT: platform/data/turnover-stone.js
    window.PF_TURNOVER_STONE = {
      projects: { "26-002": {stone_tn, unit, cell, tab, source_file, webUrl, status}, ... },
      meta: {...}
    };
    A project with no turnover doc / no tab / no line / no cache simply gets an entry
    with stone_tn=null and a status explaining why (or NO entry when there's no folder).
    The frontend fail-closes to blank -- it never fabricates.

Reuses the Graph token + env loader from /home/aiciv/tools/pf_email.py and the
Turnover-Budget resolver conventions from build-budget-actuals.py.

Thread caps are set BEFORE openpyxl import (this box ~300 pid limit; solved-box-gotchas).

Usage:
    python3 build-turnover-stone.py                # writes data/turnover-stone.js
    python3 build-turnover-stone.py --dump         # print extracted data, no write
    python3 build-turnover-stone.py --only 26-002  # restrict to one project number
"""

# --- thread caps BEFORE any heavy import (solved-box-gotchas) ---
import os
for _v in ("OPENBLAS_NUM_THREADS", "OMP_NUM_THREADS", "MKL_NUM_THREADS",
           "NUMEXPR_NUM_THREADS", "VECLIB_MAXIMUM_THREADS"):
    os.environ.setdefault(_v, "1")

import io
import re
import sys
import json
import urllib.request
import urllib.parse
import urllib.error
from datetime import datetime, timezone

# --- reuse the Graph token helper + env loader from pf_email.py ---
sys.path.insert(0, "/home/aiciv/tools")
from pf_email import _token, _env  # noqa: E402

GRAPH = "https://graph.microsoft.com/v1.0"

HERE = os.path.dirname(os.path.abspath(__file__))
PLATFORM = os.path.dirname(HERE)
DATA_DIR = os.path.join(PLATFORM, "data")
OUT_JS = os.path.join(DATA_DIR, "turnover-stone.js")

# Same folder conventions as build-budget-actuals.py (the sibling Turnover-Budget sync).
ACTIVE_PARENT = "04 - Project Management/02 - Projects"
COMPLETED_PARENT = "04 - Project Management/02 - Projects/001 - Completed Projects"
# Optional manifest to pin a specific folder/file for a job (SHARED with build-budget-actuals).
MANIFEST = os.path.join(HERE, "turnover-budget-manifest.json")

TURNOVER_RX = re.compile(r"turnover\s*budget", re.I)
# The Detailed Budget tab is named 'Detailed Budget ' (trailing space) on POET; match
# any casing/spacing via a loose regex so the space + capitalization variants survive.
DETAILED_TAB_RX = re.compile(r"detail.*budget", re.I)
# The line we copy. Label-anchored (survives row/column shifts).
STONE_LABEL_RX = re.compile(r"\btons?\s*of\s*stone\b", re.I)

# Only real project folders start with an "NN-NNN" number.
PROJNUM_RE = re.compile(r"^(\d{2}-\d{3})\b")

_env_cache = _env()
DRIVE_ID = _env_cache.get("SP_DRIVE_ID", "")


# ---------------- Graph helpers (pattern from build-budget-actuals.py) ----------------
def gget(token, url):
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    return json.loads(urllib.request.urlopen(req).read())


def list_children(token, path):
    """Return child driveItems for a folder path, or [] if missing. Handles paging."""
    p = urllib.parse.quote(path)
    url = f"{GRAPH}/drives/{DRIVE_ID}/root:/{p}:/children?$top=200"
    items = []
    try:
        while url:
            data = gget(token, url)
            items.extend(data.get("value", []))
            url = data.get("@odata.nextLink")
    except urllib.error.HTTPError:
        return []
    return items


def download_path(token, path):
    p = urllib.parse.quote(path)
    url = f"{GRAPH}/drives/{DRIVE_ID}/root:/{p}:/content"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    return urllib.request.urlopen(req).read()


def get_item_by_path(token, path):
    """Fetch a single driveItem (for its webUrl). Returns item dict or None on 404."""
    p = urllib.parse.quote(path)
    url = f"{GRAPH}/drives/{DRIVE_ID}/root:/{p}"
    try:
        return gget(token, url)
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return None
        raise


# ---------------- Turnover-Budget resolver (mirrors build-budget-actuals.py) ----------
def _year_candidates(job):
    m = re.match(r"^(\d{2})-", job)
    if not m:
        return []
    yy = int(m.group(1))
    y = 2000 + yy
    return [str(y), str(y - 1), str(y + 1)]


def _folder_matches(child_name, job):
    n = child_name.strip()
    return n.startswith(job + " ") or n.startswith(job + "-") or n == job


def _load_manifest():
    try:
        with open(MANIFEST) as f:
            return json.load(f)
    except (FileNotFoundError, ValueError):
        return {}


def resolve_workbook(token, job, manifest):
    """Return (folder_path, file_path, source, status). status in
    ok|no-folder|no-workbook. MOST-RECENTLY-MODIFIED workbook wins on multi-file jobs
    (matches build-budget-actuals.py so Estimated Stone tracks the same live workbook the
    Actuals do). Never guesses a filename."""
    # 0. manifest override (shared with the budget-actuals sync)
    if job in manifest:
        m = manifest[job]
        folder = m.get("folder", "")
        f = m.get("file", "")
        if folder and f:
            return folder, f"{folder}/{f}", "manifest", "ok"

    # 1. active parent
    folder = None
    src = None
    for child in list_children(token, ACTIVE_PARENT):
        if child.get("folder") and _folder_matches(child.get("name", ""), job):
            folder = f"{ACTIVE_PARENT}/{child['name']}"
            src = "active"
            break

    # 2. completed parent (per year)
    if not folder:
        for yr in _year_candidates(job):
            yr_parent = f"{COMPLETED_PARENT}/{yr}"
            for child in list_children(token, yr_parent):
                if child.get("folder") and _folder_matches(child.get("name", ""), job):
                    folder = f"{yr_parent}/{child['name']}"
                    src = f"completed/{yr}"
                    break
            if folder:
                break

    if not folder:
        return None, None, None, "no-folder"

    # 3. Turnover Budget workbook inside the folder (most-recent wins on ties)
    matches = [c for c in list_children(token, folder)
               if c.get("file") and TURNOVER_RX.search(c.get("name", ""))
               and c.get("name", "").lower().endswith((".xlsm", ".xlsx"))]
    if len(matches) == 0:
        return folder, None, src, "no-workbook"
    if len(matches) == 1:
        return folder, f"{folder}/{matches[0]['name']}", src, "ok"

    def _mtime(c):
        return c.get("lastModifiedDateTime") or ""

    def _ctime(c):
        fsi = c.get("fileSystemInfo") or {}
        return fsi.get("createdDateTime") or c.get("createdDateTime") or ""
    ranked = sorted(matches, key=lambda c: (_mtime(c), _ctime(c), c.get("name", "")),
                    reverse=True)
    return folder, f"{folder}/{ranked[0]['name']}", src, "ok"


# ---------------- value helpers ----------------
def _num(v):
    """Coerce a cell value to a JSON-friendly number (int when whole), else None."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return int(f) if f == int(f) else f
    s = str(v).strip().replace(",", "")
    if not s:
        return None
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except ValueError:
        return None


# ---------------- Excel extraction (label-anchored 'Tons of Stone' -> right-scan) ------
def extract_stone(raw_bytes):
    """Read the 'Tons of Stone' value from the Detailed Budget tab of a Turnover Budget
    workbook. Returns a dict:
        {status, stone_tn, unit, cell, tab}
    status in:
        ok                -> found + a sane cached numeric value
        no-tab            -> no Detailed Budget tab
        no-line           -> tab present but no 'Tons of Stone' label
        no-value          -> label found but no numeric value to its right
        formula-no-cache  -> label found, value cell is a FORMULA whose cached result is
                             absent (needs an Excel open+save) -> NEVER fabricated; blank.
    NO fabrication: any not-ok status yields stone_tn=None (frontend renders blank)."""
    import openpyxl
    # data_only=True -> Excel's cached results (the stone cell is a formula).
    wbd = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
    tab = next((s for s in wbd.sheetnames if DETAILED_TAB_RX.search(str(s))), None)
    if tab is None:
        return {"status": "no-tab", "stone_tn": None, "unit": "", "cell": "", "tab": ""}
    wsd = wbd[tab]

    # Locate the 'Tons of Stone' label cell.
    label_r = label_c = None
    for r in range(1, wsd.max_row + 1):
        for c in range(1, wsd.max_column + 1):
            v = wsd.cell(row=r, column=c).value
            if v is not None and STONE_LABEL_RX.search(str(v)):
                label_r, label_c = r, c
                break
        if label_r is not None:
            break
    if label_r is None:
        return {"status": "no-line", "stone_tn": None, "unit": "", "cell": "", "tab": tab}

    from openpyxl.utils import get_column_letter
    # Scan RIGHT of the label for the first numeric value; unit = next non-empty to its right.
    stone_val = None
    val_col = None
    for c in range(label_c + 1, wsd.max_column + 1):
        n = _num(wsd.cell(row=label_r, column=c).value)
        if n is not None:
            stone_val = n
            val_col = c
            break

    if val_col is None:
        # Was the value cell a FORMULA whose cache is missing? Check the un-cached view so we
        # can distinguish "genuinely no value" from "formula present, cache dropped".
        wbf = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=False)
        wsf = wbf[tab]
        for c in range(label_c + 1, min(label_c + 6, wsf.max_column) + 1):
            fv = wsf.cell(row=label_r, column=c).value
            if isinstance(fv, str) and fv.startswith("="):
                cellref = f"{get_column_letter(c)}{label_r}"
                return {"status": "formula-no-cache", "stone_tn": None, "unit": "",
                        "cell": cellref, "tab": tab}
        return {"status": "no-value", "stone_tn": None, "unit": "",
                "cell": f"{get_column_letter(label_c)}{label_r}", "tab": tab}

    unit = ""
    for c in range(val_col + 1, min(val_col + 3, wsd.max_column) + 1):
        uv = wsd.cell(row=label_r, column=c).value
        if uv is not None and str(uv).strip():
            unit = str(uv).strip()
            break

    cellref = f"{get_column_letter(val_col)}{label_r}"
    return {"status": "ok", "stone_tn": stone_val, "unit": unit, "cell": cellref, "tab": tab}


# ---------------- per-project build ----------------
def resolve_projects(token, only=None):
    """Enumerate active project folders and yield (project_number, folder_name)."""
    for it in list_children(token, ACTIVE_PARENT):
        if not it.get("folder"):
            continue
        name = str(it.get("name", ""))
        m = PROJNUM_RE.match(name)
        if not m:
            continue
        projnum = m.group(1)
        if only and projnum != only:
            continue
        yield projnum, name


def build(token, only=None, verbose=True):
    manifest = _load_manifest()
    projects = {}
    report = []  # (projnum, status, detail)
    for projnum, folder_name in resolve_projects(token, only=only):
        folder, wbpath, src, status = resolve_workbook(token, projnum, manifest)
        if status != "ok" or not wbpath:
            report.append((projnum, status, folder_name))
            continue
        try:
            raw = download_path(token, wbpath)
        except (urllib.error.HTTPError, urllib.error.URLError) as e:
            report.append((projnum, "download-error", f"{wbpath}: {e}"))
            continue
        try:
            ext = extract_stone(raw)
        except Exception as e:  # noqa: BLE001 -- never let one bad file break the whole sync
            report.append((projnum, "parse-error", f"{wbpath}: {str(e)[:160]}"))
            continue

        src_name = os.path.basename(wbpath)
        item = get_item_by_path(token, wbpath) or {}
        entry = {
            "stone_tn": ext["stone_tn"],
            "unit": ext.get("unit", ""),
            "tab": ext.get("tab", ""),
            "cell": ext.get("cell", ""),
            "source_file": src_name,
            "source_path": wbpath,
            "webUrl": item.get("webUrl", "") or "",
            "status": ext["status"],
        }
        projects[projnum] = entry
        report.append((projnum, ext["status"],
                       f"{src_name} [{ext.get('tab','')}!{ext.get('cell','')}] "
                       f"= {ext['stone_tn']} {ext.get('unit','')}"))

    data = {
        "projects": projects,
        "meta": {
            "generated": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            "source": ("Each project's '*Turnover Budget*.xlsm' (most-recently-modified on "
                       "multi-file jobs), tab 'Detailed Budget', label-anchored 'Tons of "
                       "Stone' -> first numeric cell to its right. QuickBooks-independent."),
            "note": ("stone_tn is Excel's CACHED formula result (the Tons of Stone cell is a "
                     "formula). status='formula-no-cache' => cache dropped (needs Excel "
                     "open+save); stone_tn left null, NEVER fabricated. Fail-closed."),
            "project_count": len(projects),
            "ok_count": sum(1 for e in projects.values() if e["status"] == "ok"),
        },
    }
    if verbose:
        print("Turnover-Stone sync:")
        for projnum, status, detail in report:
            print(f"  {projnum}: {status}  {detail}")
        print(f"  -> {len(projects)} project(s) with a turnover doc; "
              f"{data['meta']['ok_count']} with a stone value")
    return data


def write_js(data):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(OUT_JS, "w") as f:
        f.write("// AUTO-GENERATED by sync/build-turnover-stone.py -- do not edit by hand.\n")
        f.write("// Per-project 'Estimated Stone (TN)' pulled from each project's Turnover\n")
        f.write("// Budget .xlsm, 'Detailed Budget' tab, 'Tons of Stone' line (cached formula\n")
        f.write("// value). Fail-closed: null when it can't be read (frontend renders blank).\n")
        f.write("window.PF_TURNOVER_STONE = ")
        json.dump(data, f, indent=2)
        f.write(";\n")
    return OUT_JS


def main():
    only = None
    dump = False
    args = sys.argv[1:]
    i = 0
    while i < len(args):
        a = args[i]
        if a == "--dump":
            dump = True
        elif a == "--only" and i + 1 < len(args):
            only = args[i + 1]
            i += 1
        i += 1

    if not DRIVE_ID:
        print("ERROR: SP_DRIVE_ID not set in /home/aiciv/.env", file=sys.stderr)
        return 2
    token = _token()
    data = build(token, only=only)
    if dump:
        print(json.dumps(data, indent=2))
        return 0
    path = write_js(data)
    print(f"Wrote {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
