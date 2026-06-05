#!/usr/bin/env python3
"""
Extract clean project + bid data from the PF master spreadsheets into
data/projects-data.js (single source of truth the dashboard renders).

Sources:
  PF Project Master.xlsx  -> "2026 WIP & Completed Projects"  (WIP + Completed)
  Project Bid Log.xlsx    -> "Agg Pier Bid Log"               (Outstanding/Upcoming/Recent bids)

Run as part of the daily 5pm SharePoint sync. For now reads local masters in
portal_uploads/pf-masters/. Computes dashboard KPIs in Python so the frontend
stays dumb and there is exactly one place truth is derived.

Terminology (Jonathan, FINAL 6/5):
  Outstanding Bids = bids submitted / in-flight (bid log: Submitted + Submitted-Edging), count + total value
  Pipeline Value   = AWARDED projects not yet completed (WIP tab)
  Completed        = completed projects in the FY (calendar year)
  FY26 Revenue     = completed-in-26 OR AR  (we use completed contract value + AR)
  Total FY26 Sales = FY26 Revenue + FY26 Pipeline Value
  Upcoming Bids    = bid log entries with bid-due date within next 7 days
  Recent bids      = bid log Submitted + Submitted-Edging only
"""
import os, json, datetime
import openpyxl

MASTERS = "/home/aiciv/portal_uploads/pf-masters"
OUT = "/home/aiciv/PF-Platform/platform/data/projects-data.js"
PROJ_XLSX = os.path.join(MASTERS, "PF Project Master.xlsx")
BID_XLSX = os.path.join(MASTERS, "Project Bid Log.xlsx")

# As-of date drives "upcoming bids in next 7 days". Override with PF_ASOF=YYYY-MM-DD.
ASOF = os.getenv("PF_ASOF")
ASOF = datetime.date.fromisoformat(ASOF) if ASOF else datetime.date(2026, 6, 5)


def s(v):
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    return str(v).strip()


def num(v):
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return 0.0


def read_project_rows(ws, r0, r1):
    """Read project rows from a section using the standard column map."""
    cols = {
        2: "projectNo", 3: "status", 4: "name", 5: "scope", 6: "address",
        7: "cityState", 8: "value", 9: "paid", 10: "unpaid", 11: "projectedIncome",
        12: "invoiceDue", 13: "retainPct", 14: "retainAmt", 15: "retainSubmitted",
        16: "retainPaid", 17: "pctComplete", 18: "gc", 19: "gcAddress",
        20: "gcPmName", 21: "gcPmPhone", 22: "gcPmEmail", 23: "gcSuperName",
        24: "gcSuperPhone", 25: "gcSuperEmail", 26: "startDate", 27: "durationDays",
        28: "completionDate", 29: "incomePerDay",
    }
    money = {"value", "paid", "unpaid", "projectedIncome", "retainAmt", "incomePerDay"}
    out = []
    for r in range(r0, r1 + 1):
        rec = {}
        for c, key in cols.items():
            v = ws.cell(r, c).value
            rec[key] = num(v) if key in money else s(v)
        rec["pctComplete"] = num(ws.cell(r, 17).value)  # fraction 0-1
        rec["retainPct"] = num(ws.cell(r, 13).value)
        rec["durationDays"] = num(ws.cell(r, 27).value)
        if not rec["projectNo"] and not rec["name"]:
            continue
        # normalize multiline address junk
        for k in ("address", "gcAddress", "cityState"):
            rec[k] = " ".join(rec[k].split())
        out.append(rec)
    return out


def main():
    wb = openpyxl.load_workbook(PROJ_XLSX, data_only=True)
    ws = wb["2026 WIP & Completed Projects"]

    # Sections located by reading the dump: WIP rows 5-15, Completed rows 28-33.
    wip = read_project_rows(ws, 5, 20)
    completed = read_project_rows(ws, 28, 58)

    # Classify WIP: phase = 'active' (work started / FE & mobilizing) vs 'awarded' (LOI/contract, future)
    for p in wip:
        st = p["status"].lower()
        active = p["pctComplete"] > 0 or "subcontract fe" in st
        p["phase"] = "active" if active else "awarded"
    for p in completed:
        p["phase"] = "completed"

    # ---- KPI rollups (single source of truth) ----
    pipeline_value = round(sum(p["value"] for p in wip), 2)          # awarded not completed
    completed_value = round(sum(p["value"] for p in completed), 2)   # FY26 completed
    ar_total = round(sum(p["unpaid"] for p in completed), 2)         # completed-but-unpaid = AR
    paid_total = round(sum(p["paid"] for p in (wip + completed)), 2)
    retain_due = round(sum(p["retainAmt"] for p in (wip + completed) if p["retainPaid"].upper().startswith("N")), 2)
    fy26_revenue = completed_value                                   # completed in '26 (contract value of completed)
    total_fy26_sales = round(fy26_revenue + pipeline_value, 2)
    active_projects = [p for p in wip if p["phase"] == "active"]

    # ---- Bids ----
    bwb = openpyxl.load_workbook(BID_XLSX, data_only=True)
    bws = bwb["Agg Pier Bid Log"]
    bids = read_bids(bws)

    def parse(d):
        try:
            return datetime.date.fromisoformat(d)
        except (ValueError, TypeError):
            return None

    # Submitted / Submitted-Edging = bids we put a number on. Many old ones are never
    # moved to "Not Awarded", so headline "outstanding" uses a recency window on the
    # decision (due date) to mean genuinely-in-flight. Full set kept for transparency.
    WINDOW_DAYS = int(os.getenv("PF_BID_WINDOW_DAYS", "90"))
    cutoff = ASOF - datetime.timedelta(days=WINDOW_DAYS)
    submitted = [b for b in bids if b["statusNorm"] in ("submitted", "submitted - edging")]
    for b in submitted:
        dd = parse(b["dueDate"])
        b["live"] = bool(dd and dd >= cutoff)
    outstanding = sorted([b for b in submitted if b["live"]],
                         key=lambda b: b["dueDate"] or "", reverse=True)
    outstanding_value = round(sum(b["value"] for b in outstanding), 2)
    outstanding_all_value = round(sum(b["value"] for b in submitted), 2)

    # Upcoming = any bid with a due date within next 7 days (regardless of status).
    upcoming = []
    for b in bids:
        dd = parse(b["dueDate"])
        if dd and ASOF <= dd <= ASOF + datetime.timedelta(days=7):
            upcoming.append(b)
    upcoming.sort(key=lambda b: b["dueDate"])

    # Win rate from the full bid log: awarded / (awarded + not-awarded). Decided bids only.
    won = sum(1 for b in bids if b["statusNorm"] in ("awarded", "completed"))
    lost = sum(1 for b in bids if "not awarded" in b["statusNorm"])
    win_rate = round(won / (won + lost) * 100) if (won + lost) else 0

    data = {
        "asOf": ASOF.isoformat(),
        "generatedAt": ASOF.isoformat(),
        "kpis": {
            "pipelineValue": pipeline_value,
            "pipelineCount": len(wip),
            "completedValue": completed_value,
            "completedCount": len(completed),
            "fy26Revenue": fy26_revenue,
            "totalFy26Sales": total_fy26_sales,
            "arTotal": ar_total,
            "paidTotal": paid_total,
            "retainDue": retain_due,
            "outstandingBidsValue": outstanding_value,
            "outstandingBidsCount": len(outstanding),
            "outstandingAllValue": outstanding_all_value,
            "outstandingAllCount": len(submitted),
            "bidWindowDays": WINDOW_DAYS,
            "activeProjectsCount": len(active_projects),
            "winRate": win_rate,
            "wonCount": won,
            "lostCount": lost,
            "fy26Goal": 5250000,
        },
        "wip": wip,
        "completed": completed,
        "outstandingBids": outstanding,
        "upcomingBids": upcoming,
    }

    with open(OUT, "w") as f:
        f.write("// AUTO-GENERATED by sync/extract-projects.py — do not edit by hand.\n")
        f.write("// Source: PF Project Master.xlsx + Project Bid Log.xlsx\n")
        f.write("window.PF_PROJECTS = " + js_safe(data) + ";\n")


def js_safe(obj):
    """json.dumps but defanged so spreadsheet text can't break out of a <script>.
    json.dumps does NOT escape '/', so a project/GC name containing </script>
    would terminate an inline script tag (stored XSS). Escaping the slash is
    semantically a no-op inside a JS string but stops the HTML parser."""
    s = json.dumps(obj, indent=2)
    return (s.replace("</", "<\\/")
             .replace(" ", "\\u2028")   # JS line separator
             .replace(" ", "\\u2029")   # JS paragraph separator
             .replace("<!--", "<\\!--"))

    print("Wrote", OUT)
    print("  WIP/awarded projects:", len(wip), "| completed:", len(completed))
    print("  Pipeline value: $%s" % format(pipeline_value,",.0f"))
    print("  Completed value (FY26 rev): $%s" % format(completed_value,",.0f"))
    print("  Total FY26 sales: $%s" % format(total_fy26_sales,",.0f"))
    print("  AR: $%s | Retainage due: $%s" % (format(ar_total,",.0f"), format(retain_due,",.0f")))
    print("  Outstanding bids: %d ($%s)" % (len(outstanding), format(outstanding_value,",.0f")))
    print("  Upcoming bids (<=7d):", len(upcoming))


def read_bids(ws):
    """Agg Pier Bid Log: header row 6, data from row 7. Fixed column map (verified 6/5).
    B=Project#  C=Name  D=City/State  H=Total LF  I=Columns  Q=Due Date  R=Date Submitted
    U=Bid Status  V=Bid Total Value  AA=General Contractor."""
    DATA_START = 7
    C = {"projectNo": 2, "name": 3, "city": 4, "lf": 8, "columns": 9,
         "inviteDate": 16, "dueDate": 17, "dateSubmitted": 18, "followUp": 20,
         "status": 21, "value": 22, "gc": 27}
    out = []
    for r in range(DATA_START, ws.max_row + 1):
        name = s(ws.cell(r, C["name"]).value)
        status = s(ws.cell(r, C["status"]).value)
        if not name and not status:
            continue
        rec = {
            "name": name,
            "projectNo": s(ws.cell(r, C["projectNo"]).value),
            "city": s(ws.cell(r, C["city"]).value),
            "status": status,
            "statusNorm": status.strip().lower().replace("–", "-").replace("  ", " "),
            "value": num(ws.cell(r, C["value"]).value),
            "dueDate": s(ws.cell(r, C["dueDate"]).value),
            "dateSubmitted": s(ws.cell(r, C["dateSubmitted"]).value),
            "gc": s(ws.cell(r, C["gc"]).value),
        }
        out.append(rec)
    return out


if __name__ == "__main__":
    main()
