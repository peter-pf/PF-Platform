#!/usr/bin/env python3
"""
SharePoint Live Data Sync for PF Operations Platform
Downloads bid log and project master from SP, extracts data to JSON
for the platform modules to consume.

Usage: python3 sp-sync.py
Output: platform/data/*.json files
"""

import json
import os
import sys
import urllib.request
import urllib.parse
from datetime import datetime

# ---- CONFIG (from environment or .env) ----
def _load_env():
    """Load .env file if it exists."""
    env_path = os.path.expanduser("~/.env")
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, val = line.split("=", 1)
                    os.environ.setdefault(key.strip(), val.strip())

_load_env()

TENANT_ID = os.environ.get("AZURE_TENANT_ID", "")
CLIENT_ID = os.environ.get("AZURE_CLIENT_ID", "")
CLIENT_SECRET = os.environ.get("AZURE_CLIENT_SECRET", "")
DRIVE_ID = os.environ.get("SP_DRIVE_ID", "")

# SharePoint file IDs (from index)
FILES = {
    "bid_log": "016ISVH6Y7M7KQIB5C5FDLNKI5H3IZFXRK",
    "project_master": "016ISVH64J5UAFQWEW6NC3FJBZVMTVLLX6",
    "estimate_template": "016ISVH63NJBJ4O6UAUBGJ66O246RALYBJ",
    "bd_master": "016ISVH67CTBZWYBZMDREITDFK2XCIWMKM",
    "production_calcs": "016ISVH6Y5EAOIMTDZL5D2IGNUKRZ7LFAC",
    "project_readiness": "016ISVH6Z7SPMXRXECJZALGGG5UEGBVKJU",
}

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")
DOWNLOAD_DIR = os.path.join(os.path.dirname(__file__), "downloads")


def get_token():
    """Get OAuth2 token from Microsoft."""
    data = urllib.parse.urlencode({
        "client_id": CLIENT_ID,
        "scope": "https://graph.microsoft.com/.default",
        "client_secret": CLIENT_SECRET,
        "grant_type": "client_credentials",
    }).encode()
    req = urllib.request.Request(
        f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token",
        data=data,
        headers={"Content-Type": "application/x-www-form-urlencoded"}
    )
    return json.loads(urllib.request.urlopen(req).read())["access_token"]


def download_file(token, item_id, filename):
    """Download a file from SharePoint."""
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{item_id}/content"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    filepath = os.path.join(DOWNLOAD_DIR, filename)
    with urllib.request.urlopen(req) as resp:
        with open(filepath, "wb") as f:
            f.write(resp.read())
    print(f"  Downloaded: {filename} ({os.path.getsize(filepath):,} bytes)")
    return filepath


def extract_bid_log(filepath):
    """Extract bid log data to JSON."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # ---- Agg Pier Bid Log ----
    ws = wb["Agg Pier Bid Log"]
    bids = []
    for row in range(7, ws.max_row + 1):
        name = ws.cell(row=row, column=3).value
        if not name or not str(name).strip():
            continue
        bid = {
            "row": row,
            "number": ws.cell(row=row, column=2).value,
            "name": str(name).strip(),
            "city_state": ws.cell(row=row, column=4).value,
            "address": ws.cell(row=row, column=5).value,
            "total_sf": ws.cell(row=row, column=7).value,
            "total_lf": ws.cell(row=row, column=8).value,
            "total_columns": ws.cell(row=row, column=9).value,
            "total_stone_tn": ws.cell(row=row, column=10).value,
            "duration_days": ws.cell(row=row, column=11).value,
            "feed_type": ws.cell(row=row, column=12).value,
            "tax": ws.cell(row=row, column=13).value,
            "invite_date": str(ws.cell(row=row, column=16).value or ""),
            "due_date": str(ws.cell(row=row, column=17).value or ""),
            "date_submitted": str(ws.cell(row=row, column=18).value or ""),
            "bid_status": ws.cell(row=row, column=21).value,
            "bid_value": ws.cell(row=row, column=22).value,
            "gc_name": ws.cell(row=row, column=27).value,
            "gc_contact": ws.cell(row=row, column=28).value,
            "gc_email": ws.cell(row=row, column=29).value,
            "gc_phone": ws.cell(row=row, column=30).value,
            "notes": ws.cell(row=row, column=31).value,
            "engineer": ws.cell(row=row, column=32).value,
            "prelim_fee": ws.cell(row=row, column=33).value,
        }
        # Clean up numeric fields
        for k in ["total_sf", "total_lf", "total_columns", "duration_days", "bid_value"]:
            if isinstance(bid[k], str):
                bid[k] = None
            if isinstance(bid[k], float):
                bid[k] = round(bid[k], 2)
        bids.append(bid)

    # ---- Stone Costs ----
    stone_costs = []
    if "Stone Costs Per Area" in wb.sheetnames:
        ws2 = wb["Stone Costs Per Area"]
        for row in range(2, ws2.max_row + 1):
            supplier = ws2.cell(row=row, column=1).value
            if not supplier:
                continue
            stone_costs.append({
                "supplier": supplier,
                "city": ws2.cell(row=row, column=2).value,
                "state": ws2.cell(row=row, column=3).value,
                "stone_rate": ws2.cell(row=row, column=4).value,
                "haul_rate": ws2.cell(row=row, column=5).value,
                "tax": ws2.cell(row=row, column=6).value,
                "total": ws2.cell(row=row, column=7).value,
            })

    # ---- FY Goals ----
    fy_goals = {}
    if "FY $ Goal Tracker" in wb.sheetnames:
        ws3 = wb["FY $ Goal Tracker"]
        for row in range(1, ws3.max_row + 1):
            for col in range(1, ws3.max_column + 1):
                val = ws3.cell(row=row, column=col).value
                if val and "Goal" in str(val):
                    fy_goals[str(val)] = ws3.cell(row=row, column=col+1).value

    return {
        "bids": bids,
        "stone_costs": stone_costs,
        "fy_goals": fy_goals,
        "sync_time": datetime.utcnow().isoformat() + "Z",
        "source": "SharePoint/01 - Admin/13 - Master Spreadsheets/Project Bid Log.xlsx",
    }


def extract_project_master(filepath):
    """Extract project master data to JSON."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # ---- 2026 WIP ----
    projects = []
    if "2026 WIP & Completed Projects" in wb.sheetnames:
        ws = wb["2026 WIP & Completed Projects"]
        for row in range(5, ws.max_row + 1):
            proj_num = ws.cell(row=row, column=2).value
            name = ws.cell(row=row, column=4).value
            if not name or not str(name).strip():
                continue
            projects.append({
                "project_number": proj_num,
                "contract_status": ws.cell(row=row, column=3).value,
                "name": str(name).strip(),
                "scope": ws.cell(row=row, column=5).value,
                "address": ws.cell(row=row, column=6).value,
                "city_state": ws.cell(row=row, column=7).value,
                "subcontract_value": ws.cell(row=row, column=8).value,
                "paid": ws.cell(row=row, column=9).value,
                "unpaid": ws.cell(row=row, column=10).value,
                "retain_pct": ws.cell(row=row, column=13).value,
                "retainage": ws.cell(row=row, column=14).value,
                "work_pct_complete": ws.cell(row=row, column=17).value,
                "gc_name": ws.cell(row=row, column=18).value,
                "gc_pm_name": ws.cell(row=row, column=20).value,
                "gc_pm_email": ws.cell(row=row, column=22).value,
                "projected_start": str(ws.cell(row=row, column=26).value or ""),
                "duration_days": ws.cell(row=row, column=27).value,
            })

    # ---- Cost Codes ----
    cost_codes = []
    if "PF Cost Codes" in wb.sheetnames:
        ws2 = wb["PF Cost Codes"]
        for row in range(1, ws2.max_row + 1):
            code = ws2.cell(row=row, column=1).value
            name = ws2.cell(row=row, column=2).value or ws2.cell(row=row, column=3).value
            if code and name:
                cost_codes.append({"code": code, "name": str(name).strip()})

    # ---- KPIs (PF Dashboard) ----
    kpis = {}
    if "PF Dashboard" in wb.sheetnames:
        ws3 = wb["PF Dashboard"]
        labels = []
        for row in range(7, 23):
            label = ws3.cell(row=row, column=2).value
            if label:
                labels.append({"row": row, "label": str(label).strip()})
        # Read January data as sample (columns C-G)
        for item in labels:
            actual = ws3.cell(row=item["row"], column=3).value
            goal = ws3.cell(row=item["row"], column=4).value
            kpis[item["label"]] = {"actual": actual, "goal": goal}

    return {
        "projects": projects,
        "cost_codes": cost_codes,
        "kpis": kpis,
        "sync_time": datetime.utcnow().isoformat() + "Z",
        "source": "SharePoint/01 - Admin/13 - Master Spreadsheets/PF Project Master.xlsx",
    }


def extract_estimate_template(filepath):
    """Extract estimate template structure and default rates to JSON."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)

    rates = {}
    line_items = []

    # Handle trailing spaces in sheet names
    budget_sheet = next((s for s in wb.sheetnames if s.strip() == "Detailed Budget"), None)
    if budget_sheet:
        ws = wb[budget_sheet]
        for row in range(1, ws.max_row + 1):
            code = ws.cell(row=row, column=1).value
            desc = ws.cell(row=row, column=7).value or ws.cell(row=row, column=3).value
            qty = ws.cell(row=row, column=10).value
            unit = ws.cell(row=row, column=11).value
            unit_cost = ws.cell(row=row, column=12).value
            amount = ws.cell(row=row, column=16).value
            if code and desc:
                item = {
                    "code": code,
                    "description": str(desc).strip(),
                    "quantity": qty,
                    "unit": unit,
                    "unit_cost": round(unit_cost, 2) if isinstance(unit_cost, float) else unit_cost,
                    "amount": round(amount, 2) if isinstance(amount, float) else amount,
                }
                line_items.append(item)

    # Extract OH calcs if present
    oh_items = []
    if "OH Calcs" in wb.sheetnames:
        ws2 = wb["OH Calcs"]
        for row in range(1, ws2.max_row + 1):
            code = ws2.cell(row=row, column=1).value
            desc = ws2.cell(row=row, column=2).value
            qty = ws2.cell(row=row, column=3).value
            unit = ws2.cell(row=row, column=4).value
            cost = ws2.cell(row=row, column=5).value
            pct = ws2.cell(row=row, column=6).value
            total = ws2.cell(row=row, column=7).value
            if desc and str(desc).strip():
                oh_items.append({
                    "code": code,
                    "description": str(desc).strip(),
                    "quantity": qty,
                    "unit": unit,
                    "cost_per_unit": cost,
                    "pct_dedicated": pct,
                    "total": total,
                })

    return {
        "line_items": line_items,
        "oh_items": oh_items,
        "sheet_names": wb.sheetnames,
        "sync_time": datetime.utcnow().isoformat() + "Z",
        "source": "SharePoint/03 - Estimating/01 - Aggregate Piers/26-0422 Master Budget Estimate Template PF.xlsm",
    }


def extract_bd_master(filepath):
    """Extract BD master data to JSON."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)

    contacts = []
    gc_relationships = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Read header row to understand columns
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                headers.append({"col": col, "name": str(val).strip()})

        if not headers:
            continue

        # Read data rows
        for row in range(2, min(ws.max_row + 1, 500)):
            row_data = {}
            has_data = False
            for h in headers:
                val = ws.cell(row=row, column=h["col"]).value
                if val is not None:
                    has_data = True
                row_data[h["name"]] = val
            if has_data and any(v for v in row_data.values() if v):
                row_data["_sheet"] = sheet_name
                contacts.append(row_data)

    return {
        "contacts": contacts,
        "sheet_names": wb.sheetnames,
        "total_records": len(contacts),
        "sync_time": datetime.utcnow().isoformat() + "Z",
        "source": "SharePoint/01 - Admin/13 - Master Spreadsheets/PF BD Master.xlsm",
    }


def main():
    print(f"PF SharePoint Sync — {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC")
    print("=" * 50)

    # Get token
    print("Authenticating...")
    token = get_token()
    print("  Token acquired")

    # Download all 4 master files
    print("\nDownloading from SharePoint...")
    bid_log_path = download_file(token, FILES["bid_log"], "Project_Bid_Log.xlsx")
    project_master_path = download_file(token, FILES["project_master"], "PF_Project_Master.xlsx")
    estimate_path = download_file(token, FILES["estimate_template"], "Master_Budget_Estimate_Template.xlsm")
    bd_path = download_file(token, FILES["bd_master"], "PF_BD_Master.xlsm")

    # Extract data from all 4
    print("\nExtracting bid log data...")
    bid_data = extract_bid_log(bid_log_path)
    print(f"  {len(bid_data['bids'])} bids, {len(bid_data['stone_costs'])} stone suppliers")

    print("Extracting project master data...")
    project_data = extract_project_master(project_master_path)
    print(f"  {len(project_data['projects'])} projects, {len(project_data['cost_codes'])} cost codes")

    print("Extracting estimate template...")
    estimate_data = extract_estimate_template(estimate_path)
    print(f"  {len(estimate_data['line_items'])} line items, {len(estimate_data['oh_items'])} OH items")

    print("Extracting BD master...")
    bd_data = extract_bd_master(bd_path)
    print(f"  {bd_data['total_records']} BD records across {len(bd_data['sheet_names'])} sheets")

    # Write JSON output
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    with open(os.path.join(OUTPUT_DIR, "bid-log.json"), "w") as f:
        json.dump(bid_data, f, indent=2, default=str)
    print(f"\n  Wrote: data/bid-log.json")

    with open(os.path.join(OUTPUT_DIR, "project-master.json"), "w") as f:
        json.dump(project_data, f, indent=2, default=str)
    print(f"  Wrote: data/project-master.json")

    with open(os.path.join(OUTPUT_DIR, "estimate-template.json"), "w") as f:
        json.dump(estimate_data, f, indent=2, default=str)
    print(f"  Wrote: data/estimate-template.json")

    with open(os.path.join(OUTPUT_DIR, "bd-master.json"), "w") as f:
        json.dump(bd_data, f, indent=2, default=str)
    print(f"  Wrote: data/bd-master.json")

    # Write sync metadata
    meta = {
        "last_sync": datetime.utcnow().isoformat() + "Z",
        "files_synced": list(FILES.keys()),
        "bid_count": len(bid_data["bids"]),
        "project_count": len(project_data["projects"]),
        "stone_supplier_count": len(bid_data["stone_costs"]),
        "estimate_line_items": len(estimate_data["line_items"]),
        "bd_records": bd_data["total_records"],
    }
    with open(os.path.join(OUTPUT_DIR, "sync-meta.json"), "w") as f:
        json.dump(meta, f, indent=2)
    print(f"  Wrote: data/sync-meta.json")

    print("\nSync complete.")


if __name__ == "__main__":
    main()
