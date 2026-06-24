#!/usr/bin/env python3
"""
BD Dashboard ingest (build-bd-dashboard.py)
===========================================
Derives PERIOD (month / quarter / annual) KPIs for the Business Development
view from the PF BD Master workbook (the same source bd-master uses).

Output: platform/data/bd-dashboard.js  (window.PF_BD_DASH = {...})
shaped EXACTLY like data/pf-dashboard.js so it renders through the SHARED
window.PFDashboard.render(...) engine in index.html.

SOURCE (in priority order):
  1. --local <path.xlsm>            : an explicit PF BD Master workbook
  2. sync/downloads/PF_BD_Master.xlsm (DEFAULT if present)
  3. SharePoint download (needs Graph token)                  [if no local file]

WHAT THE SOURCE ACTUALLY HAS (verified) — and what it does NOT:
  - "Interactions & Actions" sheet HAS a real per-row "Interaction Date" -> this
    is the ONLY dated BD activity. Real PERIOD KPIs come from it:
        * Interactions Logged   (count)  : interactions in the period
        * Companies Contacted   (count)  : DISTINCT companies interacted-with in
                                           the period
  - "Organizations" (companies) and "Contacts" sheets have NO "date added"
    column -> "New Companies Added" / "New Contacts Added" PER PERIOD cannot be
    derived without fabricating dates. They are therefore OMITTED as monthly
    metrics and surfaced only as ANNUAL snapshot totals:
        * Total Companies   (count, annual snapshot)
        * Total Contacts    (count, annual snapshot)
  - "Opportunities" sheet exists (71 rows) but its Status / Stage columns are
    currently 100% BLANK -> Active / Won / Lost opportunities CANNOT be derived.
    OMITTED (no source data). Total Opportunities surfaced as an annual snapshot.

  Snapshot metrics carry data ONLY in the latest populated month (so the monthly
  / quarterly views show real dated activity, and the snapshot count appears once
  on the Annual rollup + the most recent month — never spread/duplicated across
  months as if they were monthly adds).

GATING: BD's own tool data. Area = `business_development` BD area. auth.js
DATA_FILE_AREAS maps /data/bd-dashboard.js -> 'business_dev'. admin + partner +
business_dev see it; field_ops BLOCKED.

Thread caps set BEFORE importing openpyxl (solved-box-gotchas).

Usage:
  python3 build-bd-dashboard.py
  python3 build-bd-dashboard.py --year 2026
  python3 build-bd-dashboard.py --local /path/PF_BD_Master.xlsm
"""

# --- thread caps BEFORE any heavy import ---
import os
for _v in ("OPENBLAS_NUM_THREADS", "OMP_NUM_THREADS", "MKL_NUM_THREADS",
           "NUMEXPR_NUM_THREADS", "VECLIB_MAXIMUM_THREADS"):
    os.environ.setdefault(_v, "1")

import sys
import json
import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
PLATFORM = os.path.dirname(HERE)
DATA_DIR = os.path.join(PLATFORM, "data")
OUT_JS = os.path.join(DATA_DIR, "bd-dashboard.js")
DEFAULT_XLSM = os.path.join(HERE, "downloads", "PF_BD_Master.xlsm")

SOURCE_FILE = "PF_BD_Master.xlsm"
SOURCE_TAB = "Interactions & Actions (+ Organizations / Contacts / Opportunities)"

MONTH_ORDER = ['January', 'February', 'March', 'April', 'May', 'June',
               'July', 'August', 'September', 'October', 'November', 'December']
QUARTERS = {
    'Q1': ['January', 'February', 'March'],
    'Q2': ['April', 'May', 'June'],
    'Q3': ['July', 'August', 'September'],
    'Q4': ['October', 'November', 'December'],
}

# Real per-period (dated) metrics first, then annual snapshots.
PERIOD_METRICS = [
    ('Interactions Logged', 'count'),
    ('Companies Contacted', 'count'),
]
SNAPSHOT_METRICS = [
    ('Total Companies', 'count'),
    ('Total Contacts', 'count'),
    ('Total Opportunities', 'count'),
]
METRICS = PERIOD_METRICS + SNAPSHOT_METRICS

HEADER_ROW = 4  # row 4 (1-indexed) holds the column headers on these sheets


def _token():
    sys.path.insert(0, "/home/aiciv/tools")
    from pf_email import _token as t
    return t()


def download_workbook():
    import urllib.request
    import urllib.parse
    tok = _token()
    drive = os.environ["SP_DRIVE_ID"]
    q = urllib.parse.quote("PF_BD_Master")
    url = "https://graph.microsoft.com/v1.0/drives/%s/root/search(q='%s')" % (drive, q)
    req = urllib.request.Request(url, headers={"Authorization": "Bearer " + tok})
    res = json.load(urllib.request.urlopen(req, timeout=120))
    fid = None
    for it in res.get("value", []):
        if str(it.get("name", "")).lower().startswith("pf_bd_master"):
            fid = it.get("id"); break
    if not fid:
        raise SystemExit("Could not find PF_BD_Master in SharePoint drive")
    curl = "https://graph.microsoft.com/v1.0/drives/%s/items/%s/content" % (drive, fid)
    creq = urllib.request.Request(curl, headers={"Authorization": "Bearer " + tok})
    return urllib.request.urlopen(creq, timeout=180).read()


def col_index(header_row, name_startswith):
    target = name_startswith.strip().lower()
    for i, c in enumerate(header_row):
        if c is not None and str(c).strip().lower().startswith(target):
            return i
    return None


def sheet_rows(ws):
    return list(ws.iter_rows(min_row=1, values_only=True))


def build(xlsm_bytes, forced_year):
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsm_bytes), data_only=True, read_only=True)

    # ---- Interactions & Actions: the dated activity ----
    inter_by_month = {y: [0] * 12 for y in range(2024, 2031)}
    companies_by_month = {y: [set() for _ in range(12)] for y in range(2024, 2031)}
    ws = wb["Interactions & Actions"]
    rows = sheet_rows(ws)
    hdr = rows[HEADER_ROW - 1]
    di = col_index(hdr, "Interaction Date")
    ci = col_index(hdr, "Company")
    ii = col_index(hdr, "Interaction")
    for r in rows[HEADER_ROW:]:
        d = r[di] if (di is not None and len(r) > di) else None
        comp = r[ci] if (ci is not None and len(r) > ci) else None
        inter = r[ii] if (ii is not None and len(r) > ii) else None
        if not isinstance(d, datetime.datetime):
            continue
        if not (inter or comp):
            continue
        y = d.year; mi = d.month - 1
        if y not in inter_by_month:
            continue
        inter_by_month[y][mi] += 1
        if comp and str(comp).strip():
            companies_by_month[y][mi].add(str(comp).strip())

    # ---- snapshot totals (no per-row dates) ----
    def count_named(sheet, name_col_idx=2):
        ws2 = wb[sheet]
        n = 0
        for r in ws2.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            if len(r) > name_col_idx and r[name_col_idx] and str(r[name_col_idx]).strip():
                n += 1
        return n

    total_companies = count_named("Organizations")
    total_contacts = count_named("Contacts")
    total_opps = count_named("Opportunities")
    wb.close()

    # ---- choose the dashboard year (most recent year with interactions) ----
    if forced_year:
        year = int(forced_year)
    else:
        years_with = [y for y in inter_by_month if sum(inter_by_month[y]) > 0]
        year = max(years_with) if years_with else datetime.datetime.now(datetime.timezone.utc).year

    inter = inter_by_month.get(year, [0] * 12)
    comps = companies_by_month.get(year, [set() for _ in range(12)])

    populated_idx = [mi for mi in range(12) if inter[mi] > 0]
    last_populated = populated_idx[-1] if populated_idx else None

    months = {}
    month_has_data = {}
    for mi, mname in enumerate(MONTH_ORDER):
        rows_out = []
        any_data = inter[mi] > 0
        # period metrics
        rows_out.append({
            'metric': 'Interactions Logged', 'type': 'count',
            'qtyActual': (inter[mi] if any_data else None),
            'qtyGoal': None, 'amtActual': None, 'amtBudget': None, 'delta': None,
        })
        rows_out.append({
            'metric': 'Companies Contacted', 'type': 'count',
            'qtyActual': (len(comps[mi]) if any_data else None),
            'qtyGoal': None, 'amtActual': None, 'amtBudget': None, 'delta': None,
        })
        # snapshot metrics: value only in the latest populated month (so Annual
        # rollup + the most recent month show the total once — never summed across
        # months as if they were monthly adds).
        show_snap = (mi == last_populated)
        for name, val in (('Total Companies', total_companies),
                          ('Total Contacts', total_contacts),
                          ('Total Opportunities', total_opps)):
            rows_out.append({
                'metric': name, 'type': 'count',
                'qtyActual': (val if show_snap else None),
                'qtyGoal': None, 'amtActual': None, 'amtBudget': None, 'delta': None,
            })
        months[mname] = {'rows': rows_out}
        month_has_data[mname] = any_data

    out = {
        'year': year,
        'sourceFile': SOURCE_FILE,
        'sourceTab': SOURCE_TAB,
        'syncedAt': datetime.datetime.now(datetime.timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
        'isTemplateData': False,
        'monthOrder': MONTH_ORDER,
        'quarters': QUARTERS,
        'monthHasData': month_has_data,
        'metricsMeta': [{'metric': n, 'type': t} for (n, t) in METRICS],
        'months': months,
        '_omitted': [
            'New Companies Added (per period): Organizations sheet has no date-added column',
            'New Contacts Added (per period): Contacts sheet has no date-added column',
            'Active / Won / Lost Opportunities: Opportunities Status + Stage columns are blank',
        ],
    }
    return out


def write_js(data):
    note = (
        "// AUTO-GENERATED by sync/build-bd-dashboard.py — DO NOT EDIT BY HAND.\n"
        "// Source: %s  synced %s\n"
        "// BD PERIOD KPIs. ONLY dated activity (Interactions) is per-period:\n"
        "//   Interactions Logged, Companies Contacted (distinct).\n"
        "// Total Companies/Contacts/Opportunities are ANNUAL SNAPSHOTS (no\n"
        "// date-added columns exist in the source). New-X-Added and Won/Lost are\n"
        "// OMITTED — the source has no dates / the status columns are blank.\n"
        "// ACCESS: business_dev — admin/partner/business_dev (field_ops BLOCKED).\n"
        % (data['sourceFile'], data['syncedAt'])
    )
    body = "window.PF_BD_DASH = " + json.dumps(data, indent=2) + ";\n"
    with open(OUT_JS, "w") as f:
        f.write(note + body)
    populated = [m for m in data['monthOrder'] if data['monthHasData'][m]]
    print("wrote", OUT_JS,
          "(year %s, %d metrics, %d months with interactions: %s)"
          % (data['year'], len(data['metricsMeta']), len(populated), ", ".join(populated)))
    print("  snapshot totals -> companies=%s contacts=%s opportunities=%s"
          % (
              _snap(data, 'Total Companies'),
              _snap(data, 'Total Contacts'),
              _snap(data, 'Total Opportunities'),
          ))
    for o in data.get('_omitted', []):
        print("  OMITTED:", o)


def _snap(data, metric):
    for mo in data['monthOrder']:
        for r in data['months'][mo]['rows']:
            if r['metric'] == metric and r['qtyActual'] is not None:
                return r['qtyActual']
    return None


def main():
    forced_year = None
    if "--year" in sys.argv:
        forced_year = sys.argv[sys.argv.index("--year") + 1]

    if "--local" in sys.argv:
        path = sys.argv[sys.argv.index("--local") + 1]
    elif os.path.exists(DEFAULT_XLSM):
        path = DEFAULT_XLSM
    else:
        path = None

    if path:
        with open(path, "rb") as f:
            xlsm = f.read()
        print("source:", path)
    else:
        xlsm = download_workbook()
        print("source: SharePoint download")

    data = build(xlsm, forced_year)
    write_js(data)


if __name__ == "__main__":
    main()
