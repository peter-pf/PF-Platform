#!/usr/bin/env python3
"""
build-pf-coi.py — PF standard COI coverages for the PF Platform.

Pulls PF's own standing insurance coverages from the reliable structured source
(the xlsx) and writes data/pf-coi.js (window.PF_COI = {...}) consumed by the
PF COI panel (Project Management -> Insurance -> PF COI) in index.html.

Source of truth: SharePoint
  `01 - Admin / 05 - Insurance / MJ Insurance / 01 - Insurance Policies / PF Insurance Policies.xlsx`
  sheet "26-27 Policies" (one row per coverage line; carrier, agg/occ limits,
  policy #, effective dates, broker, cost code).

The ACORD COI PDF text-extracts poorly (labels without values), so the xlsx is
authoritative. If a limit isn't on the row, it is left blank — never guessed.

The panel also links to the actual COI PDF
  `02 - PF COI / PF - 2026 Proof of Insurance COI.pdf` via its SharePoint webUrl.

Run with thread caps BEFORE importing openpyxl (Box thread guard):
  OMP_NUM_THREADS=1 OPENBLAS_NUM_THREADS=1 MKL_NUM_THREADS=1 python3 build-pf-coi.py
"""

import os
for _v in ("OMP_NUM_THREADS", "OPENBLAS_NUM_THREADS", "MKL_NUM_THREADS",
           "NUMEXPR_NUM_THREADS", "VECLIB_MAXIMUM_THREADS"):
    os.environ.setdefault(_v, "1")

import sys
import json
import urllib.request
from datetime import datetime, timezone

sys.path.insert(0, "/home/aiciv/tools")
import pf_email  # noqa: E402

import openpyxl  # noqa: E402

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.normpath(os.path.join(HERE, "..", "data"))
DL_DIR = os.path.join(HERE, "downloads")

POLICIES_ITEM = "016ISVH64VDRG3WFCR2JDY6GJVR6PI6MLW"  # PF Insurance Policies.xlsx
POLICIES_WEBURL = (
    "https://pierfoundations.sharepoint.com/sites/pf/_layouts/15/Doc.aspx?"
    "sourcedoc=%7BBB4D1C95-5114-47D2-8F19-358F9E8F3176%7D&file=PF%20Insurance%20Policies.xlsx"
    "&action=default&mobileredirect=true"
)
# 02 - PF COI / PF - 2026 Proof of Insurance COI.pdf
COI_PDF_WEBURL = (
    "https://pierfoundations.sharepoint.com/sites/pf/Shared%20Documents/01%20-%20Admin/"
    "05%20-%20Insurance/MJ%20Insurance/02%20-%20PF%20COI/"
    "PF%20-%202026%20Proof%20of%20Insurance%20COI.pdf"
)

POLICIES_SHEET = "26-27 Policies"

# The structured policy subfolders in "01 - Insurance Policies" (context for the panel).
POLICY_FOLDERS = [
    "7110 - Equipment Insurance",
    "7120 - General Liability",
    "7120 - Hired & Non-Owned Auto (HNOA) Liability",
    "7125 - Commercial $3M Umbrella Policy",
    "7125 - Commercial $5M Excess Umbrella Liability",
    "7130 - Workers Comp Insurance",
    "7140 - Professional Liability (D&O)",
    "7150 - Contractors Pollution Liability",
]


def download(item, name):
    os.makedirs(DL_DIR, exist_ok=True)
    fp = os.path.join(DL_DIR, name)
    url = "https://graph.microsoft.com/v1.0/drives/%s/items/%s/content" % (
        pf_email._env()["SP_DRIVE_ID"], item)
    req = urllib.request.Request(url, headers={"Authorization": "Bearer " + pf_email._token()})
    with urllib.request.urlopen(req, timeout=120) as r, open(fp, "wb") as f:
        f.write(r.read())
    print("  downloaded %s (%d bytes)" % (name, os.path.getsize(fp)))
    return fp


def s(v):
    """Cell -> clean string. Dates -> M/D/YYYY. Ints kept as ints. Blank stays blank."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return "%d/%d/%d" % (v.month, v.day, v.year)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def header_map(ws, hr=2):
    m = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=hr, column=c).value
        if v not in (None, ""):
            m[str(v).strip()] = c
    return m


def extract_coverages(fp):
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    if POLICIES_SHEET not in wb.sheetnames:
        raise RuntimeError("Sheet '%s' not found (have: %s)" % (POLICIES_SHEET, wb.sheetnames))
    ws = wb[POLICIES_SHEET]
    h = header_map(ws, 2)  # header row is row 2 (row 1 is the title)

    c_code = h.get("PF Cost Code")
    c_type = h.get("Policy Type")
    c_broker = h.get("Insurance Broker")
    c_carrier = h.get("Insurance Underwriter / Carrier")
    c_agg = h.get("Aggregate Limit")
    c_occ = h.get("Each Occurence Limit") or h.get("Each Occurrence Limit")
    c_polno = h.get("Policy Number")
    c_eff = h.get("Policy Effective Dates")

    if not (c_type and c_polno):
        raise RuntimeError("Required columns not found in '%s'" % POLICIES_SHEET)

    lines = []
    for r in range(3, ws.max_row + 1):
        ptype = s(ws.cell(row=r, column=c_type).value) if c_type else ""
        polno = s(ws.cell(row=r, column=c_polno).value) if c_polno else ""
        # A real coverage line needs at least a policy type + a policy number.
        if not ptype or not polno:
            continue
        lines.append({
            "cost_code": s(ws.cell(row=r, column=c_code).value) if c_code else "",
            "type": ptype,
            "broker": s(ws.cell(row=r, column=c_broker).value) if c_broker else "",
            "carrier": s(ws.cell(row=r, column=c_carrier).value) if c_carrier else "",
            "aggregate": s(ws.cell(row=r, column=c_agg).value) if c_agg else "",
            "each_occurrence": s(ws.cell(row=r, column=c_occ).value) if c_occ else "",
            "policy_number": polno,
            "effective": s(ws.cell(row=r, column=c_eff).value) if c_eff else "",
        })
    return lines


def main():
    print("PF COI build — %s UTC" % datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"))
    print("Authenticating + downloading PF Insurance Policies.xlsx...")
    fp = download(POLICIES_ITEM, "PF_Insurance_Policies.xlsx")
    coverages = extract_coverages(fp)
    print("  parsed %d coverage lines" % len(coverages))

    payload = {
        "generated": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M") + " UTC",
        "source": ("SharePoint/01 - Admin/05 - Insurance/MJ Insurance/"
                   "01 - Insurance Policies/PF Insurance Policies.xlsx (sheet '%s')" % POLICIES_SHEET),
        "source_url": POLICIES_WEBURL,
        "coi_pdf_url": COI_PDF_WEBURL,
        "broker": "The MJ Companies (MJ Insurance)",
        "policy_folders": POLICY_FOLDERS,
        "coverages": coverages,
    }

    os.makedirs(DATA_DIR, exist_ok=True)
    out_path = os.path.join(DATA_DIR, "pf-coi.js")
    with open(out_path, "w") as f:
        f.write("// AUTO-GENERATED by sync/build-pf-coi.py — do not edit by hand.\n")
        f.write("// PF standard COI coverages from PF Insurance Policies.xlsx ('26-27 Policies').\n")
        f.write("window.PF_COI = ")
        json.dump(payload, f, indent=2)
        f.write(";\n")

    print("\nWrote %s (%d coverages)" % (out_path, len(coverages)))


if __name__ == "__main__":
    main()
