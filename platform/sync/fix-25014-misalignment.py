#!/usr/bin/env python3
"""fix-25014-misalignment.py — surgical fix for the ONE misaligned project (25-014).

ROOT CAUSE (verified against the source, not guessed):
  A cross-workbook PROJECT-NUMBER COLLISION compounded by a corrupt source row.
  * Metadata `Agg Pier Metadata` row 10 for number "25-014" is
    **Cleveland West Veterans Project** (Cleveland OH, GC = Ozanne Construction,
    Stage = Budget Pricing) — a genuine, distinct project.
  * The `2025 WIP & Completed Projects` sheet's ONLY "25-014" row (row 22) is a
    DIFFERENT project — "Fostoria Schools - Pk-6 Addt'n" (GC = Touchstone CPM,
    PM = Amanda Burgess) — AND that row is itself hand-entered one column LEFT of
    its headers (Subcontract Value under City/State, GC PM Name under GC Address,
    etc.), which is why the consolidated values looked shifted (Work % Complete =
    "Touchstone CPM", GC PM Name = a phone, GC PM Phone = an email).
  * Cleveland West / Ozanne has NO WIP record at all. So the consolidation matched
    purely on the number "25-014" and pulled the unrelated, corrupt Fostoria row
    onto the Cleveland West row.

CORRECT FIX (no fabrication):
  Do NOT attach the Touchstone/Amanda contacts to Cleveland West — that would be a
  FALSE GC association (they belong to a different project). Instead CLEAR the 17
  WIP-sourced cells on the Cleveland West row so it shows no financials/contacts,
  which is accurate for a Budget-Pricing bid to Ozanne with no WIP record. The
  trustworthy metadata-native identity fields (Project Name, GC = Ozanne, Stage,
  Bid Total, etc.) are left UNTOUCHED. Only row 10, only these 17 new columns.

  A guard is also added to the consolidation builder (add-projectmgmt-columns.py)
  so a future re-run never re-pulls this collision.

Locked workbook (HTTP 423) -> log + STOP (report pending). TEST workbook only.

Run: OMP_NUM_THREADS=1 OPENBLAS_NUM_THREADS=1 MKL_NUM_THREADS=1 python3 fix-25014-misalignment.py [--dry-run]
"""
import os
for _v in ("OMP_NUM_THREADS", "OPENBLAS_NUM_THREADS", "MKL_NUM_THREADS",
           "NUMEXPR_NUM_THREADS", "VECLIB_MAXIMUM_THREADS"):
    os.environ.setdefault(_v, "1")

import io
import sys
import urllib.request
import urllib.error

sys.path.insert(0, "/home/aiciv")
from tools import inbox_check_daemon as _d  # noqa: E402
import openpyxl  # noqa: E402
from openpyxl.utils import get_column_letter as gl  # noqa: E402

DRIVE = "b!ogeNU-bvwUevFyKNf9PvlnJJzsEhnrxMv1zdx5x3u8NS2DUHVpM_Q7YocCSzzqgA"  # TEST
META_ITEM = "016ISVH66XMEC5VHR24BD2SER5BIH2PMJL"
TAB = "Agg Pier Metadata"
DRY = "--dry-run" in sys.argv

# The 17 folded-in columns to CLEAR on the 25-014 row (all WIP-sourced from the
# unrelated/corrupt Fostoria row). Identity + metadata-native fields are NOT here.
CLEAR_HEADERS = [
    "Contract Status", "Subcontract Value", "Work % Complete", "Scheduled Completion",
    "Paid", "Unpaid", "Projected PA #1 Income", "Invoice Due By Date", "Retain %",
    "Retainage Amount", "Retainage Submitted", "Retain Paid",
    "GC Address", "GC PM Name", "GC PM Phone", "GC PM Email",
    "GC Super Name", "GC Super Phone", "GC Super Email",
]
TARGET_NUM = "25-014"
EXPECT_NAME = "Cleveland West Veterans Project"   # sanity guard


def graph_get(item):
    tok = _d.token(_d.load_env())
    u = f"https://graph.microsoft.com/v1.0/drives/{DRIVE}/items/{item}/content"
    return urllib.request.urlopen(
        urllib.request.Request(u, headers={"Authorization": "Bearer " + tok}), timeout=180).read()


def graph_put(item, data):
    tok = _d.token(_d.load_env())
    u = f"https://graph.microsoft.com/v1.0/drives/{DRIVE}/items/{item}/content"
    req = urllib.request.Request(u, data=data, method="PUT",
                                 headers={"Authorization": "Bearer " + tok,
                                          "Content-Type": "application/octet-stream"})
    return urllib.request.urlopen(req, timeout=180)


def main():
    print("FIX 25-014: clear WIP-sourced cells wrongly pulled from an unrelated corrupt row")
    wb = openpyxl.load_workbook(io.BytesIO(graph_get(META_ITEM)))  # keep formulas/styles
    ws = wb[TAB]
    h = {str(ws.cell(2, c).value).strip(): c
         for c in range(1, ws.max_column + 1) if ws.cell(2, c).value}
    pnc = h["Project Number"]

    row = None
    for r in range(3, ws.max_row + 1):
        if str(ws.cell(r, pnc).value or "").strip() == TARGET_NUM:
            row = r
            break
    if not row:
        raise SystemExit("  FATAL: 25-014 row not found")

    name = str(ws.cell(row, h["Project Name"]).value or "").strip()
    print("  target row %d: Project Name=%r  GC=%r" %
          (row, name, ws.cell(row, h["General Contractor"]).value))
    if name != EXPECT_NAME:
        raise SystemExit("  ABORT: row %d name %r != expected %r — not touching."
                         % (row, name, EXPECT_NAME))

    cleared = []
    for hdr in CLEAR_HEADERS:
        c = h.get(hdr)
        if not c:
            continue
        old = ws.cell(row, c).value
        if old not in (None, ""):
            if not DRY:
                ws.cell(row, c).value = None
            cleared.append((gl(c), hdr, old))

    print("\n  CLEARING %d cells on row %d (all sourced from the unrelated corrupt "
          "'Fostoria Schools' WIP row):" % (len(cleared), row))
    for (col, hdr, old) in cleared:
        print("     %-4s %-26s was %r" % (col + str(row), hdr, old))

    if DRY:
        print("\n  --dry-run: NOT writing.")
        return

    buf = io.BytesIO()
    wb.save(buf)
    print("\n  Writing back (TEST) ...")
    try:
        graph_put(META_ITEM, buf.getvalue())
        print("  OK: cleared %d cells on 25-014. Other 21 projects untouched." % len(cleared))
    except urllib.error.HTTPError as ex:
        if ex.code == 423:
            print("  LOCKED (HTTP 423): workbook open. NOTHING written. PENDING re-run.")
            sys.exit(42)
        raise


if __name__ == "__main__":
    main()
