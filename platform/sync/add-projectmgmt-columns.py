#!/usr/bin/env python3
"""add-projectmgmt-columns.py — STEP 1 of the "one source per project" project.

Extends the FLAT "Agg Pier Metadata" tab of *PETER TEST - Bidding Metadata.xlsx*
(TEST workbook only) with the PROJECT MANAGEMENT / financial / contract /
engineering fields, folding in the data that used to live in the separate
*PETER TEST - PF Project Master.xlsx* so the portal has ONE source per project.

Design decisions (grounded, no fabrication):
  * NEW columns are APPENDED after the existing last column ("Notes"). Appending
    (not inserting) means NO existing column index shifts, so the price FORMULAS
    (=Z3/I3 ...) and every data-validation range (E3:E43 ...) stay valid. New
    category bands are added to the RIGHT of the existing 10 bands.
  * Existing 45 columns, their data, dropdowns, formats, formulas: UNTOUCHED.
  * Columns already present in the sheet are SKIPPED (no duplicates).
  * POPULATE by Project Number match:
      - General Info + LOI/NOI Date + Prelim Completed By come from the
        Project Master "Project Dashboard" tab (TRANSPOSED: read DOWN a column).
      - Financials + Contract Status + GC contacts + Scheduled Completion come
        from the "2026/2025 WIP & Completed Projects" tabs (one row per project).
  * ENGINEERING/SUBMITTALS DATE columns: the Dashboard's Contract-Info and
    Engineering rows (rows 27-56) DO NOT hold dates for most fields — they hold
    RESPONSIBILITY LABELS ("Garbin", "PF", "FA Wilhelm", "FedEx", "MLS", ...).
    Writing those as dates would FABRICATE data, so those date columns are added
    as STRUCTURE but left BLANK (populated later when a real date source exists).
    Only "LOI / NOI Date" and "Prelim Completed By" have trustworthy Dashboard
    values and are populated.
  * Locked workbook (HTTP 423 on PUT) -> log clearly and STOP (Derek has it open).

Run with thread caps (Box guard):
  OMP_NUM_THREADS=1 OPENBLAS_NUM_THREADS=1 MKL_NUM_THREADS=1 python3 add-projectmgmt-columns.py [--dry-run]
"""
import os
for _v in ("OMP_NUM_THREADS", "OPENBLAS_NUM_THREADS", "MKL_NUM_THREADS",
           "NUMEXPR_NUM_THREADS", "VECLIB_MAXIMUM_THREADS"):
    os.environ.setdefault(_v, "1")

import io
import sys
import urllib.request
import urllib.error
from datetime import datetime, date

sys.path.insert(0, "/home/aiciv")
from tools import inbox_check_daemon as _d  # noqa: E402
import openpyxl  # noqa: E402
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402
from openpyxl.utils import get_column_letter as gl  # noqa: E402

DRIVE = "b!ogeNU-bvwUevFyKNf9PvlnJJzsEhnrxMv1zdx5x3u8NS2DUHVpM_Q7YocCSzzqgA"  # TEST drive
META_ITEM = "016ISVH66XMEC5VHR24BD2SER5BIH2PMJL"    # PETER TEST - Bidding Metadata.xlsx
PM_ITEM = "016ISVH63AII5Q2GCARZFLRNWRE4TC4XCS"      # PETER TEST - PF Project Master.xlsx
TAB = "Agg Pier Metadata"

DRY = "--dry-run" in sys.argv

# ---------------------------------------------------------------------------
# Formats (mirror the sheet's existing conventions).
FMT_MONEY = "$#,##0.00"
FMT_DATE = "m/d/yyyy"
FMT_PHONE = "(###) ###-####"
FMT_QTY1 = "#,##0.0"
FMT_PCT = "0%"           # Work % Complete / Retain % stored as fractions in source
FMT_TEXT = "General"

# Category-band colors (new bands, echoing the sheet's saturated palette).
BAND_STYLE = {
    "CONTRACT":        "00344E86",
    "FINANCIALS":      "00196F3D",
    "PROJECT TEAM":    "007D3C98",
    "SITE":            "008A6D3B",
    "GC CONTACTS":     "005D6D7E",
    "ENG / SUBMITTALS": "004A5568",
}

# ---------------------------------------------------------------------------
# NEW COLUMNS. Each: (band, header, fmt, source_spec)
# source_spec:
#   ("dash", <dashboard attribute label>)  -> read DOWN the project's column
#   ("wip",  <wip header>)                  -> read the project's WIP row
#   ("none", None)                          -> structural only (no trustworthy source)
NEW_COLUMNS = [
    # CONTRACT
    ("CONTRACT", "Contract Status",              FMT_TEXT,  ("wip", "Contract Status")),
    ("CONTRACT", "Subcontract Value",            FMT_MONEY, ("wip", "Subcontract Value")),
    ("CONTRACT", "LOI / NOI Date",               FMT_DATE,  ("dash", "LOI / NOI Date")),
    ("CONTRACT", "Fully Executed Contract Date", FMT_DATE,  ("none", None)),   # Dashboard row = responsibility label, not a date
    ("CONTRACT", "Prelim Completed By",          FMT_TEXT,  ("dash", "Prelim Completed By")),
    ("CONTRACT", "Work % Complete",              FMT_PCT,   ("wip", "Work % Complete")),
    ("CONTRACT", "Scheduled Completion",         FMT_DATE,  ("wip", "Scheduled Completion")),
    # FINANCIALS
    ("FINANCIALS", "Paid",                    FMT_MONEY, ("wip", "Paid")),
    ("FINANCIALS", "Unpaid",                  FMT_MONEY, ("wip", "Unpaid")),
    ("FINANCIALS", "Projected PA #1 Income",  FMT_MONEY, ("wip", "Projected    \nPA #1 Income")),
    ("FINANCIALS", "Invoice Due By Date",     FMT_TEXT,  ("wip", "Invoice Due By Date")),  # source is free-text ("20th"), keep text
    ("FINANCIALS", "Retain %",                FMT_PCT,   ("wip", "Retain %")),
    ("FINANCIALS", "Retainage Amount",        FMT_MONEY, ("wip", "Retainage Amount")),
    ("FINANCIALS", "Retainage Submitted",     FMT_TEXT,  ("wip", "Retainage Submitted")),
    ("FINANCIALS", "Retain Paid",             FMT_TEXT,  ("wip", "Retain Paid            (Y or N)")),  # Yes/No dropdown added below
    # PROJECT TEAM
    ("PROJECT TEAM", "Project Manager",     FMT_TEXT, ("dash", "Project Manager")),
    ("PROJECT TEAM", "Field Operation Mgr", FMT_TEXT, ("dash", "Field Operation Mgr")),
    ("PROJECT TEAM", "Operator 2",          FMT_TEXT, ("dash", "Operator 2")),
    ("PROJECT TEAM", "Operator 3",          FMT_TEXT, ("dash", "Operator 3")),
    ("PROJECT TEAM", "Equipment",           FMT_TEXT, ("dash", "Equipment")),
    ("PROJECT TEAM", "Precon / Estimating", FMT_TEXT, ("dash", "Precon / Estimating")),
    # SITE
    ("SITE", "County",                FMT_TEXT,  ("dash", "County")),
    ("SITE", "Township",              FMT_TEXT,  ("dash", "Township")),
    ("SITE", "Project Scope",         FMT_TEXT,  ("dash", "Project Scope")),
    ("SITE", "Column Diameter",       FMT_TEXT,  ("dash", "Column Diameter")),
    ("SITE", "Estimated Spoils (CY)", FMT_QTY1,  ("dash", "Estimated Spoils (CY)")),
    # GC CONTACTS
    ("GC CONTACTS", "GC Address",       FMT_TEXT,  ("wip", "GC Address")),
    ("GC CONTACTS", "GC PM Name",       FMT_TEXT,  ("wip", "GC PM Name")),
    ("GC CONTACTS", "GC PM Phone",      FMT_PHONE, ("wip", "GC PM Phone")),
    ("GC CONTACTS", "GC PM Email",      FMT_TEXT,  ("wip", "GC PM Email")),
    ("GC CONTACTS", "GC Super Name",    FMT_TEXT,  ("wip", "GC Super Name")),
    ("GC CONTACTS", "GC Super Phone",   FMT_PHONE, ("wip", "GC Super Phone")),
    ("GC CONTACTS", "GC Super Email",   FMT_TEXT,  ("wip", "GC Super Email")),
    # ENGINEERING / SUBMITTALS — Dashboard rows hold responsibility labels, NOT
    # dates, so these are added as STRUCTURE (blank) to avoid fabricating dates.
    ("ENG / SUBMITTALS", "Release Date to Start Submittals",   FMT_DATE, ("none", None)),
    ("ENG / SUBMITTALS", "Items Needed for Submittal Completion", FMT_TEXT, ("none", None)),
    ("ENG / SUBMITTALS", "Submittals Received from Engineer",  FMT_DATE, ("none", None)),
    ("ENG / SUBMITTALS", "Submittals Sent to GC",              FMT_DATE, ("none", None)),
    ("ENG / SUBMITTALS", "Submittals Approved & Returned",     FMT_DATE, ("none", None)),
    ("ENG / SUBMITTALS", "Submittals & Shop Dwgs Saved",       FMT_DATE, ("none", None)),
    ("ENG / SUBMITTALS", "Submittals & Shop Dwgs Sent to GC",  FMT_DATE, ("none", None)),
    ("ENG / SUBMITTALS", "Shop Drawings Ready for Pickup",     FMT_DATE, ("none", None)),
    ("ENG / SUBMITTALS", "Docs Sent to Surveyor / Layout",     FMT_DATE, ("none", None)),
]

# Columns already present in the metadata sheet -> we do NOT re-add (report as skipped).
# (Comparison done live against actual row-2 headers below; this list documents the
#  known duplicates from the spec for the report.)
KNOWN_DUPES = [
    "Project Name", "Address", "Total LF", "Total Columns", "Estimated Stone (TN)",
    "Top vs. Bottom Feed", "General Contractor", "Projected Start Date",
    "Invite Date", "Due Date", "Project Duration (Days)", "City / State",
    "Award Date", "GC Email", "GC Phone", "Bidding GCs", "Awarded GC",
]

# WIP-match EXCLUSIONS: project numbers where the metadata project and the WIP
# "same-number" row are DIFFERENT projects (cross-workbook number collision), so
# pulling WIP financials/contacts would attach the WRONG project's data.
#   25-014: metadata = "Cleveland West Veterans Project" (Ozanne, Budget Pricing,
#           no WIP record); the 2025-WIP "25-014" is a mis-keyed, column-shifted
#           "Fostoria Schools" row (Touchstone CPM). Verified 2026-07-20.
WIP_MATCH_EXCLUDE = {"25-014"}


def graph_get(item):
    tok = _d.token(_d.load_env())
    u = f"https://graph.microsoft.com/v1.0/drives/{DRIVE}/items/{item}/content"
    req = urllib.request.Request(u, headers={"Authorization": "Bearer " + tok})
    return urllib.request.urlopen(req, timeout=180).read()


def graph_put(item, data):
    tok = _d.token(_d.load_env())
    u = f"https://graph.microsoft.com/v1.0/drives/{DRIVE}/items/{item}/content"
    req = urllib.request.Request(u, data=data, method="PUT",
                                 headers={"Authorization": "Bearer " + tok,
                                          "Content-Type": "application/octet-stream"})
    return urllib.request.urlopen(req, timeout=180)


def norm_num(v):
    """Normalize a project number for matching (strip, lower). Blank -> ''."""
    return str(v or "").strip().lower()


def load_project_master(raw):
    """Return two dicts keyed by normalized project number:
       dash[num]  = {attribute_label: value}     (from transposed Project Dashboard)
       wip[num]   = {wip_header: value}           (merged 2026 + 2025 WIP rows)
    """
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)

    # --- Project Dashboard (TRANSPOSED) ---
    ws = wb["Project Dashboard"]
    # attribute label per row (col B holds the label; col A holds section headers)
    labels = {}
    for r in range(3, ws.max_row + 1):
        b = ws.cell(r, 2).value
        if b not in (None, ""):
            labels[r] = str(b).strip()
    # project number per column (row 2)
    dash = {}
    for c in range(3, ws.max_column + 1):
        pn = ws.cell(2, c).value
        if pn in (None, ""):
            continue
        key = norm_num(pn)
        if not key or key in ("26-xxx",):
            continue
        attrs = {}
        for r, lbl in labels.items():
            v = ws.cell(r, c).value
            if v not in (None, ""):
                attrs[lbl] = v
        dash[key] = attrs

    # --- WIP sheets (one row per project); 2026 first, 2025 fills gaps ---
    wip = {}
    for sheet in ("2025 WIP & Completed Projects", "2026 WIP & Completed Projects"):
        if sheet not in wb.sheetnames:
            continue
        w = wb[sheet]
        hdr = {str(w.cell(3, c).value).strip(): c
               for c in range(1, w.max_column + 1) if w.cell(3, c).value}
        pc = hdr.get("Project #")
        if not pc:
            continue
        for r in range(4, w.max_row + 1):
            pn = w.cell(r, pc).value
            key = norm_num(pn)
            if not key or key == "project #":
                continue
            row = {}
            for hn, c in hdr.items():
                v = w.cell(r, c).value
                if v not in (None, ""):
                    row[hn] = v
            # 2026 overrides 2025 (later loop wins) — most current data
            wip[key] = row
    return dash, wip


def coerce(fmt, val):
    """Coerce a source value to fit the target column format. Never fabricate:
    if a DATE column gets a non-date (responsibility text), return the text as-is
    (rare here since 'none' sources are blank; kept for robustness)."""
    if val is None:
        return None
    if fmt == FMT_DATE:
        if isinstance(val, (datetime, date)):
            return val
        # non-date text in a date slot -> keep text (won't be date-formatted cleanly)
        return val
    if fmt in (FMT_MONEY, FMT_QTY1, FMT_PCT):
        try:
            return float(val)
        except (TypeError, ValueError):
            return val  # non-numeric (e.g. "TBD") passes through as text
    if fmt == FMT_PHONE:
        # phone format only renders for a pure numeric; strings ("317.691.5568",
        # "TBD", multi-number) pass through as text so nothing is lost/garbled.
        if isinstance(val, (int, float)):
            return int(val)
        return str(val).strip()
    return str(val).strip() if not isinstance(val, (int, float)) else val


def main():
    print("STEP 1: extend Agg Pier Metadata with PM/financial/contract/eng columns")
    print("Downloading TEST workbooks (fresh) ...")
    meta_raw = graph_get(META_ITEM)
    pm_raw = graph_get(PM_ITEM)
    print("  metadata %d bytes | project master %d bytes" % (len(meta_raw), len(pm_raw)))

    dash, wip = load_project_master(pm_raw)
    print("  project master: %d dashboard projects, %d WIP projects" % (len(dash), len(wip)))

    wb = openpyxl.load_workbook(io.BytesIO(meta_raw))  # keep formulas + styles
    ws = wb[TAB]

    existing = {str(ws.cell(2, c).value).strip(): c
                for c in range(1, ws.max_column + 1) if ws.cell(2, c).value}
    print("  existing columns: %d" % len(existing))

    # metadata Project Number column (col 1) + name for reporting
    pn_col = existing.get("Project Number")
    name_col = existing.get("Project Name")

    # style templates from existing cells
    band_font = Font(bold=True, color="FFFFFFFF", size=10)
    band_align = Alignment(horizontal="center", vertical="center")
    hdr_fill = PatternFill(patternType="solid", fgColor="00EEF2F6")
    hdr_font = Font(bold=True, color="001c1f23", size=9)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    hdr_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    last_row = ws.max_row  # data rows 3..last_row (43)
    start_col = ws.max_column + 1
    col = start_col

    added = []
    skipped = []
    retain_paid_col = None
    prev_band = None
    band_start = None
    band_ranges = []  # (band_label, first_col, last_col)

    def close_band():
        if prev_band is not None and band_start is not None:
            band_ranges.append((prev_band, band_start, col - 1))

    for (band, header, fmt, source) in NEW_COLUMNS:
        if header in existing:
            skipped.append(header)
            continue

        if band != prev_band:
            close_band()
            prev_band = band
            band_start = col

        # header cell (row 2)
        hc = ws.cell(2, col)
        hc.value = header
        hc.fill = hdr_fill
        hc.font = hdr_font
        hc.alignment = hdr_align
        hc.border = hdr_border

        # width heuristic
        w = max(12, min(30, len(header) + 2))
        if fmt == FMT_MONEY:
            w = max(w, 14)
        if header in ("GC Address", "Items Needed for Submittal Completion", "Project Scope"):
            w = 30
        ws.column_dimensions[gl(col)].width = w

        # populate + format each data row
        populated_here = 0
        for r in range(3, last_row + 1):
            pnum = norm_num(ws.cell(r, pn_col).value) if pn_col else ""
            cell = ws.cell(r, col)
            cell.number_format = fmt
            if fmt == FMT_TEXT:
                cell.number_format = "General"
            if not pnum:
                continue  # bidding-stage rows without a number -> leave blank (expected)
            kind, key = source
            val = None
            if kind == "dash":
                val = dash.get(pnum, {}).get(key)
            elif kind == "wip":
                # Skip WIP for known cross-workbook number collisions (wrong project).
                if pnum not in WIP_MATCH_EXCLUDE:
                    val = wip.get(pnum, {}).get(key)
            # kind == "none" -> structural, stays blank
            if val is not None:
                cell.value = coerce(fmt, val)
                populated_here += 1

        if header == "Retain Paid":
            retain_paid_col = col

        added.append((band, header, fmt, source[0], populated_here))
        col += 1

    close_band()
    end_col = col - 1

    # category bands (row 1): merge + style each new band's range
    for (label, c0, c1) in band_ranges:
        fill = PatternFill(patternType="solid", fgColor=BAND_STYLE[label])
        if c1 > c0:
            ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c1)
        top = ws.cell(1, c0)
        top.value = label
        top.fill = fill
        top.font = band_font
        top.alignment = band_align
        # ensure merged-under cells carry the fill too
        for cc in range(c0, c1 + 1):
            ws.cell(1, cc).fill = fill

    # Retain Paid -> Yes/No dropdown across data rows
    if retain_paid_col:
        letter = gl(retain_paid_col)
        dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        dv.sqref = f"{letter}3:{letter}{last_row}"
        ws.add_data_validation(dv)

    print("\n  ADDED %d columns (cols %s..%s):" % (len(added), gl(start_col), gl(end_col)))
    for (band, header, fmt, kind, n) in added:
        print("     [%-15s] %-34s fmt=%-11s src=%-4s populated=%d" %
              (band, header, fmt, kind, n))
    print("\n  SKIPPED (already present) %d:" % len(skipped))
    for h in skipped:
        print("     - %s" % h)

    # projects that got any PM data
    matched = set()
    for r in range(3, last_row + 1):
        pnum = norm_num(ws.cell(r, pn_col).value)
        if pnum and (pnum in dash or pnum in wip):
            matched.add(pnum)
    print("\n  Projects matched to PM data by number: %d" % len(matched))
    print("     %s" % ", ".join(sorted(matched)))

    if DRY:
        print("\n  --dry-run: NOT writing. (would append %d cols)" % len(added))
        return

    # write back (skip cleanly on 423 lock)
    buf = io.BytesIO()
    wb.save(buf)
    print("\n  Writing workbook back to SharePoint (TEST) ...")
    try:
        graph_put(META_ITEM, buf.getvalue())
        print("  OK: wrote %d new columns to '%s'." % (len(added), TAB))
    except urllib.error.HTTPError as ex:
        if ex.code == 423:
            print("  LOCKED (HTTP 423): workbook is open (Derek). NOTHING written. "
                  "STEP 1 is PENDING until the workbook is closed. Re-run then.")
            sys.exit(42)
        raise


if __name__ == "__main__":
    main()
