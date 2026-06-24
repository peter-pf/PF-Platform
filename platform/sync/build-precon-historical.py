#!/usr/bin/env python3
"""
Historical Bid Log ingest (build-precon-historical.py)
======================================================
Reads the per-year Win/Loss tabs of the Project Bid Log workbook FULLY and emits
a read-only historical feed for the Preconstruction > Historical Bid Log view.

INSPECTED STRUCTURE of "2025 Win Loss Log" (verified, NOT assumed):
  - Rows 1-19: a SUMMARY/STATS block (win/loss category counts: Budget, Awarded,
    Will Not Bid, Not Awarded - *, plus Total Projects + Win Rate ratios). Col B =
    label, C = Total Count, D = Prelim Used, E = Prelim %. Captured as `summary`.
  - Row 22: a single block title ("Budget Pricing") + grouped section headers.
  - Row 23: the COLUMN HEADER row ("Number", "Project Number", "Project Name",
    ... "Date Paid") = 35 columns. "Number" (col A) is a 1..N rank index, NOT a
    data field, so it is dropped from the emitted fields.
  - Rows 24+: the per-bid rows (168). A row is a bid if it has a Project Name.
    Repeated header rows (Bid Status == "Bid Status") are SKIPPED. Some cells hold
    spreadsheet errors (#REF!, #DIV/0!); these are kept VERBATIM, never computed.
  - The win/loss OUTCOME is derived from the Bid Status column:
      Completed / Awarded            -> 'win'
      Not Awarded - * / Will Not Bid -> 'loss'
      anything else                  -> 'other'

Output: platform/data/precon-historical.js
  window.PF_PRECON_HISTORICAL = {
    years: [ {
      year, tab,
      columns: [ ...header names, excluding the rank index... ],
      rows: [ {id, outcome, fields:{...all columns...}} ],
      summary: [ {label, ...summaryCols} ],   // the top stats block (if present)
    } ],
    generated, source, source_url
  }

BID ID SCHEME (consistent with precon-pipeline.js + /api/precon-log):
  bidId = "Project Number" when present, else a hash of (Project Name + GC).
  Collision-safe (first occurrence keeps the plain id; Nth dup suffixed).

GATING: /data/precon-historical.js -> preconstruction (admin/partner/business_dev;
field_ops BLOCKED).

Thread caps set BEFORE importing openpyxl (solved-box-gotchas).

Usage:
  python3 build-precon-historical.py
  python3 build-precon-historical.py --local sync/downloads/Project_Bid_Log.xlsx
"""

# --- thread caps BEFORE any heavy import ---
import os
for _v in ("OPENBLAS_NUM_THREADS", "OMP_NUM_THREADS", "MKL_NUM_THREADS",
           "NUMEXPR_NUM_THREADS", "VECLIB_MAXIMUM_THREADS"):
    os.environ.setdefault(_v, "1")

import sys
import re
import json
import hashlib
import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
PLATFORM = os.path.dirname(HERE)
DATA_DIR = os.path.join(PLATFORM, "data")
OUT_JS = os.path.join(DATA_DIR, "precon-historical.js")
DEFAULT_XLSX = os.path.join(HERE, "downloads", "Project_Bid_Log.xlsx")

SOURCE_FILE = "Project Bid Log.xlsx"
SOURCE_DESC = "SharePoint/01 - Admin/13 - Master Spreadsheets/Project Bid Log.xlsx"
BID_LOG_WEBURL = (
    "https://pierfoundations.sharepoint.com/sites/pf/_layouts/15/Doc.aspx?"
    "sourcedoc=%7B04D5671F-A207-46E9-B6A9-1D3ED192DE2A%7D&file=Project%20Bid%20Log.xlsx"
    "&action=default&mobileredirect=true"
)

WIN_TAB_RE = re.compile(r"^(\d{4})\s+win\s*loss\s*log$", re.IGNORECASE)
NAME_HEADER = "Project Name"
NUM_HEADER = "Project Number"
GC_HEADER = "General Contractor"
RANK_HEADER = "Number"          # col A rank index, dropped from fields
MAX_COLS = 60                   # ignore stray far-right cells (sheet reports 16384)

WIN_STATUSES = ("completed", "awarded")
LOSS_PREFIXES = ("not awarded", "will not bid", "lost")


def _token():
    sys.path.insert(0, "/home/aiciv/tools")
    from pf_email import _token as t
    return t()


def download_bid_log():
    import urllib.request
    tok = _token()
    drive = os.environ["SP_DRIVE_ID"]
    import urllib.parse
    q = urllib.parse.quote("Project Bid Log")
    url = "https://graph.microsoft.com/v1.0/drives/%s/root/search(q='%s')" % (drive, q)
    req = urllib.request.Request(url, headers={"Authorization": "Bearer " + tok})
    res = json.load(urllib.request.urlopen(req, timeout=120))
    fid = None
    for it in res.get("value", []):
        if it.get("name") == "Project Bid Log.xlsx":
            fid = it.get("id"); break
    if not fid:
        raise SystemExit("Could not find Project Bid Log.xlsx in SharePoint drive")
    curl = "https://graph.microsoft.com/v1.0/drives/%s/items/%s/content" % (drive, fid)
    creq = urllib.request.Request(curl, headers={"Authorization": "Bearer " + tok})
    return urllib.request.urlopen(creq, timeout=180).read()


def cellval(v):
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def norm(s):
    return re.sub(r"\s+", " ", (s or "").strip()).lower()


def collision_id(prefix, seed, seen):
    n = seen.get(seed, 0) + 1
    seen[seed] = n
    eff = seed if n == 1 else (seed + "#" + str(n))
    h = hashlib.sha1(eff.encode("utf-8")).hexdigest()[:10]
    return prefix + "_" + h, (n > 1)


def find_header_row(rows):
    """The header row is the one containing 'Project Name' + 'Bid Status'."""
    for i, r in enumerate(rows):
        vals = [norm(cellval(c)) for c in r[:MAX_COLS]]
        if "project name" in vals and "bid status" in vals:
            return i
    return None


def classify(status):
    s = norm(status)
    if not s:
        return "other"
    if s in WIN_STATUSES:
        return "win"
    for p in LOSS_PREFIXES:
        if s.startswith(p):
            return "loss"
    return "other"


def build_year(ws, year, omitted, id_warnings):
    rows = list(ws.iter_rows(values_only=True))
    hr = find_header_row(rows)
    if hr is None:
        omitted.append("%s: no header row (Project Name + Bid Status) found" % ws.title)
        return None

    # ordered headers (col_index, name); drop the rank-index "Number" column.
    headers = []
    for j, c in enumerate(rows[hr][:MAX_COLS]):
        name = cellval(c)
        if name and norm(name) != norm(RANK_HEADER):
            headers.append((j, name))
    col_names = [n for (_, n) in headers]

    name_col = next((j for (j, n) in headers if n == NAME_HEADER), None)
    status_col = next((j for (j, n) in headers if norm(n) == "bid status"), None)
    num_col = next((j for (j, n) in headers if n == NUM_HEADER), None)
    gc_col = next((j for (j, n) in headers if n == GC_HEADER), None)

    # ---- summary stats block ABOVE the header row (rows 0..hr-1) ----
    summary = []
    for i in range(0, hr):
        r = rows[i]
        label = cellval(r[1]) if len(r) > 1 else ""
        if not label:
            continue
        # capture the up-to-3 numeric/text columns that follow (C,D,E)
        entry = {"label": label}
        for k, key in ((2, "totalCount"), (3, "prelimUsed"), (4, "prelimPct")):
            if len(r) > k and r[k] is not None and str(r[k]).strip():
                entry[key] = cellval(r[k])
        if len(entry) > 1:
            summary.append(entry)

    # ---- bid rows BELOW the header ----
    out_rows = []
    seen = {}
    for i in range(hr + 1, len(rows)):
        r = rows[i]
        name = cellval(r[name_col]) if (name_col is not None and len(r) > name_col) else ""
        if not name:
            continue
        status = cellval(r[status_col]) if (status_col is not None and len(r) > status_col) else ""
        # skip repeated header rows (a row whose Bid Status literally reads "Bid Status")
        if norm(status) == "bid status" or norm(name) == "project name":
            continue
        fields = {}
        for (j, nm) in headers:
            fields[nm] = cellval(r[j]) if j < len(r) else ""
        num = cellval(r[num_col]) if (num_col is not None and len(r) > num_col) else ""
        gc = cellval(r[gc_col]) if (gc_col is not None and len(r) > gc_col) else ""
        # BID ID: Project Number when present, else hash of name+GC.
        if num:
            seed = "num:" + norm(num)
        else:
            seed = "ng:" + norm(name) + "|" + norm(gc)
        bid_id, collided = collision_id("hbid", seed, seen)
        if collided:
            id_warnings.append("%s: duplicate bid seed -> suffixed (%r / %r)" % (ws.title, name, num))
        out_rows.append({"id": bid_id, "outcome": classify(status), "fields": fields})

    return {
        "year": year,
        "tab": ws.title,
        "columns": col_names,
        "rows": out_rows,
        "summary": summary,
    }


def build(xlsx_bytes):
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    omitted = []
    id_warnings = []
    years = []
    win_tabs = [s for s in wb.sheetnames if WIN_TAB_RE.match(s or "")]
    if not win_tabs:
        omitted.append("No '<year> Win Loss Log' tab found in workbook")
    for tab in win_tabs:
        m = WIN_TAB_RE.match(tab)
        year = int(m.group(1))
        yd = build_year(wb[tab], year, omitted, id_warnings)
        if yd:
            years.append(yd)
    wb.close()
    years.sort(key=lambda y: y["year"], reverse=True)  # newest first
    out = {
        "years": years,
        "generated": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source": SOURCE_DESC,
        "source_url": BID_LOG_WEBURL,
    }
    if omitted:
        out["_omitted"] = omitted
    return out, omitted, id_warnings


def write_js(data):
    note = (
        "// AUTO-GENERATED by sync/build-precon-historical.py — DO NOT EDIT BY HAND.\n"
        "// Source: %s  generated %s\n"
        "// Historical Win/Loss bid log per year (read-only, all columns). Cells\n"
        "// with spreadsheet errors (#REF!/#DIV/0!) are kept verbatim, not computed.\n"
        "// ACCESS: preconstruction — admin/partner/business_dev (field_ops BLOCKED).\n"
        % (SOURCE_DESC, data["generated"])
    )
    with open(OUT_JS, "w") as f:
        f.write(note + "window.PF_PRECON_HISTORICAL = " + json.dumps(data, indent=2, ensure_ascii=False) + ";\n")
    print("wrote", OUT_JS)
    for y in data["years"]:
        wins = sum(1 for r in y["rows"] if r["outcome"] == "win")
        losses = sum(1 for r in y["rows"] if r["outcome"] == "loss")
        other = len(y["rows"]) - wins - losses
        print("  %d [%s]: %d rows (win %d / loss %d / other %d), %d columns, %d summary lines"
              % (y["year"], y["tab"], len(y["rows"]), wins, losses, other, len(y["columns"]), len(y["summary"])))


def main():
    if "--local" in sys.argv:
        path = sys.argv[sys.argv.index("--local") + 1]
    elif os.path.exists(DEFAULT_XLSX):
        path = DEFAULT_XLSX
    else:
        path = None
    if path:
        with open(path, "rb") as f:
            xlsx = f.read()
        print("source:", path)
    else:
        xlsx = download_bid_log()
        print("source: SharePoint download")

    data, omitted, id_warnings = build(xlsx)
    write_js(data)
    if omitted:
        for o in omitted:
            print("  OMITTED/NOTE:", o)
    else:
        print("  OMITTED: none")
    if id_warnings:
        print("  ID WARNINGS: %d duplicate bid seeds collision-suffixed" % len(id_warnings))
    # verify id uniqueness within each year
    ok = True
    for y in data["years"]:
        ids = [r["id"] for r in y["rows"]]
        if len(set(ids)) != len(ids):
            ok = False
            print("  ID UNIQUENESS: FAIL in %d" % y["year"])
    if not ok:
        sys.exit(1)
    total = sum(len(y["rows"]) for y in data["years"])
    print("  ID UNIQUENESS: OK — %d ids unique within each year" % total)


if __name__ == "__main__":
    main()
