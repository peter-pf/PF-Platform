#!/usr/bin/env python3
"""
Project Dashboard tracked-items ingest (build-project-dashboard.py)
==================================================================
Reads the PF Project Master "Project Dashboard" tab and emits per-project tracked
items (the "what is needed to complete this project" checklist). INSPECTED
STRUCTURE (verified, NOT assumed):

  The tab is TRANSPOSED: each PROJECT is a COLUMN (row 2 holds the PF Project
  Number per column: D=26-002, F=26-007, ...), and the TRACKED ITEMS are the
  ROWS. The item LABEL is in cols A-C (col A = section marker, col B = sub-label,
  col C sometimes continues it). Each project column holds that item's value.

  Sections (col A markers): General Info, PF Team, Contract Info, Engineering &
  Design, Project Safety, Site Readiness, Equipment, Material, QA / QC,
  Financials / Invoicing / Pay Application, Project Closeout.

  FINANCIALS OMITTED: the "Financials / Invoicing / Pay Application" section
  (G702/G703, retainage billing/paid, invoice dates) is DELIBERATELY NOT
  ingested - field_ops sees zero financials and budget-vs-costs is deferred per
  Brad. Those rows are skipped and reported in _omitted.

Output: platform/data/project-dashboard.js
  window.PF_PROJECT_DASHBOARD = {
    projects: [ { projectNumber, name, sections:[ {name, items:[{label, value}] } ] } ],
    sections: [ ...section names, in order, excluding Financials... ],
    generated, source, sourceTab
  }

GATING: /data/project-dashboard.js -> financials (PROJECT-level, admin/partner/
business_dev; field_ops BLOCKED). It carries per-project operational tracking,
no company-wide financials, and the Financials section is omitted entirely.

Thread caps set BEFORE importing openpyxl (solved-box-gotchas).

Usage:
  python3 build-project-dashboard.py
  python3 build-project-dashboard.py --local sync/downloads/PF_Project_Master.xlsx
"""

import os
for _v in ("OPENBLAS_NUM_THREADS", "OMP_NUM_THREADS", "MKL_NUM_THREADS",
           "NUMEXPR_NUM_THREADS", "VECLIB_MAXIMUM_THREADS"):
    os.environ.setdefault(_v, "1")

import sys
import re
import json
import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
PLATFORM = os.path.dirname(HERE)
DATA_DIR = os.path.join(PLATFORM, "data")
OUT_JS = os.path.join(DATA_DIR, "project-dashboard.js")
DEFAULT_XLSX = os.path.join(HERE, "downloads", "PF_Project_Master.xlsx")

SOURCE_FILE = "PF Project Master.xlsx"
TAB = "Project Dashboard"
NUM_ROW = 2          # row 2 (1-indexed) holds the PF Project Number per column
NAME_ROW = 3         # row 3 holds the Project Name per column
DATA_START_ROW = 4   # tracked items start at row 4

# Section names we OMIT (financials / budget-vs-costs deferred + field_ops safety)
OMIT_SECTIONS = {"financials"}


def _token():
    sys.path.insert(0, "/home/aiciv/tools")
    from pf_email import _token as t
    return t()


def download_master():
    import urllib.request
    import urllib.parse
    tok = _token()
    drive = os.environ["SP_DRIVE_ID"]
    q = urllib.parse.quote("PF Project Master")
    url = "https://graph.microsoft.com/v1.0/drives/%s/root/search(q='%s')" % (drive, q)
    req = urllib.request.Request(url, headers={"Authorization": "Bearer " + tok})
    res = json.load(urllib.request.urlopen(req, timeout=120))
    fid = None
    for it in res.get("value", []):
        if str(it.get("name", "")).lower().startswith("pf project master") or it.get("name") == "PF Project Master.xlsx":
            fid = it.get("id"); break
    if not fid:
        raise SystemExit("Could not find PF Project Master.xlsx in SharePoint drive")
    curl = "https://graph.microsoft.com/v1.0/drives/%s/items/%s/content" % (drive, fid)
    creq = urllib.request.Request(curl, headers={"Authorization": "Bearer " + tok})
    return urllib.request.urlopen(creq, timeout=180).read()


def cellval(v):
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def label_for(r):
    """Item label from cols A/B/C (idx 0,1,2). Col A is a section marker; the
    item sub-label is col B (+ col C continuation)."""
    parts = []
    for k in (1, 2):  # B + C make the item label; A is the section
        if k < len(r):
            v = cellval(r[k])
            if v:
                parts.append(v)
    return " ".join(parts).strip()


def build(xlsx_bytes):
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    omitted = []
    if TAB not in wb.sheetnames:
        wb.close()
        return ({"projects": [], "sections": [],
                 "generated": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                 "source": SOURCE_FILE, "sourceTab": TAB, "_omitted": ["Tab missing: %r" % TAB]},
                ["Tab missing: %r" % TAB])

    ws = wb[TAB]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # project columns: row 2 (idx1) cells that hold a real number (skip label col
    # + the "26-XXX" placeholder).
    numrow = rows[NUM_ROW - 1] if len(rows) >= NUM_ROW else ()
    namerow = rows[NAME_ROW - 1] if len(rows) >= NAME_ROW else ()
    proj_cols = []  # (col_idx, number, name)
    for j, c in enumerate(numrow):
        num = cellval(c)
        if not num or num.upper().endswith("XXX") or num == "PF Project Number":
            continue
        if not re.match(r"^\d{2}-\d", num):
            continue
        name = cellval(namerow[j]) if j < len(namerow) else ""
        proj_cols.append((j, num, name))

    # walk rows from DATA_START_ROW; track the current section from col A (idx0).
    # Build, per project, sections -> items (skip omitted sections + blank labels).
    sections_order = []
    seen_sections = set()
    # per-project accumulator
    per_proj = {num: {"projectNumber": num, "name": name, "sections": [], "_byname": {}}
                for (_, num, name) in proj_cols}

    cur_section = None
    cur_section_omit = False
    for i in range(DATA_START_ROW - 1, len(rows)):
        r = rows[i]
        a = cellval(r[0]) if len(r) > 0 else ""
        if a:
            cur_section = a
            cur_section_omit = (a.strip().lower().split(" ")[0] in OMIT_SECTIONS
                                or a.strip().lower().startswith("financ"))
            if cur_section_omit:
                if a not in [o for o in omitted]:
                    omitted.append("Section omitted (financials/budget deferred): %r" % a)
                continue
            if a not in seen_sections:
                seen_sections.add(a)
                sections_order.append(a)
        if cur_section_omit or cur_section is None:
            continue
        label = label_for(r)
        if not label:
            continue
        for (j, num, _name) in proj_cols:
            val = cellval(r[j]) if j < len(r) else ""
            if not val:
                continue  # skip empty cells (only show items with a value)
            pp = per_proj[num]
            sec = pp["_byname"].get(cur_section)
            if sec is None:
                sec = {"name": cur_section, "items": []}
                pp["_byname"][cur_section] = sec
                pp["sections"].append(sec)
            sec["items"].append({"label": label, "value": val})

    projects = []
    for (_, num, _name) in proj_cols:
        pp = per_proj[num]
        del pp["_byname"]
        projects.append(pp)

    out = {
        "projects": projects,
        "sections": sections_order,
        "generated": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source": SOURCE_FILE,
        "sourceTab": TAB,
    }
    if omitted:
        out["_omitted"] = omitted
    return out, omitted


def write_js(data):
    note = (
        "// AUTO-GENERATED by sync/build-project-dashboard.py - DO NOT EDIT BY HAND.\n"
        "// Source: %s :: '%s' tab (TRANSPOSED: projects are columns, tracked items\n"
        "// are rows).  generated %s\n"
        "// Per-project tracked items (what is needed to complete). The Financials /\n"
        "// Invoicing section is OMITTED (deferred + field_ops safety).\n"
        "// ACCESS: financials (PROJECT-level) - admin/partner/business_dev; field_ops BLOCKED.\n"
        % (SOURCE_FILE, TAB, data["generated"])
    )
    with open(OUT_JS, "w") as f:
        f.write(note + "window.PF_PROJECT_DASHBOARD = " + json.dumps(data, indent=2, ensure_ascii=False) + ";\n")
    print("wrote", OUT_JS)
    print("  projects: %d" % len(data["projects"]))
    print("  sections (ingested): %s" % ", ".join(data["sections"]))
    for p in data["projects"][:5]:
        ni = sum(len(s["items"]) for s in p["sections"])
        print("    %s %s: %d tracked items across %d sections"
              % (p["projectNumber"], (p["name"] or "")[:24], ni, len(p["sections"])))


def main():
    if "--local" in sys.argv:
        path = sys.argv[sys.argv.index("--local") + 1]
    elif os.path.exists(DEFAULT_XLSX):
        path = DEFAULT_XLSX
    else:
        path = None
    if path:
        with open(path, "rb") as f:
            xlsx = f.read()
        print("source:", path)
    else:
        xlsx = download_master()
        print("source: SharePoint download")

    data, omitted = build(xlsx)
    write_js(data)
    if omitted:
        for o in omitted:
            print("  OMITTED:", o)
    else:
        print("  OMITTED: none")


if __name__ == "__main__":
    main()
