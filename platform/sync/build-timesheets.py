#!/usr/bin/env python3
"""
Build TimeSheets — Field Operations weekly timesheet allocations  [READ-ONLY]
=============================================================================
Pulls the PF weekly timesheet workbook from SharePoint and writes
platform/data/timesheets.js (window.PF_TIMESHEETS = {...}), which the
Field Operations -> TimeSheets panel reads.

SOURCE (real, no fabrication):
  SharePoint `05 - Field Operations / Timesheets / 2026 PF Timesheets.xlsx`.
  ONE workbook, ONE SHEET PER WEEK named by date range ("2.1-2.7", "6.7-6.13", ...).
  Two NON-week sheets are SKIPPED: "Timesheet Template" and "PF Cost Codes".

WEEKLY SHEET LAYOUT (verified 2026-06-17):
  A1 "Pier Foundations Weekly Timesheet"
  A2 "Crew Number: NN"
  A4 "Week Starting:"  D4 <date>   F4 "Week Ending:"  H4 <date>
  Then STACKED EMPLOYEE BLOCKS (~15 rows each):
    "Employee Name:" (col A, value col D)   "Title:" (col F, value col H)
    "Employee Number:" (val D)              "Status:" (val H)
    "Department:" (val D)                   "Supervisor:" (val H)
    header row: Day|Date|Job #|Cost Code|Activities|Start|End|Regular|OT|Total|Per Diem  (cols A..K)
    7 day rows (Sunday..Saturday): Job# (C), Cost Code (D), Activity (E),
        Start (F), End (G), Regular (H), OT (I), Total (J), Per Diem (K)
    "Totals:" row (col G label; H/I/J totals; K per-diem count)
    "Employee Notes:" row
  Blocks are detected by scanning for "Employee Name:" rows (NOT hard-coded offsets).
  An employee block with no name is skipped (blank template slot).

Per week we compute AGGREGATES:
  - hours by Cost Code, hours by Job #, and overall regular/OT/total + per-diem nights.

Reuses the Graph token helper + env loader from pf_email.py (same pattern as
build-project-record.py). Thread caps set BEFORE openpyxl import (box ~300 pid
limit; solved-box-gotchas).

Usage:
  python3 build-timesheets.py            # writes data/timesheets.js
  python3 build-timesheets.py --dump     # print the latest-week-with-hours summary, no write
"""

# --- thread caps BEFORE any heavy import (solved-box-gotchas) ---
import os
for _v in ("OPENBLAS_NUM_THREADS", "OMP_NUM_THREADS", "MKL_NUM_THREADS",
           "NUMEXPR_NUM_THREADS", "VECLIB_MAXIMUM_THREADS"):
    os.environ.setdefault(_v, "1")

import io
import re
import sys
import json
import urllib.request
import urllib.parse
import urllib.error
from datetime import datetime, timezone, date, time

# --- reuse the Graph token helper + env loader from pf_email.py ---
sys.path.insert(0, "/home/aiciv/tools")
from pf_email import _token, _env  # noqa: E402

GRAPH = "https://graph.microsoft.com/v1.0"

HERE = os.path.dirname(os.path.abspath(__file__))
PLATFORM = os.path.dirname(HERE)
DATA_DIR = os.path.join(PLATFORM, "data")
OUT_JS = os.path.join(DATA_DIR, "timesheets.js")

SP_FILE = "05 - Field Operations/Timesheets/2026 PF Timesheets.xlsx"
COST_CODE_SHEET = "PF Cost Codes"
SKIP_SHEETS = {"Timesheet Template", "PF Cost Codes"}

_env_cache = _env()
DRIVE_ID = _env_cache.get("SP_DRIVE_ID", "")

DAYS = {"sunday", "monday", "tuesday", "wednesday", "thursday", "friday", "saturday"}


# ---------------- Graph helpers (pattern from build-project-record.py) ----------------
def gget(token, url):
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    return json.loads(urllib.request.urlopen(req).read())


def get_item_by_path(token, path):
    p = urllib.parse.quote(path)
    url = f"{GRAPH}/drives/{DRIVE_ID}/root:/{p}"
    try:
        return gget(token, url)
    except urllib.error.HTTPError:
        return None


def download_path(token, path):
    p = urllib.parse.quote(path)
    url = f"{GRAPH}/drives/{DRIVE_ID}/root:/{p}:/content"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    return urllib.request.urlopen(req).read()


# ---------------- cell helpers ----------------
def _clean(v):
    """Trim strings, drop N/A placeholders. Keeps numbers/None for caller to handle."""
    if v is None:
        return ""
    if isinstance(v, str):
        s = v.strip()
        if s.upper() in ("N/A", "NA", "TBD"):
            return ""
        return s
    return v


def _str(v):
    c = _clean(v)
    return "" if c == "" else str(c).strip()


def _num(v):
    """Return a float for a numeric cell, else 0.0 (blank stays blank/zero per spec)."""
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _date_str(v):
    """Format a real spreadsheet date as YYYY-MM-DD. Reject the 1900-epoch
    placeholder dates Excel leaves in empty template rows (year < 2015)."""
    if isinstance(v, datetime):
        if v.year < 2015:
            return ""
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        if v.year < 2015:
            return ""
        return v.strftime("%Y-%m-%d")
    if isinstance(v, time):
        return ""
    s = _str(v)
    return s


def _time_str(v):
    """Format a start/end time cell as HH:MM (24h). Blank stays blank."""
    if isinstance(v, time):
        return v.strftime("%H:%M")
    if isinstance(v, datetime):
        return v.strftime("%H:%M") if (v.hour or v.minute) else ""
    return _str(v)


def _job(v):
    """Job # is a string like '26-005'. Numbers get stringified plainly."""
    c = _clean(v)
    if c == "":
        return ""
    if isinstance(c, float) and c.is_integer():
        return str(int(c))
    return str(c).strip()


def _code(v):
    """Cost code is a labor code like 5220. Render as a plain integer string."""
    c = _clean(v)
    if c == "":
        return ""
    if isinstance(c, (int, float)) and float(c).is_integer():
        return str(int(c))
    return str(c).strip()


# ---------------- PF Cost Codes map ----------------
def parse_cost_codes(wb):
    """Build {code(str): name(str)} from the 'PF Cost Codes' sheet.
    Codes live in col1 (parent groups) or col2 (subcodes); the human name is in
    the next non-empty description column (col2 for parents, col3 for subcodes)."""
    out = {}
    if COST_CODE_SHEET not in wb.sheetnames:
        return out
    ws = wb[COST_CODE_SHEET]
    for r in range(1, ws.max_row + 1):
        c1 = ws.cell(row=r, column=1).value
        c2 = ws.cell(row=r, column=2).value
        c3 = ws.cell(row=r, column=3).value
        # parent group row: code in col1, name in col2
        if isinstance(c1, (int, float)) and float(c1).is_integer():
            name = _str(c2)
            if name:
                out[str(int(c1))] = name
        # subcode row: code in col2, name in col3
        if isinstance(c2, (int, float)) and float(c2).is_integer():
            name = _str(c3)
            if name:
                out[str(int(c2))] = name
    return out


# ---------------- weekly sheet parse ----------------
def _is_week_sheet(name):
    if name in SKIP_SHEETS:
        return False
    # week sheets look like "2.1-2.7" or "2.22 - 2.28" (dot-dates joined by a dash)
    return bool(re.match(r"^\s*\d{1,2}\.\d{1,2}\s*-\s*\d{1,2}\.\d{1,2}\s*$", name))


def parse_week(ws, label):
    """Parse one weekly sheet into a structured dict."""
    maxr = ws.max_row

    crew = ""
    week_start = ""
    week_end = ""

    # Scan the top for crew + week dates (don't hard-code rows).
    for r in range(1, min(maxr, 8) + 1):
        a = _str(ws.cell(row=r, column=1).value)
        f = _str(ws.cell(row=r, column=6).value)
        if a.lower().startswith("crew number"):
            # value may be inline after the colon, or in col D
            m = re.search(r"crew number:?\s*(.+)$", a, re.I)
            inline = m.group(1).strip() if m and m.group(1).strip() else ""
            crew = inline or _str(ws.cell(row=r, column=4).value)
        if a.lower().startswith("week starting"):
            week_start = _date_str(ws.cell(row=r, column=4).value)
        if f.lower().startswith("week ending"):
            week_end = _date_str(ws.cell(row=r, column=8).value)

    # Find employee-block start rows by scanning col A for "Employee Name:".
    block_starts = []
    for r in range(1, maxr + 1):
        a = _str(ws.cell(row=r, column=1).value)
        if a.lower().startswith("employee name"):
            block_starts.append(r)

    employees = []
    for r0 in block_starts:
        emp = _parse_block(ws, r0, maxr)
        if emp:
            employees.append(emp)

    # ---- aggregates ----
    by_code = {}   # code -> {regular, ot, total}
    by_job = {}    # job  -> {regular, ot, total}
    tot_reg = tot_ot = tot_total = 0.0
    per_diem_nights = 0

    for emp in employees:
        for d in emp["days"]:
            tot_reg += d["regular"]
            tot_ot += d["ot"]
            tot_total += d["total"]
            if str(d.get("per_diem", "")).strip().upper() == "Y":
                per_diem_nights += 1
            code = d["cost_code"] or "(uncoded)"
            job = d["job"] or "(no job)"
            if d["total"] or d["regular"] or d["ot"]:
                bc = by_code.setdefault(code, {"regular": 0.0, "ot": 0.0, "total": 0.0})
                bc["regular"] += d["regular"]; bc["ot"] += d["ot"]; bc["total"] += d["total"]
                bj = by_job.setdefault(job, {"regular": 0.0, "ot": 0.0, "total": 0.0})
                bj["regular"] += d["regular"]; bj["ot"] += d["ot"]; bj["total"] += d["total"]

    def _round(x):
        return round(x, 2)

    by_code_list = [{"code": k, "regular": _round(v["regular"]), "ot": _round(v["ot"]),
                     "total": _round(v["total"])} for k, v in by_code.items()]
    by_code_list.sort(key=lambda x: (-x["total"], x["code"]))
    by_job_list = [{"job": k, "regular": _round(v["regular"]), "ot": _round(v["ot"]),
                    "total": _round(v["total"])} for k, v in by_job.items()]
    by_job_list.sort(key=lambda x: (-x["total"], x["job"]))

    has_hours = tot_total > 0 or tot_reg > 0 or tot_ot > 0

    return {
        "week_label": label,
        "week_start": week_start,
        "week_end": week_end,
        "crew": crew,
        "employees": employees,
        "by_cost_code": by_code_list,
        "by_job": by_job_list,
        "totals": {"regular": _round(tot_reg), "ot": _round(tot_ot),
                   "total": _round(tot_total), "per_diem_nights": per_diem_nights},
        "has_hours": has_hours,
        "employee_count": len([e for e in employees if e["days_with_hours"] > 0]),
    }


def _parse_block(ws, r0, maxr):
    """Parse one employee block starting at the 'Employee Name:' row r0.
    Returns an employee dict, or None if the block has no employee name."""
    name = _str(ws.cell(row=r0, column=4).value)
    title = _str(ws.cell(row=r0, column=8).value)
    if not name:
        return None  # blank template slot

    number = title2 = status = department = supervisor = ""
    # The next few rows hold Employee Number / Department / Title / Status / Supervisor.
    for r in range(r0 + 1, min(r0 + 4, maxr) + 1):
        a = _str(ws.cell(row=r, column=1).value).lower()
        f = _str(ws.cell(row=r, column=6).value).lower()
        if a.startswith("employee number"):
            number = _str(ws.cell(row=r, column=4).value)
        elif a.startswith("department"):
            department = _str(ws.cell(row=r, column=4).value)
        if f.startswith("status"):
            status = _str(ws.cell(row=r, column=8).value)
        elif f.startswith("supervisor"):
            supervisor = _str(ws.cell(row=r, column=8).value)

    # Find this block's day-header row ("Day" in col A) below r0.
    header_row = None
    for r in range(r0 + 1, min(r0 + 8, maxr) + 1):
        if _str(ws.cell(row=r, column=1).value).lower() == "day":
            header_row = r
            break
    days = []
    emp_reg = emp_ot = emp_total = 0.0
    emp_per_diem = 0
    if header_row:
        for r in range(header_row + 1, maxr + 1):
            day = _str(ws.cell(row=r, column=1).value)
            if day.lower() not in DAYS:
                break  # reached Totals / Notes / next block
            d_date = _date_str(ws.cell(row=r, column=2).value)
            job = _job(ws.cell(row=r, column=3).value)
            code = _code(ws.cell(row=r, column=4).value)
            activity = _str(ws.cell(row=r, column=5).value)
            start = _time_str(ws.cell(row=r, column=6).value)
            end = _time_str(ws.cell(row=r, column=7).value)
            reg = _num(ws.cell(row=r, column=8).value)
            ot = _num(ws.cell(row=r, column=9).value)
            total = _num(ws.cell(row=r, column=10).value)
            per_diem = _str(ws.cell(row=r, column=11).value)
            days.append({
                "day": day, "date": d_date, "job": job, "cost_code": code,
                "activity": activity, "start": start, "end": end,
                "regular": round(reg, 2), "ot": round(ot, 2), "total": round(total, 2),
                "per_diem": per_diem,
            })
            emp_reg += reg; emp_ot += ot; emp_total += total
            if per_diem.strip().upper() == "Y":
                emp_per_diem += 1

    days_with_hours = sum(1 for d in days if d["total"] or d["regular"] or d["ot"])
    return {
        "name": name, "number": number, "title": title, "status": status,
        "department": department, "supervisor": supervisor,
        "days": days,
        "totals": {"regular": round(emp_reg, 2), "ot": round(emp_ot, 2),
                   "total": round(emp_total, 2), "per_diem_nights": emp_per_diem},
        "days_with_hours": days_with_hours,
    }


def aggregate_by_employee_year(weeks):
    """Sum across ALL weekly sheets, per employee (keyed by name + number), for the
    whole year: Regular, OT, Total hours and Per Diem nights. Returns a list of
    employee rows (sorted by Total desc) plus a grand-total row."""
    acc = {}      # key -> row dict
    order = []    # preserve first-seen order for stable tie-breaks
    g_reg = g_ot = g_total = 0.0
    g_pd = 0

    for w in weeks:
        for e in w.get("employees", []):
            name = (e.get("name") or "").strip()
            if not name:
                continue
            number = (e.get("number") or "").strip()
            t = e.get("totals", {}) or {}
            reg = float(t.get("regular", 0) or 0)
            ot = float(t.get("ot", 0) or 0)
            total = float(t.get("total", 0) or 0)
            pd = int(t.get("per_diem_nights", 0) or 0)

            key = (name.lower(), number.lower())
            if key not in acc:
                acc[key] = {"name": name, "number": number, "title": e.get("title", "") or "",
                            "regular": 0.0, "ot": 0.0, "total": 0.0, "per_diem_nights": 0}
                order.append(key)
            row = acc[key]
            row["regular"] += reg
            row["ot"] += ot
            row["total"] += total
            row["per_diem_nights"] += pd
            # keep the first non-empty title we encounter for this employee
            if not row["title"] and (e.get("title") or "").strip():
                row["title"] = e["title"].strip()

            g_reg += reg
            g_ot += ot
            g_total += total
            g_pd += pd

    rows = []
    for key in order:
        r = acc[key]
        rows.append({"name": r["name"], "number": r["number"], "title": r["title"],
                     "regular": round(r["regular"], 2), "ot": round(r["ot"], 2),
                     "total": round(r["total"], 2),
                     "per_diem_nights": r["per_diem_nights"]})
    rows.sort(key=lambda x: (-x["total"], x["name"].lower()))

    grand = {"regular": round(g_reg, 2), "ot": round(g_ot, 2),
             "total": round(g_total, 2), "per_diem_nights": g_pd}
    return rows, grand


def _week_sort_key(label):
    """Chronological sort key from a 'M.D-M.D' label using the START date.
    Assumes calendar year of the workbook (2026); only relative order matters."""
    m = re.match(r"^\s*(\d{1,2})\.(\d{1,2})", label)
    if not m:
        return (99, 99)
    return (int(m.group(1)), int(m.group(2)))


# ---------------- assemble + write ----------------
def assemble(token):
    import openpyxl
    raw = download_path(token, SP_FILE)
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)

    cost_codes = parse_cost_codes(wb)

    weeks = []
    for name in wb.sheetnames:
        if not _is_week_sheet(name):
            continue
        weeks.append(parse_week(wb[name], name.strip()))

    weeks.sort(key=lambda w: _week_sort_key(w["week_label"]))

    # latest week that actually has hours (chronologically last with has_hours)
    latest_with_hours = ""
    for w in weeks:
        if w["has_hours"]:
            latest_with_hours = w["week_label"]

    item = get_item_by_path(token, SP_FILE)
    source_url = (item or {}).get("webUrl", "")

    # Annual per-employee aggregation across ALL weekly sheets.
    by_employee_year, year_grand_total = aggregate_by_employee_year(weeks)

    # Year label: derive from the workbook file name ("2026 PF Timesheets.xlsx"),
    # falling back to a week_start year, then the current UTC year.
    year_label = ""
    m = re.match(r"\s*(\d{4})", "2026 PF Timesheets.xlsx")
    if m:
        year_label = m.group(1)
    if not year_label:
        for w in weeks:
            ws = w.get("week_start", "")
            ym = re.match(r"^(\d{4})-", ws)
            if ym:
                year_label = ym.group(1)
                break
    if not year_label:
        year_label = str(datetime.now(timezone.utc).year)

    return {
        "generated": datetime.now(timezone.utc).isoformat() + "Z",
        "source_file": "2026 PF Timesheets.xlsx",
        "source_url": source_url,
        "cost_code_names": cost_codes,
        "latest_week_with_hours": latest_with_hours,
        "year": year_label,
        "by_employee_year": by_employee_year,
        "year_grand_total": year_grand_total,
        "weeks": weeks,
    }


def write_js(data):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(OUT_JS, "w") as f:
        f.write("// AUTO-GENERATED by sync/build-timesheets.py — do not edit by hand.\n")
        f.write("// Field Operations weekly timesheets (hours by Cost Code + Job #).\n")
        f.write("// Source: SharePoint 05 - Field Operations / Timesheets / 2026 PF Timesheets.xlsx\n")
        f.write("window.PF_TIMESHEETS = ")
        json.dump(data, f, indent=2)
        f.write(";\n")
    return OUT_JS


def _summary(data, only_latest=True):
    print(f"TimeSheets assembled: {len(data['weeks'])} week sheets, "
          f"{len(data['cost_code_names'])} cost codes mapped.")
    print(f"Latest week WITH hours: {data['latest_week_with_hours']!r}")

    # ---- ANNUAL PER-EMPLOYEE SUMMARY (proof for self-check) ----
    print(f"\n=== ANNUAL SUMMARY — {data.get('year','')} (across all {len(data['weeks'])} weeks) ===")
    print(f"  {'Employee':<26} {'#':<8} {'Reg':>9} {'OT':>8} {'Total':>9} {'PD':>4}")
    for r in data.get("by_employee_year", []):
        print(f"  {r['name']:<26} {str(r['number']):<8} "
              f"{r['regular']:>9} {r['ot']:>8} {r['total']:>9} {r['per_diem_nights']:>4}")
    g = data.get("year_grand_total", {})
    print(f"  {'GRAND TOTAL':<26} {'':<8} "
          f"{g.get('regular',0):>9} {g.get('ot',0):>8} {g.get('total',0):>9} {g.get('per_diem_nights',0):>4}")
    # cross-check: sum of weekly totals should equal the grand total
    wk_reg = round(sum(w['totals']['regular'] for w in data['weeks']), 2)
    wk_ot = round(sum(w['totals']['ot'] for w in data['weeks']), 2)
    wk_total = round(sum(w['totals']['total'] for w in data['weeks']), 2)
    wk_pd = sum(w['totals']['per_diem_nights'] for w in data['weeks'])
    print(f"  SUM-OF-WEEKLY-TOTALS:        "
          f"{wk_reg:>9} {wk_ot:>8} {wk_total:>9} {wk_pd:>4}  "
          f"-> match={'YES' if (wk_reg==g.get('regular') and wk_ot==g.get('ot') and wk_total==g.get('total') and wk_pd==g.get('per_diem_nights')) else 'NO'}")
    target = data["latest_week_with_hours"]
    for w in data["weeks"]:
        if only_latest and w["week_label"] != target:
            continue
        print(f"\n=== WEEK {w['week_label']}  ({w['week_start']} -> {w['week_end']})  "
              f"Crew {w['crew']}  has_hours={w['has_hours']} ===")
        t = w["totals"]
        print(f"  WEEK TOTALS: regular={t['regular']} ot={t['ot']} total={t['total']} "
              f"per_diem_nights={t['per_diem_nights']}  employees_with_hours={w['employee_count']}")
        print("  EMPLOYEES:")
        for e in w["employees"]:
            et = e["totals"]
            print(f"    - {e['name']} (#{e['number']}, {e['title']}): "
                  f"reg={et['regular']} ot={et['ot']} total={et['total']} "
                  f"per_diem={et['per_diem_nights']}  days_with_hours={e['days_with_hours']}")
        print("  BY COST CODE:")
        for c in w["by_cost_code"]:
            nm = data["cost_code_names"].get(c["code"], "")
            print(f"    [{c['code']}] {nm}: reg={c['regular']} ot={c['ot']} total={c['total']}")
        print("  BY JOB #:")
        for j in w["by_job"]:
            print(f"    {j['job']}: reg={j['regular']} ot={j['ot']} total={j['total']}")


def main():
    dump = "--dump" in sys.argv[1:]
    if not DRIVE_ID:
        print("ERROR: SP_DRIVE_ID not set in /home/aiciv/.env", file=sys.stderr)
        sys.exit(1)
    print(f"build-timesheets — {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC",
          file=sys.stderr)
    token = _token()
    data = assemble(token)
    _summary(data)
    if dump:
        return
    out = write_js(data)
    print(f"\nWrote: {out}", file=sys.stderr)


if __name__ == "__main__":
    main()
