#!/usr/bin/env python3
"""
Build Project Record — POET (26-002)  [Phase 2 v1, READ-ONLY]
=============================================================
Pulls the POET project's real data from SharePoint and writes
platform/data/project-record-poet.js (window.PF_PROJECT_POET = {...}),
which the per-project RECORD DETAIL VIEW reads.

This is the FIRST concrete project record. POET is the reference/template
for all projects (the 11-section schema in docs/portal-rebuild/PROJECT-RECORD-SCHEMA.md).

DATA SOURCED (real, no fabrication):
  - Contacts (Owner, GC, Engineering, PF Team, Vendors) + General Info GC/Engineering
    from the POET "...Project Info.xlsx" sheet "Project Name Project Contacts"
    (14 columns: Scope/Company/Primary Contact/Address/Business Phone/Cell Phone/Email/Website/.../Notes).
  - Section document links (Graph webUrl) for the relevant POET subfolders/files so each
    card can link to its docs (Engineering & Design, Subcontract, Invoicing, Safety, Field, GC Drawings).
  - QA/QC installed columns + installed LF from the auto-progress output
    (data/progress-data.js -> window.PF_PROGRESS["26-002"]). Baseline still pending -> no %.
  - Any field with no source -> rendered blank (em-dash) by the view. DO NOT fabricate.

Reuses the Graph auth + drive patterns from sp-sync.py / build-progress.py
(token via /home/aiciv/tools/pf_email.py _token(); SP_DRIVE_ID from /home/aiciv/.env).

Thread caps set BEFORE openpyxl import (this box, ~300 pid limit; solved-box-gotchas).

Usage:
  python3 build-project-record.py            # writes data/project-record-poet.js
  python3 build-project-record.py --dump     # print the assembled record to stdout, no write
"""

# --- thread caps BEFORE any heavy import (solved-box-gotchas) ---
import os
for _v in ("OPENBLAS_NUM_THREADS", "OMP_NUM_THREADS", "MKL_NUM_THREADS",
           "NUMEXPR_NUM_THREADS", "VECLIB_MAXIMUM_THREADS"):
    os.environ.setdefault(_v, "1")

import io
import re
import sys
import json
import urllib.request
import urllib.parse
import urllib.error
from datetime import datetime, timezone

# --- reuse the Graph token helper + env loader from pf_email.py ---
sys.path.insert(0, "/home/aiciv/tools")
from pf_email import _token, _env  # noqa: E402

GRAPH = "https://graph.microsoft.com/v1.0"

HERE = os.path.dirname(os.path.abspath(__file__))
PLATFORM = os.path.dirname(HERE)
DATA_DIR = os.path.join(PLATFORM, "data")
OUT_JS = os.path.join(DATA_DIR, "project-record-poet.js")
PROGRESS_JS = os.path.join(DATA_DIR, "progress-data.js")

# POET project identifiers (confirmed via SharePoint exploration 2026-06-17)
PROJECT_NUMBER = "26-002"
PROJECT_NAME = "POET Biosciences"
PROJECT_LOCATION = "Shelbyville, IN"
SP_PROJECT_FOLDER = "04 - Project Management/02 - Projects/26-002 - POET Projects - POET"
CONTACTS_FILE = "26-0330 - 26-002 - POET Biosciences - Project Info.xlsx"
CONTACTS_SHEET = "Project Name Project Contacts"

_env_cache = _env()
DRIVE_ID = _env_cache.get("SP_DRIVE_ID", "")


# ---------------- Graph helpers (pattern from sp-sync.py) ----------------
def gget(token, url):
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    return json.loads(urllib.request.urlopen(req).read())


def list_children_by_path(token, path):
    p = urllib.parse.quote(path)
    url = f"{GRAPH}/drives/{DRIVE_ID}/root:/{p}:/children"
    items = []
    while url:
        data = gget(token, url)
        items.extend(data.get("value", []))
        url = data.get("@odata.nextLink")
    return items


def get_item_by_path(token, path):
    """Return the driveItem metadata (incl webUrl) for a folder/file path, or None."""
    p = urllib.parse.quote(path)
    url = f"{GRAPH}/drives/{DRIVE_ID}/root:/{p}"
    try:
        return gget(token, url)
    except urllib.error.HTTPError:
        return None


def download_path(token, path):
    p = urllib.parse.quote(path)
    url = f"{GRAPH}/drives/{DRIVE_ID}/root:/{p}:/content"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    return urllib.request.urlopen(req).read()


# ---------------- contacts parse ----------------
def _clean(v):
    if v is None:
        return ""
    s = str(v).strip()
    # treat explicit N/A placeholders as empty so the view shows a blank, not "N/A"
    if s.upper() in ("N/A", "NA", "TBD"):
        return ""
    return s


def _contact(ws, r):
    """Build a contact dict from a data row (column indices are 1-based)."""
    return {
        "scope": _clean(ws.cell(row=r, column=2).value),
        "company": _clean(ws.cell(row=r, column=3).value),
        "name": _clean(ws.cell(row=r, column=4).value),
        "address": _clean(ws.cell(row=r, column=5).value),
        "phone": _clean(ws.cell(row=r, column=6).value),
        "cell": _clean(ws.cell(row=r, column=7).value),
        "email": _clean(ws.cell(row=r, column=8).value),
        "website": _clean(ws.cell(row=r, column=9).value),
        "notes": _clean(ws.cell(row=r, column=12).value),
    }


def _has_contact(c):
    return bool(c["company"] or c["name"] or c["email"])


def parse_contacts(token):
    """Read the POET contact log. Returns dict grouped by category."""
    import openpyxl
    path = f"{SP_PROJECT_FOLDER}/{CONTACTS_FILE}"
    raw = download_path(token, path)
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    if CONTACTS_SHEET not in wb.sheetnames:
        # fall back to the only sheet if exact name drifts
        ws = wb[wb.sheetnames[0]]
    else:
        ws = wb[CONTACTS_SHEET]

    # The sheet uses section header rows in column A (e.g. "Owner",
    # "General Contractor", "Engineering & Design", "Pier Foundations Project Team",
    # "Project Vendor Contacts"). Walk rows, track the current section, and bucket
    # each contact row into our schema groups.
    groups = {
        "owner": [],
        "gc": [],
        "engineering": [],
        "pf_team": [],
        "vendors": [],
    }

    SECTION_MAP = [
        (re.compile(r"\bowner\b", re.I), "owner"),
        (re.compile(r"general contractor", re.I), "gc"),
        (re.compile(r"engineering", re.I), "engineering"),
        (re.compile(r"pier foundations", re.I), "pf_team"),
        (re.compile(r"vendor", re.I), "vendors"),
    ]

    current = None
    title = _clean(ws.cell(row=1, column=1).value)
    for r in range(2, ws.max_row + 1):
        a = _clean(ws.cell(row=r, column=1).value)
        if a:
            matched = None
            for rx, key in SECTION_MAP:
                if rx.search(a):
                    matched = key
                    break
            # ignore non-section header rows like "Professional Sources"
            if matched:
                current = matched
            continue
        if current is None:
            continue
        c = _contact(ws, r)
        if _has_contact(c):
            groups[current].append(c)

    return {"title": title, "groups": groups, "source_file": CONTACTS_FILE}


# ---------------- bid log (General Info + Contract Info metrics/dates/award) ----------------
BID_LOG_FILE = "01 - Admin/13 - Master Spreadsheets/Project Bid Log.xlsx"
BID_LOG_SHEET = "Agg Pier Bid Log"  # AP jobs; "Helical Pier Bid Log" is the helical equivalent (later)
BID_PROJECT_NUMBER_COL = 2  # xl col B (0-based col1) = Project Number


def _xlclean(v):
    """Clean a cell value: trim strings, drop N/A placeholders, normalize datetimes to date."""
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        try:
            return v.strftime("%Y-%m-%d")
        except Exception:
            return str(v)
    s = str(v).strip()
    if s.upper() in ("N/A", "NA", "TBD"):
        return ""
    # trim ' 00:00:00' off any stringified datetime
    if s.endswith(" 00:00:00"):
        s = s[:-9]
    return s


def parse_bid_log(token):
    """Pull POET's row from the Project Bid Log (Agg Pier Bid Log sheet).
    Returns a flat dict of bid-log-sourced fields, or None if POET row not found.
    Column map confirmed against the live sheet 2026-06-17 (xl row 104).
    Field headers on xl row 6; data from xl row 7+. Match Project Number == '26-002'."""
    import openpyxl
    raw = download_path(token, BID_LOG_FILE)
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    if BID_LOG_SHEET not in wb.sheetnames:
        return None
    ws = wb[BID_LOG_SHEET]

    target_row = None
    for r in range(7, ws.max_row + 1):
        pn = ws.cell(row=r, column=BID_PROJECT_NUMBER_COL).value
        if pn and str(pn).strip() == PROJECT_NUMBER:
            target_row = r
            break
    if target_row is None:
        return None

    def cell(xl_col):
        return _xlclean(ws.cell(row=target_row, column=xl_col).value)

    # xl_col = 0-based-index + 1
    return {
        "row": target_row,
        # General Info
        "project_name": cell(3),         # col2 Project Name
        "city_state": cell(4),           # col3 City / State
        "address": cell(5),              # col4 Address
        "site_acres": cell(6),           # col5 Site Size (Acres)
        "total_sf": cell(7),             # col6 Total SF (Bldg Pad)
        "total_lf": cell(8),             # col7 Total LF
        "total_columns": cell(9),        # col8 Total Columns
        "total_stone_tn": cell(10),      # col9 Total Stone (TN)
        "duration_days": cell(11),       # col10 Project Duration (Days)
        "feed_type": cell(12),           # col11 Top vs Bottom Feed
        # Contract Info
        "tax": cell(13),                 # col12 Tax
        "min_insurance_req": cell(14),   # col13 Min Insurance Req
        "wage_req": cell(15),            # col14 Wage Req.
        "invite_date": cell(16),         # col15 Invite Date
        "due_date": cell(17),            # col16 Due Date
        "date_submitted": cell(18),      # col17 Date Submitted
        "projected_start": cell(19),     # col18 Projected Start Date
        "follow_up_date": cell(20),      # col19 Follow Up Date
        "bid_status": cell(21),          # col20 Bid Status (POET = Awarded)
        "bid_total_value": cell(22),     # col21 Bid Total Value
        # GC (fallback / award contact)
        "gc_name": cell(27),             # col26 General Contractor
        "gc_contact": cell(28),          # col27 Contact Name
        "gc_email": cell(29),            # col28 GC Email
        "gc_phone": cell(30),            # col29 GC Phone
        # Engineering
        "engineer_firm": cell(32),       # col31 Engineer / Design Firm
        "prelim_design_fee": cell(33),   # col32 Prelim Design Fee
        "design_completed_date": cell(34),  # col33 Design Completed Date
        "design_paid_date": cell(35),    # col34 Date Paid
        "source_file": BID_LOG_FILE,
        "source_sheet": BID_LOG_SHEET,
    }


# ---------------- subcontract (fully-executed) extraction ----------------
# POET has NO separate LOI; the fully-executed subcontract is the post-award source
# that fills General Info + Contract Info gaps the Bid Log/contacts did not provide.
# Precedence (DATA-SOURCES.md): Bid Log WINS. The subcontract only FILLS blanks or
# CONFIRMS. If the subcontract disagrees with a Bid-Log value, we keep the Bid-Log
# value and record the disagreement under "discrepancies" (we never silently overwrite).
# Every field here is tagged source "Subcontract" so the UI shows provenance.
SUBCONTRACT_FILE = (f"{SP_PROJECT_FOLDER}/02 - Project Management/"
                    "Subcontract Agreement/26-0331 - POET Subcontract Agmt FE.pdf")


# Crew One on POET (per the project timesheets / Crew 01 assignment). The POET
# contact log only carries John Willis (the crew lead); his crew works the job
# with him. Add the rest of Crew 1 as known PF Team members so the record
# reflects who is actually on site. Name + crew role only — we do NOT fabricate
# phone/email we don't have (left blank, rendered as em-dash by the view).
# Source: POET timesheets (Crew 01). John Willis already comes from the contact log.
POET_CREW_ONE = [
    {
        "scope": "Crew 1 — Operator",
        "company": "Pier Foundations, LLC",
        "name": "Seth Willis",
        "address": "",
        "phone": "",
        "cell": "",
        "email": "",
        "website": "www.pierfoundations.com",
        "notes": "Crew 1 (added from POET timesheets — Crew 01)",
    },
    {
        "scope": "Crew 1 — Operator",
        "company": "Pier Foundations, LLC",
        "name": "Jordan Lemay",
        "address": "",
        "phone": "",
        "cell": "",
        "email": "",
        "website": "www.pierfoundations.com",
        "notes": "Crew 1 (added from POET timesheets — Crew 01)",
    },
]


def add_crew_one(contacts):
    """Ensure POET's PF Team includes Crew 1 (John Willis + Seth Willis + Jordan
    Lemay). John Willis already comes from the contact log; append Seth and Jordan
    only if a member of the same name is not already present (idempotent)."""
    team = contacts.get("groups", {}).get("pf_team", [])
    existing = {(c.get("name") or "").strip().lower() for c in team}
    for crew in POET_CREW_ONE:
        if crew["name"].strip().lower() not in existing:
            team.append(dict(crew))
            existing.add(crew["name"].strip().lower())
    return contacts


def resolve_subcontract_item_id(token):
    """Resolve the Graph drive-item id for the POET fully-executed subcontract PDF.
    This id is baked into the record so the portal can embed the LIVE file inline
    via the /api/doc proxy (<iframe src="/api/doc?item=<id>">). Returns '' if not
    resolvable (the view then falls back to the SharePoint link)."""
    item = get_item_by_path(token, SUBCONTRACT_FILE)
    if item and item.get("id"):
        return item["id"]
    return ""


def parse_subcontract(token):
    """Download + parse the POET fully-executed subcontract PDF (AIA A142-2004
    Design-Builder/Contractor agreement) and return:
        {"fields": {field: value}, "snippets": {field: supporting_text},
         "source_file": ..., "pages": N, "scanned_pages": [...]}.
    Only fields GENUINELY present in the document are returned. No fabrication.
    Every returned field is rendered with source "Subcontract" by the view."""
    try:
        from pypdf import PdfReader
    except ImportError:
        return None
    try:
        raw = download_path(token, SUBCONTRACT_FILE)
    except Exception:
        return None

    reader = PdfReader(io.BytesIO(raw))
    pages = [(pg.extract_text() or "") for pg in reader.pages]
    full = "\n".join(pages)
    scanned = [i + 1 for i, t in enumerate(pages) if len(t.strip()) < 20]

    def grab(pattern, flags=re.S | re.I):
        m = re.search(pattern, full, flags)
        return m.group(1).strip() if m else ""

    fields = {}
    snippets = {}

    def setf(key, value, snippet):
        """Record a field only if a value was actually found in the doc."""
        if value:
            fields[key] = value
            snippets[key] = " ".join(snippet.split())[:400]

    # --- Subcontract / Contract Number (POET's contract id, from the executed Notice to Proceed) ---
    # The blank AIA template forms contain placeholder text ("PROJECT NUMBER: Project Code",
    # "CONTRACT NUMBER: SUBCONTRACT NUMBER"). The REAL filled values appear together in the
    # executed Notice to Proceed: "PROJECT NUMBER: SHB-03 ... CONTRACT NUMBER: SHB03E14".
    # Anchor on the block where both real ids appear adjacent (within ~80 chars) and require
    # the contract number to be a concrete code (letters+digits, no spaces/placeholder words).
    pnum = cnum = ""
    m_ids = re.search(r"PROJECT NUMBER:\s*(SHB-?\d+)\b.{0,120}?CONTRACT NUMBER:\s*([A-Z0-9]{5,})\b",
                      full, re.S | re.I)
    if m_ids:
        pnum, cnum = m_ids.group(1).strip(), m_ids.group(2).strip()
    if cnum and cnum.upper() not in ("SUBCONTRACT", "NUMBER"):
        val = cnum + (f" (POET Project No. {pnum})" if pnum else "")
        setf("subcontract_number", val,
             f"PROJECT NUMBER: {pnum}  CONTRACT NUMBER: {cnum} (executed Notice to Proceed)")

    # --- Fully Executed Contract Date (both DocuSign signatures dated 3/31/2026) ---
    # Agreement "made as of the 25th day of March 2026"; executed (signed) 3/31/2026.
    if re.search(r"3/31/2026", full):
        setf("fully_executed_date", "2026-03-31",
             "(Date) ... Brad Reinking Partner/Owner ... Docusign ... 3/31/2026 3/31/2026"
             " (both Design-Builder and Contractor signatures dated 3/31/2026)")
    agreement_date = grab(r"AGREEMENT made as of the\s+(\d+\w*\s+day of\s+\w+\s+in the year of\s+\d{4})")
    if agreement_date:
        setf("agreement_date", agreement_date,
             f"AGREEMENT made as of the {agreement_date}")

    # --- Project Address (from the agreement face) ---
    # "POET Bioprocessing - Shelbyville / 2373 West 300 North / Shelbyville, IN 46176"
    if re.search(r"2373 West 300 North", full):
        setf("project_address", "2373 West 300 North, Shelbyville, IN 46176",
             "POET Bioprocessing - Shelbyville  2373 West 300 North  Shelbyville, IN 46176")

    # --- GC / parties + address (Design-Builder = the GC here) ---
    if re.search(r"POET.{0,3}\s*Design\s*&\s*Construction", full):
        setf("gc_party", "POET Design & Construction, Inc., 4615 N Lewis Ave, Sioux Falls, SD 57104",
             "BETWEEN the Design-Builder: POET Design & Construction, Inc. 4615 N Lewis Ave"
             " Sioux Falls, SD 57104")
    if re.search(r"POET Holding Company", full):
        setf("owner_party", "POET Holding Company, LLC & Affiliated Entities, Sioux Falls, SD",
             "Owner: POET Holding Company, LLC & Affiliated Entities 4615 N. Lewis Ave. Sioux Falls, SD")

    # --- Subcontract Value (Stipulated Sum) ---
    sumv = grab(r"Stipulated Sum is[^$]*\(\$\s*([\d,]+\.\d{2})\s*\)")
    if sumv:
        setf("subcontract_value", "$" + sumv,
             f"The Stipulated Sum is Three Hundred Forty-Three Thousand Thirty-Seven and 07/100"
             f" Dollars (${sumv}), subject to additions and deductions")

    # --- Confirmed Scope ---
    if re.search(r"Aggregate Piers installation as detailed in bid package", full, re.I):
        setf("confirmed_scope", "Aggregate Piers installation (Grains/Fermentation areas) per bid package dated 2/4/2026",
             "Aggregate Piers installation as detailed in bid package dated 2/4/2026."
             " (Exhibit C, Contractor's Scope of Work)")

    # --- Contract Duration (calendar days) ---
    dur = grab(r"completed within\s+(\d+)\s+calendar days")
    if dur:
        setf("contract_duration_calendar", dur + " calendar days",
             f"entire scope of work shall be completed within {dur} calendar days")

    # --- Retainage % withheld (Stipulated-Sum progress payments) ---
    if re.search(r"less retainage of zero percent\s*\(0%\)", full, re.I):
        setf("retainage_pct", "0% (no retainage withheld)",
             "Take that portion of the Contract Sum ... less retainage of zero percent (0%) on the Work"
             " (Section 5.2.2.1, Stipulated Sum)")

    # --- Retainage Release ---
    if re.search(r"A\.?9\.8\.4.*release of applicable retainage", full, re.S | re.I) or \
       re.search(r"release of applicable retainage upon\s*Substantial Completion", full, re.S | re.I):
        setf("retainage_release",
             "Per Exhibit A Sec A.9.8.4 — release of applicable retainage upon Substantial Completion",
             "Section A.9.8.4 of Exhibit A, Terms and Conditions discusses release of applicable"
             " retainage upon Substantial Completion of Work.")

    # --- Payment Terms (net days, pay-when/if-paid) ---
    if re.search(r"not later than the 30th day of the following month", full, re.I):
        setf("payment_terms",
             "Net ~30 — App for Payment monthly; if received by month-end, paid by the 30th of the "
             "following month; otherwise within 30 days of receipt (Sec 5.1.2/5.1.3). No pay-if-paid "
             "clause found on the agreement face.",
             "The period covered by each Application for Payment will be one calendar month ... the"
             " Design-Builder will make payment to the Contractor not later than the 30th day of the"
             " following month ... otherwise not later than thirty (30) days after the Design-Builder"
             " receives the Application for Payment. (Sec 5.1.3)")

    # --- Liquidated Damages ---
    if re.search(r"\$500\.00\)?\s*per day", full, re.I) or \
       re.search(r"Five Hundred Dollars.*\(\$500\.00\)\s*per day", full, re.I):
        setf("liquidated_damages", "$500.00 per calendar day past Substantial Completion / each Milestone",
             "liquidated damages in the amount of Five Hundred Dollars and No Cents ($500.00) per day"
             " for each calendar day beyond the date of Substantial Completion or each calendar day"
             " beyond each Milestone date (Sec 3.2)")

    # --- Certified Payroll / Prevailing Wage ---
    if re.search(r"prevailing wage and apprenticeship requirements as detailed in Exhibit P", full, re.I):
        setf("prevailing_wage",
             "Yes — IRA prevailing wage + apprenticeship (Davis-Bacon-style rates per US Sec. of Labor; "
             "15% apprentice labor hours), IRC 45/48 enhanced-credit Labor Requirements (Exhibit P)",
             "compliance with the prevailing wage and apprenticeship requirements as detailed in"
             " Exhibit P. ... wages at rates not less than the prevailing rates ... in accordance with"
             " Subchapter IV of Chapter 31 of Title 40 ... not less than 15% of the total labor hours"
             " ... performed by qualified apprentices (Exhibit P)")
    if re.search(r"weekly certified payroll records.*LCP\s*Tracker", full, re.S | re.I):
        setf("certified_payroll",
             "Yes — weekly certified payroll records submitted through LCP Tracker (Exhibit P)",
             "submit weekly certified payroll records to Design Builder through LCP Tracker")

    # --- Project Working Hours ---
    if re.search(r"all Work at the site shall be performed during\s*regular working hours", full, re.I):
        setf("working_hours",
             "Regular working hours; no overtime, Saturday, Sunday, or legal-holiday work without "
             "Design-Builder's prior written consent (Sec A.3.4.4)",
             "all Work at the site shall be performed during regular working hours, and Contractor"
             " shall not permit overtime work or the performance of Work on Saturday, Sunday, or any"
             " legal holiday without Design-Builder's prior written consent. (Sec A.3.4.4)")

    # --- Surveying & Staking (who performs) ---
    if re.search(r"Contractor shall provide competent, suitably qualified personnel to survey and lay out the Work", full, re.I):
        setf("surveying_staking",
             "Contractor (PF) provides personnel to survey and lay out the Work (Sec A.3.4.1)",
             "Contractor shall provide competent, suitably qualified personnel to survey and lay out"
             " the Work, perform construction ... (Sec A.3.4.1)")

    # --- Tax exempt status (NOT definitively on the face) ---
    # Bid Form: price includes all sales/use taxes; "If applicable, an Excise Tax exemption certificate
    # will be provided with the contract"; "See Project Manual for Tax Exempt Certificates". The
    # agreement face does NOT definitively state PF's work is tax-exempt -> leave to Bid Log; note only.
    if re.search(r"See Project Manual for Tax Exempt Certificates", full, re.I):
        setf("tax_note",
             "Conditional — Bid Form: proposal includes all State/Municipal Sales & Use Taxes; an Excise "
             "Tax exemption certificate provided with the contract if applicable; tax-exempt certificates "
             "referenced in the Project Manual (not attached to this PDF). Not definitively stated on face.",
             "The Subcontractor acknowledges that the foregoing proposal includes all applicable State and"
             " Municipal Sales and Use Taxes on materials ... If applicable, an Excise Tax exemption"
             " certificate will be provided with the contract. (See Project Manual for Tax Exempt Certificates)")

    # --- Notice to Proceed / commencement date ---
    ntp = grab(r"Contract Times under the Contract will commence to run on:\s*([A-Za-z]+ \d+, \d{4})")
    if ntp:
        setf("commencement_date", ntp,
             f"You are notified that the Contract Times under the Contract will commence to run on: {ntp}")

    # NOTE: County and Township are NOT stated for the project site anywhere in the doc
    # (only "Shelbyville, IN 46176"; the only "County" references are Minnehaha County, SD =
    # the governing-law/venue, NOT the project location). Left blank — never inferred.

    return {
        "fields": fields,
        "snippets": snippets,
        "source_file": "26-0331 - POET Subcontract Agmt FE.pdf",
        "pages": len(pages),
        "scanned_pages": scanned,
    }


# ---------------- section document links ----------------
# Map each schema section to the POET subfolders/files we want it to link to.
# webUrl from Graph is an org SharePoint link (opens in browser for authenticated
# PF staff). We do NOT create anonymous share links (org policy disables sharing).
SECTION_LINKS = {
    "engineering": [
        ("Approved Shop Dwgs", f"{SP_PROJECT_FOLDER}/03 - Engineering & Design/Approved Shop Dwgs"),
        ("Garbin Prelim", f"{SP_PROJECT_FOLDER}/03 - Engineering & Design/Garbin Prelim"),
        ("Geotech Report", f"{SP_PROJECT_FOLDER}/03 - Engineering & Design/Geotech Report"),
        ("Modulus Test", f"{SP_PROJECT_FOLDER}/03 - Engineering & Design/Modulus Test"),
        ("Stamped Drawings", f"{SP_PROJECT_FOLDER}/03 - Engineering & Design/Stamped Drawings"),
        ("QAQC", f"{SP_PROJECT_FOLDER}/03 - Engineering & Design/QAQC"),
    ],
    "contract": [
        ("Subcontract Agreement", f"{SP_PROJECT_FOLDER}/02 - Project Management/Subcontract Agreement"),
    ],
    "financials": [
        ("Invoicing", f"{SP_PROJECT_FOLDER}/02 - Project Management/Invoicing"),
        ("Payroll", f"{SP_PROJECT_FOLDER}/02 - Project Management/Payroll"),
    ],
    "safety": [
        ("Site Specific Safety Plan (SSSP)", f"{SP_PROJECT_FOLDER}/SSSP-PIER-POET-2026.03.04-FINAL.pdf"),
        ("Field Safety Folder", f"{SP_PROJECT_FOLDER}/05 - Field/06 - Safety"),
    ],
    "site": [
        ("Field — Drawings", f"{SP_PROJECT_FOLDER}/05 - Field/03 - Drawings"),
        ("Field — Testing", f"{SP_PROJECT_FOLDER}/05 - Field/04 - Testing"),
        ("Field — Approved Materials", f"{SP_PROJECT_FOLDER}/05 - Field/05 - Approved Materials"),
        ("Field — Project Photos", f"{SP_PROJECT_FOLDER}/05 - Field/02 - Project Photos"),
    ],
    "general": [
        ("GC Drawings & Specs", f"{SP_PROJECT_FOLDER}/04 - GC Drawings & Specs"),
        ("Project Folder (root)", SP_PROJECT_FOLDER),
    ],
    "qaqc": [
        ("QAQC Logs (GUHMA)", f"{SP_PROJECT_FOLDER}/03 - Engineering & Design/QAQC"),
        ("Modulus Test", f"{SP_PROJECT_FOLDER}/03 - Engineering & Design/Modulus Test"),
    ],
    "material": [
        ("Field — Approved Materials", f"{SP_PROJECT_FOLDER}/05 - Field/05 - Approved Materials"),
        ("Vendor Quotes", f"{SP_PROJECT_FOLDER}/01 - Preconstruction/Vendor Quotes"),
    ],
}


def resolve_links(token):
    """Resolve each configured section link to its Graph webUrl. Returns
    {section: [ {label, url, found} ]}. Missing folders -> found:false, url:''."""
    out = {}
    for section, entries in SECTION_LINKS.items():
        resolved = []
        for label, path in entries:
            item = get_item_by_path(token, path)
            if item and item.get("webUrl"):
                resolved.append({"label": label, "url": item["webUrl"], "found": True})
            else:
                resolved.append({"label": label, "url": "", "found": False})
        out[section] = resolved
    return out


# ---------------- QA/QC from progress-data.js ----------------
def load_progress():
    """Extract window.PF_PROGRESS['26-002'] from the generated progress-data.js."""
    if not os.path.exists(PROGRESS_JS):
        return None
    with open(PROGRESS_JS) as f:
        txt = f.read()
    m = re.search(r"window\.PF_PROGRESS\s*=\s*(\{.*\});", txt, re.S)
    if not m:
        return None
    try:
        data = json.loads(m.group(1))
    except json.JSONDecodeError:
        return None
    return (data.get("projects") or {}).get(PROJECT_NUMBER)


# ---------------- assemble + write ----------------
def _norm_money(s):
    """Normalize a money string for NUMERIC comparison: strip $, commas, spaces, then
    compare by value so '398500' and '$398,500.00' are recognized as EQUAL (same
    dollar amount, different formatting) and do NOT produce a false discrepancy."""
    digits = re.sub(r"[^\d.]", "", str(s or ""))
    if not digits:
        return ""
    try:
        # canonical numeric string: drops trailing .00 / formatting differences
        return repr(float(digits))
    except ValueError:
        return digits


def detect_discrepancies(bid, sub):
    """Compare subcontract-sourced values against Bid-Log values that the Bid Log
    ALREADY set. Bid Log wins; we only NOTE disagreements (never overwrite).
    Returns a list of {field, bid_log, subcontract} dicts."""
    out = []
    if not bid or not sub:
        return out
    sf = sub.get("fields", {})

    # Subcontract value vs Bid Log Bid Total Value
    if sf.get("subcontract_value") and bid.get("bid_total_value"):
        if _norm_money(sf["subcontract_value"]) != _norm_money(bid["bid_total_value"]):
            out.append({"field": "Contract / Bid Value",
                        "bid_log": "$" + str(bid["bid_total_value"]),
                        "subcontract": sf["subcontract_value"]})

    # Duration: Bid Log holds WORKING days; subcontract holds CALENDAR days.
    # These are different units, so a numeric mismatch is EXPECTED, not a conflict.
    # We surface it as an informational note rather than a discrepancy.
    return out


def assemble(token):
    contacts = parse_contacts(token)
    contacts = add_crew_one(contacts)  # ensure Crew 1 (John + Seth + Jordan) on PF Team
    links = resolve_links(token)
    progress = load_progress()
    bid = parse_bid_log(token)
    sub = parse_subcontract(token)
    discrepancies = detect_discrepancies(bid, sub)

    # Bake the live subcontract drive-item id into the record so the portal can
    # embed the file inline via the /api/doc proxy (always current, not a copy).
    if sub is not None:
        sub["item_id"] = resolve_subcontract_item_id(token)

    record = {
        "project_number": PROJECT_NUMBER,
        "project_name": PROJECT_NAME,
        "location": PROJECT_LOCATION,
        "sp_folder": SP_PROJECT_FOLDER,
        "generated": datetime.now(timezone.utc).isoformat() + "Z",
        "data_note": ("READ-ONLY v1. Populated from SharePoint: the POET Project Info contacts "
                      "file (full contact directory), the Project Bid Log (General Info metrics, "
                      "Contract Info dates, award value, GC/engineer), the fully-executed POET "
                      "subcontract (executed date, contract no., retainage, payment terms, LDs, "
                      "prevailing wage, working hours, surveying — fills Bid-Log gaps only), folder "
                      "doc links, and the auto-progress engine (QA/QC installed qty). Editing/saving "
                      "is v2. Bid Log wins on shared fields; subcontract only fills blanks or confirms. "
                      "Fields with no confirmed source render blank — never fabricated."),
        "contacts": contacts,
        "links": links,
        "qaqc": progress,
        "bid_log": bid,
        "subcontract": sub,
        "discrepancies": discrepancies,
    }
    return record


def write_js(record):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(OUT_JS, "w") as f:
        f.write("// AUTO-GENERATED by sync/build-project-record.py — do not edit by hand.\n")
        f.write("// POET (26-002) project record — READ-ONLY v1 (saving is v2).\n")
        f.write("// Sources: POET Project Info contacts + Project Bid Log (metrics/dates/award) +\n")
        f.write("//          POET folder doc links + progress-data.js (QA/QC).\n")
        f.write("window.PF_PROJECT_POET = ")
        json.dump(record, f, indent=2)
        f.write(";\n")
    return OUT_JS


def _summary(record):
    g = record["contacts"]["groups"]
    print("POET project record assembled:")
    for k in ("owner", "gc", "engineering", "pf_team", "vendors"):
        print(f"  contacts.{k}: {len(g[k])}")
        for c in g[k]:
            who = c["name"] or "(no name)"
            print(f"      - {c['scope']}: {c['company']} / {who} / "
                  f"{c['email'] or '(no email)'} / {c['cell'] or c['phone'] or '(no phone)'}")
    print("  section links:")
    for section, lst in record["links"].items():
        found = sum(1 for x in lst if x["found"])
        print(f"      {section}: {found}/{len(lst)} resolved")
        for x in lst:
            mark = "OK " if x["found"] else "MISS"
            print(f"        [{mark}] {x['label']}")
    q = record["qaqc"]
    if q:
        print(f"  QA/QC (26-002): installed_columns={q.get('installed_columns')} "
              f"installed_lf={q.get('installed_lf')} baseline={q.get('baseline_status')} "
              f"last_log={q.get('last_log_date')}")
    else:
        print("  QA/QC: no progress entry found")
    b = record["bid_log"]
    if b:
        print(f"  Bid Log (row {b['row']}): status={b['bid_status']} value={b['bid_total_value']} "
              f"LF={b['total_lf']} cols={b['total_columns']} stone_tn={b['total_stone_tn']} "
              f"feed={b['feed_type']} start={b['projected_start']} GC={b['gc_name']} "
              f"engineer={b['engineer_firm']}")
    else:
        print("  Bid Log: POET row not found")
    s = record.get("subcontract")
    if s:
        print(f"  Subcontract ({s['source_file']}): {s['pages']} pages, "
              f"scanned={s['scanned_pages'] or 'none'}")
        for k, v in s["fields"].items():
            print(f"      [SUB] {k}: {v}")
    else:
        print("  Subcontract: not parsed (missing pypdf or file)")
    d = record.get("discrepancies") or []
    if d:
        print("  Discrepancies (Bid Log kept, subcontract noted):")
        for x in d:
            print(f"      ! {x['field']}: bidlog={x['bid_log']} vs subcontract={x['subcontract']}")
    else:
        print("  Discrepancies: none")


# =====================================================================
# PER-PROJECT DEEP SUBCONTRACT EXTRACTION  (--project NN-NNN)
# =====================================================================
# POET (above) was the FIRST deep record and lives in its own file
# (project-record-poet.js). Every OTHER awarded project lives in the
# bulk keyed file data/project-records.js (window.PF_PROJECT_RECORDS),
# built by sync/build-project-records.py with subcontract=null.
#
# This section deepens a bulk record to POET level: it extracts the SAME
# field set we extracted for POET (verbatim PDF snippets, NO fabrication,
# blanks where absent) from that project's executed subcontract, resolves
# the Graph drive-item id so the portal can embed the live file inline via
# /api/doc, and MERGES the result into the existing bulk entry — preserving
# every other field of that record AND every other record untouched.
#
# Field-extraction regexes are inherently document-specific (each GC uses a
# different contract form: POET = AIA A142; Patterson Horth = its own
# Standard Terms form). So each project gets its own extractor function,
# but ALL of them reuse the shared Graph helpers, the setf()/snippet
# discipline, the item-id resolver pattern, and the bulk-merge writer.

BULK_JS = os.path.join(DATA_DIR, "project-records.js")

# ---- per-project config: where each deep-project's subcontract lives ----
# NOTE the SharePoint folder spells the company "Schaff" (double-f) while the
# PDF filename spells it "Schaaf" — both confirmed live via Graph 2026-06-18.
PROJECT_DEEP = {
    "26-015": {
        "name": "Schaaf CPA Group",
        "location": "Westfield, IN",
        "sp_folder": ("04 - Project Management/02 - Projects/"
                      "26-015 - Schaff CPA - Patterson Horth"),
        "subcontract_path": ("04 - Project Management/02 - Projects/"
                             "26-015 - Schaff CPA - Patterson Horth/"
                             "02 - Project Management/Subcontract Agreement/"
                             "26-0617 - Schaaf CPA Subcontract Agmt rev.pdf"),
        "subcontract_file": "26-0617 - Schaaf CPA Subcontract Agmt rev.pdf",
        "extractor": "schaaf",
        # page-1 fingerprint we REQUIRE before trusting the doc (we've been
        # burned by mislabeled subcontracts) — must contain the owner + GC.
        "verify_tokens": ["Schaaf CPA Group", "Patterson Horth"],
    },
    "26-013": {
        "name": "Park & Poplar (Garbin)",
        "location": "Westfield, IN",
        "sp_folder": ("04 - Project Management/02 - Projects/"
                      "26-013 - Park & Poplar - OldTown"),
        # NOTE: this subcontract is a DRAFT (not executed). The SharePoint copy is
        # byte-identical to the local /tmp DRAFT (size 1,089,789). It uses Old Town
        # Construction's own "Subcontractor Agreement" form (§-numbered body with a
        # page-1 PROJECT/SUBCONTRACT-SUM/CONTRACTOR/SUBCONTRACTOR/OWNER block) — a
        # different form than POET (AIA A142) or Schaaf (Patterson Horth header).
        "subcontract_path": ("04 - Project Management/02 - Projects/"
                             "26-013 - Park & Poplar - OldTown/"
                             "02 - Project Management/Subcontract Agreement/"
                             "Pier Foundations - Park and Poplar Subcontract (6-11-26) DRAFT.pdf"),
        "subcontract_file": "Pier Foundations - Park and Poplar Subcontract (6-11-26) DRAFT.pdf",
        "extractor": "pp",
        # page-1 fingerprint REQUIRED before trusting the doc — owner + GC + project.
        "verify_tokens": ["Park & Poplar", "Old Town Construction"],
        # DRAFT (not executed): no signatures, blank Subcontractor reps. Carried on
        # the merged record so the view shows DRAFT status (matches analysis block).
        "execution_status": "DRAFT — needs signature",
    },
}


def _verify_subcontract(full_text, tokens):
    """Confirm the PDF is the subcontract we think it is before extracting.
    Returns (ok, missing_tokens). Matches on whitespace-collapsed text so a
    multi-word token still matches across pypdf line wraps."""
    flat = _collapse_ws(full_text).lower()
    missing = [t for t in tokens if _collapse_ws(t).lower() not in flat]
    return (len(missing) == 0, missing)


def _collapse_ws(s):
    """Collapse ALL runs of whitespace (incl. newlines from PDF line breaks)
    into single spaces, so multi-word phrases match regardless of how pypdf
    wrapped the lines."""
    return re.sub(r"\s+", " ", s)


def parse_subcontract_schaaf(full):
    """Extract the Schaaf CPA / Patterson Horth subcontract field set from the
    already-loaded full PDF text. Same return shape + snippet discipline as the
    POET parse_subcontract(). Patterson Horth uses a clean page-1 header block
    (PROJECT NO / DATE / CONTRACTOR / SUBCONTRACTOR / OWNER / RETAINAGE /
    SUBCONTRACT SUM / dates) plus a Standard Terms body. NO fabrication; a
    field is only set when a value is genuinely found in the document."""
    # Collapse pypdf line wraps so multi-word phrases (and the page-1 header
    # block) match on a single normalized string. Hand-written snippets below
    # are literals, so they stay clean regardless of how the source wrapped.
    flat = _collapse_ws(full)

    fields = {}
    snippets = {}

    def setf(key, value, snippet):
        if value:
            fields[key] = value
            snippets[key] = " ".join(snippet.split())[:400]

    def grab(pattern, flags=re.S | re.I):
        m = re.search(pattern, flat, flags)
        return m.group(1).strip() if m else ""

    # --- Subcontract / Project number (page-1 header: "PROJECT NO.:22621SC03") ---
    pno = grab(r"PROJECT NO\s*\.?\s*:\s*([A-Z0-9\-]+)")
    if pno:
        setf("subcontract_number", pno,
             f"PROJECT NO.: {pno} (Patterson Horth subcontract header, page 1)")

    # --- Agreement / contract date (page-1 header: "DATE : 06/15/2026") ---
    cdate = grab(r"\bDA\s*TE\s*:\s*(\d{1,2}/\d{1,2}/\d{4})")
    if cdate:
        setf("agreement_date", _norm_date(cdate),
             f"DATE: {cdate} (subcontract header, page 1)")

    # --- Commencement date (page-1 header: "COMMENCEMENT DATE: 06/15/2026") ---
    comm = grab(r"COMMENCEMENT\s*DA\s*TE\s*:\s*(\d{1,2}/\d{1,2}/\d{4})")
    if comm:
        setf("commencement_date", _norm_date(comm),
             f"COMMENCEMENT DATE: {comm} (subcontract header, page 1)")

    # --- Substantial & final completion (page-1 header) ---
    subc = grab(r"SUBST ?ANTIAL\s*COMPLETION\s*DA\s*TE\s*:\s*(\d{1,2}/\d{1,2}/\d{4})")
    finc = grab(r"FINAL\s*COMPLETION\s*DA\s*TE\s*:\s*(\d{1,2}/\d{1,2}/\d{4})")
    if subc:
        comp = "Substantial " + subc + (" / Final " + finc if finc else "")
        setf("completion_dates", comp,
             f"SUBSTANTIAL COMPLETION DATE: {subc}  FINAL COMPLETION DATE: {finc} (page 1)")

    # --- Parties (CONTRACTOR = GC; OWNER) ---
    if re.search(r"Patterson Horth, ?Inc\.", flat, re.I):
        setf("gc_party",
             "Patterson Horth, Inc. (Contractor/GC), 5745 Progress Rd, Indianapolis, IN 46241",
             "CONTRACTOR: Patterson Horth, Inc. 5745 Progress Rd Indianapolis, Indiana 46241 (page 1)")
    if re.search(r"Schaaf CPA Group", flat, re.I):
        setf("owner_party", "Schaaf CPA Group, LLC (Owner)",
             "OWNER: Schaaf CPA Group, LLC  PROJECT: Schaaf CPA Group (page 1)")

    # --- Project address (page-1 header) ---
    if re.search(r"250 N\.? Union Street", flat, re.I):
        setf("project_address", "250 N. Union Street, Westfield, IN 46074",
             "PROJECT LOCATION: 250 N. Union Street, Westfield, Indiana 46074 (page 1)")

    # --- Subcontract value (page-1 header: "SUBCONTRACT SUM: $68,200.00") ---
    val = grab(r"SUBCONTRACT\s*SUM\s*:\s*\$?\s*([\d,]+\.\d{2})")
    if val:
        setf("subcontract_value", "$" + val,
             f"SUBCONTRACT SUM: ${val} (page 1)")

    # --- Confirmed scope (page-1 SCOPE OF WORK) ---
    if re.search(r"BP\s*2\s*Rammed Aggregate Pier", flat, re.I):
        setf("confirmed_scope",
             "BP 2 — Rammed Aggregate Pier work (all labor, material, taxes, insurance, "
             "supervision, equipment) per the Contract Documents",
             "Provide all Labor, Material, Taxes, Insurance, Supervision, Equipment and all other "
             "necessary requirements to complete the BP 2 Rammed Aggregate Pier work per the "
             "Contract Documents. (SCOPE OF WORK, page 1)")

    # --- Retainage % (page-1 header: "RETAINAGE: 5.0%") ---
    ret = grab(r"RET\s*AINAGE\s*:\s*([\d.]+\s*%)")
    if ret:
        setf("retainage_pct", ret.replace(" ", "") + " withheld on progress payments",
             f"RETAINAGE: {ret} (page 1); 'Net costs to be defined as the cost of all field labor, "
             "materials, tools, equipment, insurance and taxes.' (Sec 9, Standard Terms)")

    # --- Retainage release / final payment (Standard Terms: up to 100% on final) ---
    if re.search(r"final payme\s*nt,?\s*not to exceed one hundred percent", flat, re.I):
        setf("retainage_release",
             "Final application up to 100% of the Subcontract amount after Subcontractor completes "
             "its obligations to the full satisfaction of Contractor, Owner and Architect "
             "(retainage released at final payment) — Standard Terms",
             "Subcontract may make an application for final payment, not to exceed one hundred "
             "percent (100%) of the Subcontract amount, after Subcontractor completes its "
             "obligations hereunder to the full satisfaction of Contractor, Owner and Architect.")

    # --- Payment terms (PAY-IF-PAID: condition precedent + 10 days after GC paid) ---
    if re.search(r"condition precedent to Subcontractor", flat, re.I):
        setf("payment_terms",
             "PAY-IF-PAID — Owner's payment to Contractor is an express CONDITION PRECEDENT to "
             "Subcontractor's right to progress payments. When the Owner pays, Subcontractor's "
             "properly submitted Applications are paid 10 days after Contractor receives the "
             "corresponding payment from the Owner. (Standard Terms, Sec on Payments)",
             "payments to Contractor by the Owner ... for periodic progress payments are a condition "
             "precedent to Subcontractor's right to periodic progress payments. Where corresponding "
             "payments by the Owner are made, Subcontractor's properly submitted Applications for "
             "Payment will be paid 10 days after Contractor receives the corresponding payment from owner.")

    # --- Liquidated damages (page-1 header says n/a; Standard Terms flow-down clause) ---
    ld_hdr = grab(r"LIQUIDA\s*TED\s*DAMAGES\s*:\s*([^\n]+?)\s*(?:RET|$)")
    if ld_hdr:
        ld_hdr_clean = ld_hdr.strip()
        flow = bool(re.search(r"same liquidated damages provisions under the Contract which binds Contractor",
                              flat, re.I))
        val_ld = (f"None scheduled (header: \"{ld_hdr_clean}\")"
                  + (" — but Subcontractor is bound to the SAME liquidated-damages provisions of the "
                     "Prime Contract that bind Contractor to the Owner, if any (flow-down, Sec 8)."
                     if flow else ""))
        setf("liquidated_damages", val_ld,
             f"LIQUIDATED DAMAGES: {ld_hdr_clean} (page 1). "
             "Subcontractor shall be bound to Contractor by the same liquidated damages provisions "
             "under the Contract which binds Contractor to the Owner ('Prime Contract'), if any. (Sec 8)")

    # --- Personal guaranty (counsel-flagged) ---
    if re.search(r"This Guaranty is an absolute, unconditional and continuing guaranty", flat, re.I):
        setf("personal_guaranty",
             "PRESENT — absolute, unconditional and continuing personal/entity guaranty of full and "
             "punctual performance of all Obligations; enforceable without first pursuing the "
             "Subcontractor or any security. (Guaranty section) — COUNSEL FLAGGED",
             "This Guaranty is an absolute, unconditional and continuing guaranty of the full and "
             "punctual performance by the Subcontractor of the Obligations and not of collectability "
             "only. Enforcement ... is in no way conditioned upon any requirement that the Owner or "
             "Patterson Horth, Inc. first attempt to collect or take any action against the Subcontractor.")

    # --- Certified payroll / labor standards ---
    if re.search(r"Certified Payroll Report.*submit weekly", flat, re.S | re.I):
        setf("certified_payroll",
             "Yes — Certified Payroll Report submitted weekly via email per 'Contract Labor Standards "
             "Requirements'; General Sales Tax Exemption Certificate also required.",
             "Certified Payroll Report and Instructions for Completing (complete a Certified Payroll "
             "Report and submit weekly via email - see 'Contract Labor Standards Requirements')")
    if re.search(r"Contract Labor Standards Requirements", flat, re.I):
        setf("prevailing_wage",
             "Labor Standards apply — 'Contract Labor Standards Requirements' referenced with weekly "
             "certified payroll; specific prevailing-wage rate not stated on the agreement face.",
             "complete a Certified Payroll Report and submit weekly via email - see 'Contract Labor "
             "Standards Requirements'")

    # --- Insurance requirements (Subcontract Insurance Requirements exhibit) ---
    if re.search(r"Commercial General Liability Limits", flat, re.I):
        setf("insurance_requirements",
             "CGL $1,000,000 each occurrence / $2,000,000 aggregate (per project); Auto $1,000,000 "
             "CSL (owned/hired/non-owned); Worker's Comp / Employer's Liability $500,000 each "
             "accident; EIFS/EFIS $1,000,000 min; $10,000 max deductible; written occurrence-basis, "
             "maintained until final payment. (Subcontract Insurance Requirements)",
             "Commercial General Liability Limits Each Occurrence $1,000,000 General Aggregate "
             "$2,000,000 ... General Aggregate Limit to apply per project ... Commercial Automobile "
             "Liability Limits Each accident $1,000,000 ... Worker's Comp/Employer's Liability Limits "
             "Each Accident $500,000 ... $10,000 maximum deductible.")

    # --- Working hours / jobsite policy ---
    if re.search(r"normal working hours and only with approval of the superintendent", flat, re.I):
        setf("working_hours",
             "Normal working hours only; no one onsite at any other time without superintendent "
             "approval. No radios/headphones/earbuds in the construction area; daily cleanup; weekly "
             "toolbox talk + daily JSA; English-speaking authorized representative required. "
             "(Standard Terms / jobsite policy)",
             "No employee of the subcontractor's or vendor's are to be on the construction site at any "
             "time other than normal working hours and only with approval of the superintendent. ... "
             "No radios shall be played during working hours in the construction area. This includes "
             "all headphones or earbuds.")

    # --- Surveying & staking / layout ---
    if re.search(r"Subcontractor shall maintain at the project a sup", flat, re.I):
        setf("surveying_staking",
             "Subcontractor must maintain a superintendent/foreman or such representative onsite "
             "authorized to act for it; no project signage without Contractor's written approval. "
             "(Sec 10, Subcontractor's On-site Presence). No separate VSC-layout add line appears in "
             "this rev of the agreement.",
             "Subcontractor shall maintain at the project a superintendent, foreman or other such "
             "representative and such individual ... is hereby authorized to make agreements for or "
             "otherwise act on behalf of Subcontractor. (Sec 10)")

    # --- Tax (page-1 scope includes Taxes; sales-tax-exemption cert required) ---
    if re.search(r"General Sales Tax Exemption Certificate", flat, re.I):
        setf("tax_note",
             "Scope price INCLUDES all Taxes (page-1 Scope of Work). A General Sales Tax Exemption "
             "Certificate is a required Additional Document (Exhibit A). Tax-exempt status not "
             "definitively stated on the face.",
             "Provide all Labor, Material, Taxes, Insurance ... (Scope of Work, page 1). "
             "General Sales Tax Exemption Certificate (Additional Documents, Exhibit A)")

    return {
        "fields": fields,
        "snippets": snippets,
    }


def parse_subcontract_pp(full):
    """Extract the Park & Poplar / Old Town Construction subcontract field set from
    the already-loaded full PDF text. Same return shape + snippet discipline as the
    POET/Schaaf parsers. This is Old Town Construction's own "Subcontractor Agreement"
    form: a page-1 block (PROJECT DATE / SUBCONTRACT SUM / CONTRACTOR / SUBCONTRACTOR /
    OWNER / ARCHITECT) plus a §-numbered body. THIS DOCUMENT IS A DRAFT (unsigned,
    blank signature lines, blank Subcontractor reps). NO fabrication; a field is only
    set when a value is genuinely found in the document."""
    # Collapse pypdf line wraps so multi-word phrases (and the page-1 block) match on
    # a single normalized string. Hand-written snippets below are literals.
    flat = _collapse_ws(full)

    fields = {}
    snippets = {}

    def setf(key, value, snippet):
        if value:
            fields[key] = value
            snippets[key] = " ".join(snippet.split())[:400]

    def grab(pattern, flags=re.S | re.I):
        m = re.search(pattern, flat, flags)
        return m.group(1).strip() if m else ""

    # --- Subcontract / Project number ---
    # This DRAFT has NO assigned subcontract/contract number (the analysis block
    # records "DRAFT (not yet numbered)"). The page-1 block carries PROJECT + DATE +
    # SUBCONTRACT SUM but no contract number field. Leave blank — never fabricate.

    # --- Contract / agreement date (page-1 block: "PROJECT DATE: 6/11/2026") ---
    cdate = grab(r"PROJECT\s*DATE\s*:\s*(\d{1,2}/\d{1,2}/\d{4})")
    if cdate:
        setf("agreement_date", _norm_date(cdate),
             f"PROJECT  DATE: {cdate} (page-1 block). Draft footer: 'Last updated 5/27/2026 (BAR)'.")

    # --- DRAFT status note (no signatures; unsigned) ---
    # Signature page (pg 19) has blank "(Signature) ... Date: Date:" lines and §1.2
    # Subcontractor Representative fields are blank — this is an unsigned DRAFT.
    if re.search(r"This Subcontract is effective as of the day and year first written above", flat, re.I):
        setf("fully_executed_date",
             "Not executed — DRAFT (signature page blank; §1.2 Subcontractor reps blank)",
             "This Subcontract is effective as of the day and year first written above. "
             "CONTRACTOR: SUBCONTRACTOR: (Signature) (Signature) (Printed name and title) "
             "(Printed name and title) Date: Date: (signature page left blank — unsigned DRAFT)")

    # --- Parties (CONTRACTOR = GC; SUBCONTRACTOR = PF; OWNER; ARCHITECT) ---
    if re.search(r"Old Town Construction", flat, re.I):
        setf("gc_party",
             "Old Town Construction, LLC (Contractor/GC), 525 North End Dr., Suite 100, "
             "Carmel, IN 46032 — Rep: Patrick Groogan / VP Construction Joel Scheele",
             "CONTRACTOR Name: Old Town Construction Street: 525 North End Dr., Suite 100 "
             "City/State/Zip: Carmel, IN, 46032. Contractor's Project Representative(s) are "
             "Vice President of Construction Joel Scheele and: Name: Patrick Groogan (§1.1, page 1)")
    if re.search(r"Park & Poplar Residential", flat, re.I):
        setf("owner_party",
             "Park & Poplar Residential LLC (Owner), 525 North End Drive, Suite 100, "
             "Carmel, IN 46032 — AFFILIATE of Contractor (§1.4)",
             "OWNER Name: Park & Poplar Residential LLC Street: 525 North End Drive, Suite 100 "
             "Carmel, IN 46032. §1.4 Owner Affiliation: Contractor and Owner may be affiliated "
             "entities ... Subcontractor waives any claim or defense based upon the relationship.")

    # --- Project address (page-1 block) ---
    if re.search(r"Westfield Blvd\s*/\s*Park Street", flat, re.I):
        setf("project_address", "Westfield Blvd / Park Street, Westfield, IN",
             "Project Name: Park & Poplar  Project Address: Westfield Blvd/Park Street, "
             "Westfield IN (page-1 block)")

    # --- Subcontract value (page-1 block: "SUBCONTRACT SUM: $398,500.00") ---
    val = grab(r"SUBCONTRACT\s*SUM\s*:\s*\$?\s*([\d,]+\.\d{2})")
    if val:
        setf("subcontract_value", "$" + val,
             f"SUBCONTRACT SUM: ${val} (page-1 block). 'The Subcontract Sum is a fixed, firm, "
             "and all-inclusive price ... includes all applicable federal, state, county, "
             "municipal and local taxes ... All price escalation is included.'")

    # --- Confirmed scope (Exhibit B Scope of Work — VSC + Helical Piers) ---
    if re.search(r"Agg and Helical Pier scopes", flat, re.I):
        setf("confirmed_scope",
             "Vibratory Stone Columns (VSC) + Galvanized Helical Piers — full design-build "
             "scope: submittals signed/sealed by licensed engineer, VSC & helical layout, "
             "furnish & install at design locations, move spoils outside building footprint, "
             "verification testing. All-inclusive of labor, material, taxes, insurance, "
             "supervision, equipment (Exhibit B Scope of Work).",
             "complete professional execution of Agg and Helical Pier scopes ... 3. Provide VSC "
             "and Helical Pier Layout ... 4. Furnish and install Vibratory Stone Columns (VSC) at "
             "design locations. 5. Furnish and install Galvanized Helical Piers at design "
             "locations. 6. Spoils will be moved outside of building footprint by this "
             "subcontractor. 7. Provide verification testing as required (Exhibit B)")

    # --- Retainage % (§4.4 progress payments: 'less retainage of ten percent (10%)') ---
    if re.search(r"less retainage of ten percent\s*\(10%\)", flat, re.I):
        setf("retainage_pct",
             "10% withheld on completed Work AND on stored materials (§4.4)",
             "Take that portion of the Subcontract Sum properly allocable to completed Work ... "
             "less retainage of ten percent (10%); (b) ... materials and equipment delivered and "
             "suitably stored ... less retainage of ten percent (10%) (§4.4)")

    # --- Retainage release / final payment conditions (§4.8, pay-when-paid) ---
    if re.search(r"Contractor receives final payment from Owner for Subcontractor", flat, re.I):
        setf("retainage_release",
             "Final payment (which releases retainage) is contingent on, among other conditions, "
             "Contractor receiving final payment from Owner for the Subcontractor's Work (§4.8(d)) "
             "— Owner is Contractor's affiliate. Retainage may be held until final payment.",
             "(d) Contractor receives final payment from Owner for Subcontractor's Work. (§4.8 "
             "final payment conditions — preceded by close-out submissions: warranties, O&M "
             "manuals, As-Builts, keys and other close-out submissions)")

    # --- Payment terms (PAY-WHEN-PAID: progress payment 10 days after GC paid by Owner) ---
    if re.search(r"within ten\s*\(10\)\s*days of Contractor.?s receipt of payment from the Owner", flat, re.I):
        setf("payment_terms",
             "PAY-WHEN-PAID — Application for Payment due by the 15th of each month (§4.3); "
             "Contractor makes progress payments within 10 days of Contractor's RECEIPT of "
             "payment from the Owner (§4.4). Final payment expressly conditioned on Contractor "
             "receiving final Owner payment (§4.8(d)). Owner is Contractor's affiliate (§1.4). "
             "If undisputed amounts unpaid 30 days past due, Subcontractor may stop work on 7 "
             "days' notice (§4.9).",
             "On or before the 15th day of each month ('Invoice Due Date'), the Subcontractor "
             "shall ... submit ... an Application for Payment (§4.3). ... Contractor shall make "
             "progress payments to the Subcontractor within ten (10) days of Contractor's receipt "
             "of payment from the Owner. (§4.4)")

    # --- Liquidated damages (proportional flow-down + Contractor's own actual delay damages) ---
    if re.search(r"Contractor may assess liquidated damages against the Subcontractor in proportion", flat, re.I):
        setf("liquidated_damages",
             "Proportional flow-down — if the Prime Contract provides LDs against Contractor and "
             "they are assessed, Contractor may assess LDs against Subcontractor in proportion to "
             "its share of delay responsibility (capped at the amount assessed against "
             "Contractor). PLUS Contractor may separately assess its OWN actual delay damages "
             "(extended general conditions, acceleration costs, other actual damages) — these are "
             "uncapped. (LDs / Delay sections). COUNSEL-FLAGGED: ask for an aggregate cap.",
             "If the Prime Contract provides for liquidated damages against Contractor, and such "
             "damages are assessed, Contractor may assess liquidated damages against the "
             "Subcontractor in proportion to the Subcontractor's share of responsibility for the "
             "delay, not to exceed the amount assessed against Contractor. ... Contractor may "
             "suffer its own damages ... including without limitation, extended general "
             "conditions, acceleration costs, and other actual damages ... assessed to "
             "Subcontractor.")

    # --- Bonds (§3.5: payment & performance bonds NOT required — 'No [X]') ---
    if re.search(r"required to provide payment and performance bonds\?.{0,60}No\s*\[\s*_*X", flat, re.I):
        setf("bonds_required",
             "No — payment & performance bonds NOT required (§3.5 'No [X]' checked).",
             "§ 3.5 Bonds: Is Subcontractor required to provide payment and performance bonds? "
             "[Check One] Yes [____] No [__X__] (bonds not required)")

    # --- Insurance requirements (§7/§9/§14.9 stack) ---
    if re.search(r"Comprehensive General Liability insurance written on an ISO occurrence form", flat, re.I):
        setf("insurance_requirements",
             "Heavier than PF baseline. CGL (ISO occurrence): $2,000,000 BI/PD per occurrence, "
             "$5,000,000 general aggregate PER PROJECT (CG 25 03 endorsement), $2,000,000 "
             "products & completed operations aggregate. Employer's Liability $1,000,000 each "
             "accident / $1,000,000 disease. Excess/Umbrella $5,000,000. Contractors Pollution "
             "Liability $1,000,000/occurrence & $2,000,000 aggregate (triggered by "
             "excavation/earthwork; 3-yr tail if claims-made). Professional Liability $2,000,000 "
             "per claim / $4,000,000 aggregate (structural/design-build) or $1,000,000 / "
             "$2,000,000 for other trades (§14.9). Best A- or better; additional-insured, "
             "primary & noncontributory, waiver of subrogation; deductibles/SIRs borne by "
             "Subcontractor.",
             "Comprehensive General Liability insurance written on an ISO occurrence form ... "
             "$2,000,000 Bodily Injury/Personal Injury per person and Property Damage per "
             "occurrence, and $5,000,000 general aggregate per project and $2,000,000 products "
             "and completed operations aggregate. ... Excess/umbrella liability ... at least "
             "$5,000,000. ... Contractors Pollution Liability ... $1,000,000 per occurrence and "
             "$2,000,000 aggregate. ... professional liability coverage with limits of not less "
             "than $2,000,000 per claim and $4,000,000 in the aggregate (§14.9)")

    # --- Certified payroll / labor standards (E-Verify; no prevailing-wage/cert-payroll clause) ---
    if re.search(r"E-\s*Verify program", flat, re.I):
        setf("certified_payroll",
             "No certified-payroll requirement found on the agreement face. Labor compliance is "
             "E-Verify + Form I-9 employment-eligibility verification (records produced within 24 "
             "hours of request), flowed down to lower-tier subs.",
             "Subcontractor shall verify the employment eligibility of all workers through proper "
             "completion and retention of Form I-9s and participation in the E-Verify program, "
             "and shall make such records available for inspection within 24 hours of request by "
             "Contractor.")
    # No prevailing-wage clause appears in this DRAFT — leave SF.prevailing_wage blank.

    # --- Project working hours / superintendent presence (§3.7) ---
    if re.search(r"English-speaking superintendent, foreman, or other site supervisor who shall be onsite", flat, re.I):
        setf("working_hours",
             "No fixed clock-hours clause on the face. Subcontractor must furnish (in writing, "
             "for Contractor approval) an English-speaking superintendent/foreman/site supervisor "
             "ONSITE whenever Subcontractor is performing Work; rep must attend all jobsite "
             "safety, production and coordination meetings starting two weeks before commencement "
             "through final completion (§3.7).",
             "§ 3.7 Superintendent: The Subcontractor shall ... furnish in writing to Contractor "
             "the name and qualifications of a proposed English-speaking superintendent, foreman, "
             "or other site supervisor who shall be onsite whenever Subcontractor is performing "
             "Work.")

    # --- Surveying & staking / layout (Exhibit B — Subcontractor provides VSC/Helical layout) ---
    if re.search(r"Provide VSC and Helical Pier Layout for the Scope of Work", flat, re.I):
        setf("surveying_staking",
             "Subcontractor (PF) provides VSC and Helical Pier layout for its scope (Exhibit B "
             "item 3). Subcontractor must also maintain an authorized onsite superintendent/"
             "foreman (§3.7).",
             "3. Provide VSC and Helical Pier Layout for the Scope of Work. (Exhibit B, Scope of "
             "Work)")

    # --- Tax (all-inclusive lump sum includes all taxes; no tax-exempt clause) ---
    if re.search(r"includes all applicable federal, state, county, municipal and local taxes", flat, re.I):
        setf("tax_note",
             "Subcontract Sum is all-inclusive and INCLUDES all applicable federal, state, "
             "county, municipal and local taxes (sales, consumer, use and similar) plus all "
             "tariffs and all price escalation (§4.1 / Scope). No tax-exempt certificate or "
             "exemption clause appears on the agreement face.",
             "The Subcontract Sum is a fixed, firm, and all-inclusive price ... includes all "
             "applicable federal, state, county, municipal and local taxes including but not "
             "limited to all applicable sales, consumer, use and similar taxes, and all tariffs "
             "applicable to the Work. All price escalation is included in the Subcontract Sum.")

    # --- Personal / performance guaranty ---
    # No personal-guaranty clause found in this DRAFT (unlike Schaaf). The analysis
    # block notes bonds are NOT required (§3.5). Leave SF.personal_guaranty blank.

    # NOTE: contract duration in calendar days is NOT stated as a single figure on the
    # agreement face; the schedule lives in Exhibit H (Project Schedule). The analysis
    # block references a ~23-day critical-path window from that schedule. Left blank
    # here (no single verbatim "N calendar days" value to quote) — never inferred.

    return {
        "fields": fields,
        "snippets": snippets,
    }


def _norm_date(mdy):
    """MM/DD/YYYY -> YYYY-MM-DD; pass through anything unexpected."""
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", mdy.strip())
    if not m:
        return mdy.strip()
    mo, da, yr = m.groups()
    return f"{yr}-{int(mo):02d}-{int(da):02d}"


EXTRACTORS = {
    "schaaf": parse_subcontract_schaaf,
    "pp": parse_subcontract_pp,
}


def deep_extract_subcontract(token, cfg):
    """Resolve + fetch + verify + extract a deep-project subcontract.
    Returns the subcontract block (POET shape: fields/snippets/source_file/
    pages/scanned_pages/item_id) or raises on a verification failure."""
    from pypdf import PdfReader

    # resolve drive-item id (for the inline /api/doc embed)
    item = get_item_by_path(token, cfg["subcontract_path"])
    if not item or not item.get("id"):
        raise RuntimeError(f"Could not resolve drive item for {cfg['subcontract_path']}")
    item_id = item["id"]
    size = item.get("size")
    mime = (item.get("file") or {}).get("mimeType")

    # fetch bytes (proof) + parse text
    raw = download_path(token, cfg["subcontract_path"])
    if raw[:5] != b"%PDF-":
        raise RuntimeError("Fetched content is not a PDF (missing %PDF- header)")
    reader = PdfReader(io.BytesIO(raw))
    pages = [(pg.extract_text() or "") for pg in reader.pages]
    # This Patterson Horth PDF uses non-breaking spaces (\xa0) between words.
    # Normalize nbsp -> space so verification + extraction regexes match.
    pages = [pg.replace("\xa0", " ") for pg in pages]
    full = "\n".join(pages)
    scanned = [i + 1 for i, t in enumerate(pages) if len(t.strip()) < 20]

    # VERIFY the doc is what we expect BEFORE trusting any extracted value
    ok, missing = _verify_subcontract(full, cfg.get("verify_tokens", []))
    if not ok:
        raise RuntimeError(f"Subcontract verification FAILED — missing tokens {missing}. "
                           "Refusing to extract (possible mislabeled subcontract).")

    extractor = EXTRACTORS[cfg["extractor"]]
    parsed = extractor(full)

    sub = {
        "fields": parsed["fields"],
        "snippets": parsed["snippets"],
        "source_file": cfg["subcontract_file"],
        "pages": len(pages),
        "scanned_pages": scanned,
        "item_id": item_id,
    }
    return sub, {"item_id": item_id, "size": size, "mime": mime,
                 "fetched_bytes": len(raw), "verified": ok}


# ---- bulk-file read/merge/write (preserves all other records) ----
def _read_bulk():
    with open(BULK_JS) as f:
        txt = f.read()
    m = re.search(r"window\.PF_PROJECT_RECORDS\s*=\s*(\{.*\})\s*;", txt, re.S)
    if not m:
        raise RuntimeError("Could not parse window.PF_PROJECT_RECORDS in project-records.js")
    return json.loads(m.group(1)), txt


def _write_bulk(data, prior_header_lines):
    with open(BULK_JS, "w") as f:
        for line in prior_header_lines:
            f.write(line + "\n")
        f.write("window.PF_PROJECT_RECORDS = ")
        json.dump(data, f, indent=2)
        f.write(";\n")


def _bulk_header_lines(orig_txt):
    """Preserve the original leading // comment lines (the file is auto-generated;
    we keep its banner verbatim and just re-emit the data object)."""
    lines = []
    for line in orig_txt.splitlines():
        if line.startswith("//"):
            lines.append(line)
        else:
            break
    return lines


def deepen_project(token, num):
    """Extract the subcontract for project `num` and MERGE it (plus the contract
    item_id + any discrepancy vs the Bid Log) into the existing bulk record,
    leaving every other field and every other record untouched."""
    cfg = PROJECT_DEEP.get(num)
    if not cfg:
        print(f"ERROR: no deep-extraction config for project {num}. "
              f"Known: {', '.join(PROJECT_DEEP)}", file=sys.stderr)
        sys.exit(2)

    data, orig_txt = _read_bulk()
    records = data.get("records", {})
    if num not in records:
        print(f"ERROR: {num} not present in project-records.js (run build-project-records.py first).",
              file=sys.stderr)
        sys.exit(2)

    sub, proof = deep_extract_subcontract(token, cfg)

    rec = records[num]

    # SURGICAL FIELD MERGE — only ADD/UPDATE the subcontract, discrepancies, item_id
    # (carried inside sub), execution_status, generated, and data_note. EVERY other
    # field of this record is left exactly as-is. In particular the rich `analysis`
    # block (verdict, risks, execution_status, key_terms, ...) migrated from the
    # standalone Subcontracts review is NOT touched — we never overwrite or drop it.
    rec["subcontract"] = sub

    # Surface a top-level execution_status on the record so the contract's
    # signed/draft state is queryable on the record itself (this DRAFT = unsigned).
    # The analysis block keeps its own execution_status independently (unchanged).
    if cfg.get("execution_status"):
        rec["execution_status"] = cfg["execution_status"]

    # discrepancy: Bid Log value vs subcontract Stipulated/Subcontract Sum (Bid Log wins)
    bid = rec.get("bid_log") or {}
    discrepancies = []
    sv = sub["fields"].get("subcontract_value")
    bv = bid.get("bid_total_value")
    if sv and bv and _norm_money(sv) != _norm_money(bv):
        discrepancies.append({
            "field": "Contract / Bid Value",
            "bid_log": "$" + str(bv),
            "subcontract": sv,
        })
    rec["discrepancies"] = discrepancies

    # refresh the data_note + generated stamp to reflect the deepening
    rec["generated"] = datetime.now(timezone.utc).isoformat() + "Z"
    draft_prefix = ("DRAFT subcontract (not executed). " if cfg.get("execution_status")
                    and "DRAFT" in cfg["execution_status"].upper() else "")
    rec["data_note"] = (
        draft_prefix +
        "READ-ONLY. Bulk-populated from data we already hold (Bid Log metrics/dates/award, "
        "Project Info contacts, folder doc links, auto-progress QA/QC) AND deepened to POET level "
        "with a fresh per-job subcontract extraction (dates, parties, value, scope, retainage, "
        "payment terms incl. pay-when/if-paid, liquidated damages, bonds, certified payroll/labor, "
        "insurance, working hours, surveying, tax) — each with a verbatim PDF snippet. Bid Log wins "
        "on shared fields; the subcontract only fills blanks or confirms, and any disagreement is "
        "noted, never silently overwritten. The migrated subcontract analysis (verdict + risks) is "
        "preserved. Fields with no confirmed source render blank — never fabricated.")

    header = _bulk_header_lines(orig_txt)
    _write_bulk(data, header)

    # ---- console summary + proof ----
    print(f"\nDEEPENED {num} ({cfg['name']}):", file=sys.stderr)
    print(f"  contract item_id : {proof['item_id']}", file=sys.stderr)
    print(f"  fetched bytes    : {proof['fetched_bytes']} | mime: {proof['mime']} | "
          f"verified: {proof['verified']}", file=sys.stderr)
    print(f"  subcontract pages: {sub['pages']} | scanned/blank: {sub['scanned_pages'] or 'none'}",
          file=sys.stderr)
    print(f"  extracted fields : {len(sub['fields'])}", file=sys.stderr)
    for k, v in sub["fields"].items():
        print(f"      [SUB] {k}: {v}", file=sys.stderr)
    if discrepancies:
        print("  discrepancies (Bid Log kept):", file=sys.stderr)
        for d in discrepancies:
            print(f"      ! {d['field']}: bidlog={d['bid_log']} vs subcontract={d['subcontract']}",
                  file=sys.stderr)
    print(f"\nWrote (merged): {BULK_JS}", file=sys.stderr)


def main():
    args = sys.argv[1:]
    dump = "--dump" in args
    if not DRIVE_ID:
        print("ERROR: SP_DRIVE_ID not set in /home/aiciv/.env", file=sys.stderr)
        sys.exit(1)

    # --project NN-NNN : deepen a bulk record (merge subcontract + embed id)
    proj = None
    if "--project" in args:
        i = args.index("--project")
        if i + 1 < len(args):
            proj = args[i + 1]
    if proj:
        print(f"build-project-record (deepen {proj}) — "
              f"{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC", file=sys.stderr)
        token = _token()
        deepen_project(token, proj)
        return

    # default: rebuild the POET deep record
    print(f"build-project-record (POET) — {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC",
          file=sys.stderr)
    token = _token()
    record = assemble(token)
    _summary(record)
    if dump:
        print(json.dumps(record, indent=2))
        return
    out = write_js(record)
    print(f"\nWrote: {out}", file=sys.stderr)


if __name__ == "__main__":
    main()
