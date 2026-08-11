#!/usr/bin/env python3
"""
Build Garbin Prelim data  ->  platform/data/garbin-prelim.js
==============================================================
For every active project under SharePoint

    04 - Project Management/02 - Projects/<project folder>/
        03 - Engineering & Design/Garbin Prelim/<AP Preliminary Design Summary ...>.xlsx

this reads the "Prelim Design Summary" tab and extracts the six values the portal's
Engineering & Design section displays:

    - LF               (Grand Total LF)
    - total_columns    (Grand Total COL)
    - nominal_dia_ft   (Nominal Diameter, RAW feet  -> frontend x12 -> inches)
    - bearing_psf      (Bearing Capacity, psf)
    - stone_tn         (Stone Required, tons)
    - webUrl           (clickable "open the workbook" link)

Output: platform/data/garbin-prelim.js
    window.PF_GARBIN_PRELIM = { projects: { "26-002": {...}, ... }, meta: {...} };

WHY window.PF_GARBIN_PRELIM (not PF_GARBIN):
    index.html already uses window.PF_GARBIN for the per-bid "sent to Garbin"
    checkbox flags (KV-backed feature flags). Reusing that name would collide, so
    this Garbin-Prelim design data gets its own distinct global.

READ DISCIPLINE (no fabrication):
    - Values are LABEL-ANCHORED, not row-hardcoded: for each target we scan column R
      for the label (e.g. "Stone Required:") and read column T on the SAME row, with
      column U as the unit. This survives row shifts across projects. If a project's
      layout differs (a label missing), that value is left null (frontend shows blank)
      and the miss is reported in the run summary + baked into the entry's "missing"
      list. Nothing is invented.

FOLDER RESOLUTION + OLD EXCLUSION (TWO possible prelim folder names):
    - Enumerate the project folders, take the leading "NN-NNN" as the project number.
    - Under "<folder>/03 - Engineering & Design", look for the prelim subfolder under
      EITHER name (case-insensitive), older-then-newer:
          "Garbin Prelim"  (older projects, e.g. 26-002 POET)
          "Bid Prelim"     (newer projects, e.g. 26-013 Park & Poplar - OldTown)
      The names live in PRELIM_SUBFOLDER_NAMES so the list is easy to extend later.
    - Use the current Excel directly IN that folder. Any "OLD"/"old" SUBFOLDER is
      DISREGARDED (superseded workbooks live there). If several .xlsx sit directly in
      the folder, prefer one whose name contains "Preliminary Design Summary", else the
      most-recently-modified.
    - If a project has BOTH prelim folders, we pick the one that actually holds a
      current .xlsx. If BOTH hold files we FLAG it (both-folders) and prefer the newer
      structure ("Bid Prelim", last in PRELIM_SUBFOLDER_NAMES).
    - Projects with no prelim folder / no Excel simply get NO entry (the frontend
      renders nothing for them, gracefully).

Reuses the Graph token + env loader from /home/aiciv/tools/pf_email.py, exactly like
build-project-record.py / sp-sync.py.

Thread caps are set BEFORE openpyxl import (this box, ~300 pid limit; solved-box-gotchas).

Usage:
    python3 build-garbin-prelim.py            # writes data/garbin-prelim.js
    python3 build-garbin-prelim.py --dump     # print extracted data to stdout, no write
    python3 build-garbin-prelim.py --only 26-002   # restrict to one project number
"""

# --- thread caps BEFORE any heavy import (solved-box-gotchas) ---
import os
for _v in ("OPENBLAS_NUM_THREADS", "OMP_NUM_THREADS", "MKL_NUM_THREADS",
           "NUMEXPR_NUM_THREADS", "VECLIB_MAXIMUM_THREADS"):
    os.environ.setdefault(_v, "1")

import io
import re
import sys
import json
import urllib.request
import urllib.parse
import urllib.error
from datetime import datetime, timezone

# --- reuse the Graph token helper + env loader from pf_email.py ---
sys.path.insert(0, "/home/aiciv/tools")
from pf_email import _token, _env  # noqa: E402

GRAPH = "https://graph.microsoft.com/v1.0"

HERE = os.path.dirname(os.path.abspath(__file__))
PLATFORM = os.path.dirname(HERE)
DATA_DIR = os.path.join(PLATFORM, "data")
OUT_JS = os.path.join(DATA_DIR, "garbin-prelim.js")

PROJECTS_BASE = "04 - Project Management/02 - Projects"
ENG_BASE = "03 - Engineering & Design"
# The prelim Excel lives in ONE of these subfolders under "03 - Engineering & Design".
# The file structure was updated on more recent projects (Brad 2026-08-11):
#   - "Garbin Prelim" = older structure (e.g. 26-002 POET)
#   - "Bid Prelim"    = newer structure (e.g. 26-013 Park & Poplar - OldTown)
# Matched case-insensitively. Easy to extend later -- just add a name here. When a
# project has BOTH folders with files, we FLAG and prefer the newer structure order
# (later entries win the tiebreak; "Bid Prelim" is listed last on purpose).
PRELIM_SUBFOLDER_NAMES = ["Garbin Prelim", "Bid Prelim"]
EXTRACT_TAB = "Prelim Design Summary"

# FALLBACK (Brad's rule of thumb, 2026-08-11): every job that reaches the active-projects
# phase HAS a prelim done, so the AP Preliminary Design Summary workbook WILL exist somewhere
# under "03 - Engineering & Design" -- most often in a "Bid Prelim" folder, but sometimes in a
# differently-named (or no) subfolder. When neither NAMED prelim folder yields a file, we walk
# the ENTIRE "03 - Engineering & Design" tree and match a workbook by NAME.
#
# The filename filter is DELIBERATELY specific so we never grab an unrelated Excel (budget,
# takeoff, submittal log, RFI log, etc.). Both known real files are named
# "AP Preliminary Design Summary <project>.xlsx". We accept, case-insensitively, a filename that
# contains any of these distinctive phrases:
PRELIM_NAME_PATTERNS = [
    re.compile(r"preliminary\s*design\s*summary", re.I),  # canonical (most known files)
    re.compile(r"prelim\s*design\s*summary", re.I),       # abbreviated variant
    # "AP Preliminary Design ..." WITHOUT the word "Summary" -- same document, older naming
    # (e.g. 26-007 "AP Preliminary Design Madison Lifestyle.xlsx"). The "AP" (aggregate pier)
    # prefix keeps this specific to the prelim workbook, not a generic "design" file.
    re.compile(r"\bap\s*preliminary\s*design\b", re.I),
    re.compile(r"\bap\s*prelim\b", re.I),                 # "AP Prelim ..." variant
]
# Any path segment containing one of these words means a stale/superseded copy -> never read it.
STALE_DIR_PATTERNS = [
    re.compile(r"\bold\b", re.I),
    re.compile(r"archive", re.I),
    re.compile(r"superseded", re.I),
]
# Cap the tree walk so a pathological deep/huge Engineering folder can't blow up the sync.
FALLBACK_MAX_DEPTH = 6


def _name_is_prelim(name):
    n = str(name or "")
    return any(rx.search(n) for rx in PRELIM_NAME_PATTERNS)


def _segment_is_stale(name):
    n = str(name or "")
    return any(rx.search(n) for rx in STALE_DIR_PATTERNS)

# Only real project folders start with an "NN-NNN" number. Skip the completed-projects
# roll-up folder, the template placeholder, and any non-project folders.
PROJNUM_RE = re.compile(r"^(\d{2}-\d{3})\b")

_env_cache = _env()
DRIVE_ID = _env_cache.get("SP_DRIVE_ID", "")

# Label-anchored targets. Each = (output_key, label_regex_on_col_R). We scan column R
# for the label and read column T (value) + column U (unit) on the SAME row.
TARGETS = [
    ("total_columns", re.compile(r"grand\s*total\s*col", re.I)),
    ("lf",            re.compile(r"grand\s*total\s*lf", re.I)),
    ("nominal_dia",   re.compile(r"nominal\s*diameter", re.I)),
    ("bearing",       re.compile(r"bearing\s*capacity", re.I)),
    ("stone",         re.compile(r"stone\s*required", re.I)),
]

R_COL = 18  # column R (1-based)
T_COL = 20  # column T (value)
U_COL = 21  # column U (unit)


# ---------------- Graph helpers (pattern from build-project-record.py) ----------------
def gget(token, url):
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    return json.loads(urllib.request.urlopen(req).read())


def list_children_by_path(token, path):
    p = urllib.parse.quote(path)
    url = f"{GRAPH}/drives/{DRIVE_ID}/root:/{p}:/children"
    items = []
    while url:
        data = gget(token, url)
        items.extend(data.get("value", []))
        url = data.get("@odata.nextLink")
    return items


def try_list_children_by_path(token, path):
    """Like list_children_by_path but returns None (not raise) on 404 -- used to probe
    for a folder that may not exist for a given project."""
    try:
        return list_children_by_path(token, path)
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return None
        raise


def list_children_by_item(token, item_id):
    """List the direct children of a driveItem by its id (used to walk a subtree)."""
    url = f"{GRAPH}/drives/{DRIVE_ID}/items/{item_id}/children"
    items = []
    while url:
        data = gget(token, url)
        items.extend(data.get("value", []))
        url = data.get("@odata.nextLink")
    return items


def download_item_content(token, item_id):
    url = f"{GRAPH}/drives/{DRIVE_ID}/items/{item_id}/content"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    return urllib.request.urlopen(req).read()


# ---------------- value helpers ----------------
def _num(v):
    """Coerce a cell value to a JSON-friendly number (int when whole), else None."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return int(f) if f == int(f) else f
    s = str(v).strip().replace(",", "")
    if not s:
        return None
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except ValueError:
        return None


def _unit(v):
    return ("" if v is None else str(v).strip())


# ---------------- Excel extraction (label-anchored col R -> col T) ----------------
def extract_from_workbook(raw_bytes):
    """Return (values_dict, missing_list). values_dict has the raw numbers keyed by
    output key; missing_list names any target whose label was not found. NO fabrication:
    a target that is not found is simply absent from values_dict (and listed in missing)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
    if EXTRACT_TAB not in wb.sheetnames:
        # layout differs -- report the tab miss; caller treats as fully-missing.
        return {}, {}, [f"tab:{EXTRACT_TAB}"]
    ws = wb[EXTRACT_TAB]

    # Build a label->row index by scanning column R once.
    values = {}
    units = {}
    missing = []
    for key, rx in TARGETS:
        found_row = None
        for r in range(1, ws.max_row + 1):
            label = ws.cell(row=r, column=R_COL).value
            if label is not None and rx.search(str(label)):
                found_row = r
                break
        if found_row is None:
            missing.append(key)
            continue
        val = _num(ws.cell(row=found_row, column=T_COL).value)
        unit = _unit(ws.cell(row=found_row, column=U_COL).value)
        if val is None:
            missing.append(key)
            continue
        values[key] = val
        units[key] = unit
    return values, units, missing


def build_entry(values, units, missing, webUrl, source_file, item_id):
    """Assemble the per-project portal entry from extracted raw numbers.
    nominal_dia is stored RAW in feet (frontend multiplies x12 -> inches).
    Only keys actually extracted are populated; the rest are null (frontend -> blank)."""
    entry = {
        "lf": values.get("lf"),
        "total_columns": values.get("total_columns"),
        # RAW nominal diameter in FEET; the frontend converts x12 and labels "(in)".
        "nominal_dia_ft": values.get("nominal_dia"),
        "bearing_psf": values.get("bearing"),
        "stone_tn": values.get("stone"),
        "webUrl": webUrl or "",
        "source_file": source_file,
        "item_id": item_id,
        "tabs": ["Prelim Design Summary", "Design Notes"],
        "units": {  # units as READ from the workbook (col U), for verification/provenance
            "lf": units.get("lf", ""),
            "total_columns": units.get("total_columns", ""),
            "nominal_dia_ft": units.get("nominal_dia", ""),
            "bearing_psf": units.get("bearing", ""),
            "stone_tn": units.get("stone", ""),
        },
        "missing": missing,
    }
    return entry


# ---------------- per-project resolution ----------------
def pick_prelim_excel(children):
    """From the direct children of a 'Garbin Prelim' folder, choose the current Excel.
    - Ignore SUBFOLDERS entirely (this excludes any OLD/old folder -- superseded files).
    - Prefer a file whose name contains 'Preliminary Design Summary'; else the most
      recently modified .xlsx. Returns the chosen driveItem, or None."""
    xlsx = [c for c in children
            if not c.get("folder") and str(c.get("name", "")).lower().endswith((".xlsx", ".xlsm"))]
    if not xlsx:
        return None
    preferred = [c for c in xlsx if "preliminary design summary" in str(c.get("name", "")).lower()]
    pool = preferred or xlsx

    def _mtime(c):
        return c.get("lastModifiedDateTime", "") or ""
    pool.sort(key=_mtime, reverse=True)
    return pool[0]


def fallback_prelim_search(token, eng_root_item):
    """FALLBACK when neither named prelim folder yields a file (Brad's rule: an active
    project HAS a prelim somewhere under '03 - Engineering & Design').

    Walk the ENTIRE Engineering & Design subtree (direct files AND subfolders) and collect
    every .xlsx/.xlsm whose NAME matches the prelim pattern (PRELIM_NAME_PATTERNS). Any path
    that passes through an OLD/archive/superseded folder is EXCLUDED (stale copies). Returns a
    list of candidate dicts: {item, path, in_named_folder, mtime}. Empty list = nothing found.

    The name filter is specific enough to skip budgets/takeoffs/submittal logs/etc. Ranking
    (which candidate wins) is done by the caller.
    """
    candidates = []
    named_lower = {n.lower() for n in PRELIM_SUBFOLDER_NAMES}
    # Stack of (item, path_segments_so_far, depth). Root's own name is not part of the path.
    stack = [(eng_root_item, [], 0)]
    while stack:
        node, segs, depth = stack.pop()
        try:
            kids = list_children_by_item(token, node.get("id", ""))
        except urllib.error.HTTPError as e:
            if e.code == 404:
                continue
            raise
        for c in kids:
            cname = str(c.get("name", ""))
            if c.get("folder"):
                if _segment_is_stale(cname):
                    continue  # do not descend into OLD/archive/superseded
                if depth + 1 <= FALLBACK_MAX_DEPTH:
                    stack.append((c, segs + [cname], depth + 1))
                continue
            # file
            if not cname.lower().endswith((".xlsx", ".xlsm")):
                continue
            if not _name_is_prelim(cname):
                continue
            in_named = any(s.lower() in named_lower for s in segs)
            candidates.append({
                "item": c,
                "path": "/".join(segs + [cname]),
                "in_named_folder": in_named,
                "mtime": c.get("lastModifiedDateTime", "") or "",
            })
    return candidates


def _get_item_by_path(token, path):
    """Fetch a single driveItem (folder or file) by path. Returns the item, or None on 404."""
    p = urllib.parse.quote(path)
    url = f"{GRAPH}/drives/{DRIVE_ID}/root:/{p}"
    try:
        return gget(token, url)
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return None
        raise


def resolve_prelim_excel(token, folder):
    """For one project folder, resolve the prelim Excel under '03 - Engineering & Design'.

    Priority order (Brad's rule: an active project HAS a prelim somewhere under E&D):
      1. NAMED prelim subfolders first (current behavior): 'Garbin Prelim' then 'Bid Prelim'
         (newer 'Bid Prelim' wins the tiebreak if both hold files).
      2. FALLBACK: if neither named folder yields a file, walk the ENTIRE E&D tree and match a
         workbook by NAME (PRELIM_NAME_PATTERNS). OLD/archive/superseded folders excluded.
         Multiple matches -> prefer one inside a named prelim folder, else newest by mtime; flag.
      3. Nothing found either way -> no entry.

    Returns a dict:
      {found: bool, item, prelim_folder, source: 'named'|'fallback'|None,
       both_flag: bool, multi_flag: bool, path, candidate_paths: [..]}
    """
    result = {
        "found": False, "item": None, "prelim_folder": None, "source": None,
        "both_flag": False, "multi_flag": False, "path": None, "candidate_paths": [],
        # ranked (item, folder_label, path) fallback candidates, best-first, so build() can
        # skip a name-matched file that turns out to lack the Prelim Design Summary tab.
        "fallback_ranked": [],
    }
    eng_root = _get_item_by_path(token, f"{PROJECTS_BASE}/{folder}/{ENG_BASE}")
    if eng_root is None:
        return result  # no Engineering & Design folder at all

    eng_kids = try_list_children_by_path(token, f"{PROJECTS_BASE}/{folder}/{ENG_BASE}")
    if eng_kids is None:
        eng_kids = []

    # ---- (1) NAMED prelim subfolders ----
    subfolders = {str(c.get("name", "")).lower(): c
                  for c in eng_kids if c.get("folder")}
    hits = []  # (order_index, actual_name, chosen_excel_item)
    eng_path = f"{PROJECTS_BASE}/{folder}/{ENG_BASE}"
    for idx, cand in enumerate(PRELIM_SUBFOLDER_NAMES):
        sub = subfolders.get(cand.lower())
        if sub is None:
            continue
        actual_name = str(sub.get("name", cand))
        kids = try_list_children_by_path(token, f"{eng_path}/{actual_name}")
        if kids is None:
            continue
        chosen = pick_prelim_excel(kids)
        if chosen is not None:
            hits.append((idx, actual_name, chosen))

    if hits:
        both = len(hits) > 1
        hits.sort(key=lambda h: h[0])
        _, actual_name, chosen = hits[-1]  # newest structure (last in list) wins
        result.update(found=True, item=chosen, prelim_folder=actual_name,
                      source="named", both_flag=both,
                      path=f"{actual_name}/{chosen.get('name', '')}")
        return result

    # ---- (2) FALLBACK: walk the whole E&D tree by filename ----
    candidates = fallback_prelim_search(token, eng_root)
    if not candidates:
        return result  # (3) genuinely nothing found

    result["candidate_paths"] = [c["path"] for c in candidates]
    # Rank: prefer a candidate INSIDE a named prelim folder, then newest by mtime.
    candidates.sort(key=lambda c: (c["in_named_folder"], c["mtime"]), reverse=True)

    def _folder_label(path):
        d = os.path.dirname(path)
        return "(fallback) " + d if d else "(fallback) E&D root"

    ranked = [(c["item"], _folder_label(c["path"]), c["path"]) for c in candidates]
    best = candidates[0]
    result.update(found=True, item=best["item"],
                  prelim_folder=_folder_label(best["path"]),
                  source="fallback", multi_flag=len(candidates) > 1,
                  path=best["path"], fallback_ranked=ranked)
    return result


def resolve_projects(token, only=None):
    """Enumerate project folders and yield (project_number, folder_name)."""
    children = list_children_by_path(token, PROJECTS_BASE)
    for it in children:
        if not it.get("folder"):
            continue
        name = str(it.get("name", ""))
        m = PROJNUM_RE.match(name)
        if not m:
            continue
        projnum = m.group(1)
        if only and projnum != only:
            continue
        yield projnum, name


def _extract_with_tab_guard(token, item):
    """Download + extract ONE candidate workbook. Returns:
      ('ok', values, units, missing)      -- has the Prelim Design Summary tab
      ('no-tab', None, None, None)         -- lacks the tab (probably the wrong file -> skip it)
      ('error', err_str, None, None)       -- download/parse failure
    NO fabrication: a file without the extraction tab is never mined for values."""
    try:
        raw = download_item_content(token, item.get("id", ""))
    except Exception as e:  # noqa: BLE001
        return ("error", f"download: {e}", None, None)
    try:
        values, units, missing = extract_from_workbook(raw)
    except Exception as e:  # noqa: BLE001
        return ("error", f"parse: {e}", None, None)
    if missing == [f"tab:{EXTRACT_TAB}"]:
        return ("no-tab", None, None, None)
    return ("ok", values, units, missing)


def build(token, only=None, verbose=True):
    projects = {}
    report = []  # (projnum, status, detail)
    for projnum, folder in resolve_projects(token, only=only):
        res = resolve_prelim_excel(token, folder)
        if not res["found"]:
            report.append((projnum, "no-prelim",
                           "no prelim in named folders OR anywhere under E&D (fallback searched)"))
            continue

        source = res["source"]
        # Build the ordered list of candidates to TRY. For a fallback with multiple matches we
        # try them best-first, skipping any that lack the Prelim Design Summary tab (wrong file).
        if source == "fallback" and res["fallback_ranked"]:
            trylist = res["fallback_ranked"]  # (item, folder_label, path)
        else:
            trylist = [(res["item"], res["prelim_folder"], res["path"])]

        skipped_no_tab = []
        chosen_tuple = None
        outcome = None
        for item, folder_label, path in trylist:
            kind, values, units, missing = _extract_with_tab_guard(token, item)
            if kind == "error":
                report.append((projnum, "download/parse-error", f"{path}: {values}"))
                outcome = "hard-error"
                break
            if kind == "no-tab":
                skipped_no_tab.append(path)
                continue
            chosen_tuple = (item, folder_label, path, values, units, missing)
            break

        if outcome == "hard-error":
            continue
        if chosen_tuple is None:
            # every candidate lacked the tab -> no entry, flag WHY.
            detail = f"candidate(s) lacked '{EXTRACT_TAB}' tab (skipped, not extracted): " \
                     + "; ".join(skipped_no_tab)
            report.append((projnum, "no-prelim (wrong-file)", detail))
            continue

        item, folder_label, path, values, units, missing = chosen_tuple
        webUrl = item.get("webUrl", "")
        src_name = item.get("name", "")
        item_id = item.get("id", "")
        entry = build_entry(values, units, missing, webUrl, src_name, item_id)
        entry["prelim_folder"] = folder_label      # which folder/structure this came from
        entry["source"] = source                   # 'named' or 'fallback'
        entry["source_path"] = path                 # provenance path under E&D
        projects[projnum] = entry

        flags = []
        if source == "named" and res["both_flag"]:
            flags.append("[BOTH named prelim folders had files -- used newer]")
        if source == "fallback":
            flags.append("[FALLBACK tree-search]")
            if res["multi_flag"]:
                flags.append(f"[{len(res['candidate_paths'])} name-matches -- picked best]")
        if skipped_no_tab:
            flags.append(f"[skipped {len(skipped_no_tab)} no-tab candidate(s)]")
        flag = ("  " + " ".join(flags)) if flags else ""
        detail = f"[{folder_label}] {src_name}{flag}"
        status = "ok" if not missing else f"ok (missing: {','.join(missing)})"
        report.append((projnum, status, detail))

    data = {
        "projects": projects,
        "meta": {
            "generated": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            "source": "SharePoint 03 - Engineering & Design/{Garbin Prelim|Bid Prelim}/<AP Preliminary Design Summary>.xlsx (or, fallback, any 'Preliminary Design Summary' workbook anywhere under E&D), tab 'Prelim Design Summary'",
            "extraction": "label-anchored col R -> value col T (unit col U)",
            "note": "nominal_dia_ft is RAW feet; frontend x12 -> inches. Prelim from named folder 'Garbin Prelim' (older) OR 'Bid Prelim' (newer); FALLBACK = whole-E&D-tree filename search when neither named folder has a file (OLD/archive/superseded folders excluded, tab-verified). OLD subfolders disregarded.",
            "project_count": len(projects),
        },
    }
    if verbose:
        print("Garbin Prelim sync:")
        for projnum, status, detail in report:
            print(f"  {projnum}: {status}  {detail}")
        print(f"  -> {len(projects)} project(s) populated")
    return data


def write_js(data):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(OUT_JS, "w") as f:
        f.write("// AUTO-GENERATED by sync/build-garbin-prelim.py -- do not edit by hand.\n")
        f.write("// Garbin Prelim design values per project (Engineering & Design section).\n")
        f.write("// Source: SharePoint '03 - Engineering & Design/Garbin Prelim' AP Preliminary\n")
        f.write("// Design Summary workbook, tab 'Prelim Design Summary'. Label-anchored reads.\n")
        f.write("// Distinct global from window.PF_GARBIN (the per-bid sent-to-Garbin flags).\n")
        f.write("window.PF_GARBIN_PRELIM = ")
        json.dump(data, f, indent=2)
        f.write(";\n")
    return OUT_JS


def main():
    only = None
    dump = False
    args = sys.argv[1:]
    i = 0
    while i < len(args):
        a = args[i]
        if a == "--dump":
            dump = True
        elif a == "--only" and i + 1 < len(args):
            only = args[i + 1]
            i += 1
        i += 1

    if not DRIVE_ID:
        print("ERROR: SP_DRIVE_ID not set in /home/aiciv/.env", file=sys.stderr)
        return 2
    token = _token()
    data = build(token, only=only)
    if dump:
        print(json.dumps(data, indent=2))
        return 0
    path = write_js(data)
    print(f"Wrote {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
