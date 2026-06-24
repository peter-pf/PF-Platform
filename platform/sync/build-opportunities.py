#!/usr/bin/env python3
"""
Opportunities ingest (build-opportunities.py)
=============================================
Reads the PF BD Master workbook "Opportunities" tab and emits the read-only
BASE feed for the BD CRM Opportunities module. The FIELD LIST is taken from the
tab's header ROW 4 (the spec'd headers). We do NOT invent fields. Missing
fields/tab are OMITTED + reported, never fabricated.

Output: platform/data/opportunities.js
  window.PF_OPPORTUNITIES = {
    opportunities: [ { id, fields:{...row4 fields...} } ],
    fields: [ ...row4 headers in order... ],
    generated, source, sourceTab
  }

This is the READ-ONLY BASE. New + edited opportunities, the feasibility
recommendation, the email-bridge flags, and the go/pass decision all live in KV
(opp_overlay_v1, written by /api/opportunity) and merge on top in the UI. A
re-sync of this file never erases KV data.

GATING: BD's own tool data. /data/opportunities.js -> business_dev area in
auth.js DATA_FILE_AREAS. admin + partner + business_dev only; field_ops BLOCKED.

Thread caps set BEFORE importing openpyxl (solved-box-gotchas).

Usage:
  python3 build-opportunities.py
  python3 build-opportunities.py --local sync/downloads/PF_BD_Master.xlsm
"""

# --- thread caps BEFORE any heavy import ---
import os
for _v in ("OPENBLAS_NUM_THREADS", "OMP_NUM_THREADS", "MKL_NUM_THREADS",
           "NUMEXPR_NUM_THREADS", "VECLIB_MAXIMUM_THREADS"):
    os.environ.setdefault(_v, "1")

import sys
import re
import json
import hashlib
import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
PLATFORM = os.path.dirname(HERE)
DATA_DIR = os.path.join(PLATFORM, "data")
OUT_JS = os.path.join(DATA_DIR, "opportunities.js")
DEFAULT_XLSM = os.path.join(HERE, "downloads", "PF_BD_Master.xlsm")

SOURCE_FILE = "PF_BD_Master.xlsm"
OPP_TAB = "Opportunities"
HEADER_ROW = 4              # field list lives in row 4 (1-indexed)
NAME_FIELD = "Project Name"  # the record label column


def _token():
    sys.path.insert(0, "/home/aiciv/tools")
    from pf_email import _token as t
    return t()


def download_workbook():
    import urllib.request
    import urllib.parse
    tok = _token()
    drive = os.environ["SP_DRIVE_ID"]
    q = urllib.parse.quote("PF_BD_Master")
    url = "https://graph.microsoft.com/v1.0/drives/%s/root/search(q='%s')" % (drive, q)
    req = urllib.request.Request(url, headers={"Authorization": "Bearer " + tok})
    res = json.load(urllib.request.urlopen(req, timeout=120))
    fid = None
    for it in res.get("value", []):
        if str(it.get("name", "")).lower().startswith("pf_bd_master"):
            fid = it.get("id"); break
    if not fid:
        raise SystemExit("Could not find PF_BD_Master in SharePoint drive")
    curl = "https://graph.microsoft.com/v1.0/drives/%s/items/%s/content" % (drive, fid)
    creq = urllib.request.Request(curl, headers={"Authorization": "Bearer " + tok})
    return urllib.request.urlopen(creq, timeout=180).read()


def cellval(v):
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def headers_for(ws):
    """Row-4 field headers as an ordered list of (col_index, name)."""
    rows = list(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
    hdr = rows[0] if rows else ()
    out = []
    for i, c in enumerate(hdr):
        name = cellval(c)
        if name:
            out.append((i, name))
    return out


def norm_key(s):
    return re.sub(r"\s+", " ", (s or "").strip()).lower()


def collision_id(prefix, seed, seen):
    """Stable, collision-ONLY id (same scheme as build-bd-records.py). First
    occurrence of a seed keeps the plain hash; the Nth duplicate hashes seed#N,
    so non-colliding ids stay stable across re-syncs and only dups shift."""
    n = seen.get(seed, 0) + 1
    seen[seed] = n
    eff = seed if n == 1 else (seed + "#" + str(n))
    h = hashlib.sha1(eff.encode("utf-8")).hexdigest()[:10]
    return prefix + "_" + h, (n > 1)


def build(xlsm_bytes):
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsm_bytes), data_only=True, read_only=True)

    omitted = []
    if OPP_TAB not in wb.sheetnames:
        wb.close()
        return ({
            "opportunities": [], "fields": [],
            "generated": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "source": SOURCE_FILE, "sourceTab": OPP_TAB,
            "_omitted": ["Tab missing: %r" % OPP_TAB],
        }, ["Tab missing: %r" % OPP_TAB], [])

    ws = wb[OPP_TAB]
    fields = headers_for(ws)
    field_names = [n for (_, n) in fields]
    if NAME_FIELD not in field_names:
        omitted.append("Opportunities row 4 has no %r field (used as the record label)" % NAME_FIELD)

    name_col = None
    for (i, n) in fields:
        if n == NAME_FIELD:
            name_col = i
            break

    opportunities = []
    seen = {}
    id_warnings = []
    if name_col is not None:
        for r in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            if len(r) <= name_col:
                continue
            name = cellval(r[name_col])
            if not name:
                continue
            rec_fields = {}
            for (i, fname) in fields:
                rec_fields[fname] = cellval(r[i]) if i < len(r) else ""
            oid, collided = collision_id("op", norm_key(name), seen)
            if collided:
                id_warnings.append("Duplicate Project Name (collision-suffixed id): %r" % name)
            opportunities.append({"id": oid, "fields": rec_fields})
    wb.close()

    out = {
        "opportunities": opportunities,
        "fields": field_names,
        "generated": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source": SOURCE_FILE,
        "sourceTab": OPP_TAB,
    }
    if omitted:
        out["_omitted"] = omitted
    return out, omitted, id_warnings


def write_js(data):
    note = (
        "// AUTO-GENERATED by sync/build-opportunities.py — DO NOT EDIT BY HAND.\n"
        "// Source: %s :: '%s' tab  (field list = row 4)  generated %s\n"
        "// Opportunities BASE (read-only). New/edited opps, feasibility\n"
        "// recommendation, email-bridge flags + go/pass decision live in KV\n"
        "// (/api/opportunity) and merge on top in the UI.\n"
        "// ACCESS: business_dev — admin/partner/business_dev (field_ops BLOCKED).\n"
        % (SOURCE_FILE, OPP_TAB, data["generated"])
    )
    body = "window.PF_OPPORTUNITIES = " + json.dumps(data, indent=2, ensure_ascii=False) + ";\n"
    with open(OUT_JS, "w") as f:
        f.write(note + body)
    print("wrote", OUT_JS)
    print("  opportunities: %d" % len(data["opportunities"]))
    print("  fields (%d): %s" % (len(data["fields"]), ", ".join(data["fields"])))


def main():
    if "--local" in sys.argv:
        path = sys.argv[sys.argv.index("--local") + 1]
    elif os.path.exists(DEFAULT_XLSM):
        path = DEFAULT_XLSM
    else:
        path = None
    if path:
        with open(path, "rb") as f:
            xlsm = f.read()
        print("source:", path)
    else:
        xlsm = download_workbook()
        print("source: SharePoint download")

    data, omitted, id_warnings = build(xlsm)
    write_js(data)
    if omitted:
        for o in omitted:
            print("  OMITTED:", o)
    else:
        print("  OMITTED: none (Opportunities tab + all row-4 fields present)")
    if id_warnings:
        for w in id_warnings:
            print("  ID WARNING:", w)
    else:
        print("  ID WARNING: none (no duplicate Project Name)")

    ids = [o["id"] for o in data["opportunities"]]
    dupes = set(x for x in ids if ids.count(x) > 1)
    if dupes:
        print("  ID UNIQUENESS: FAIL — %s" % list(dupes)); sys.exit(1)
    print("  ID UNIQUENESS: OK — %d ids all unique" % len(ids))


if __name__ == "__main__":
    main()
