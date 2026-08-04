#!/usr/bin/env python3
"""
Shared Budget-vs-Actual parser  (project-AGNOSTIC)
==================================================
Extracted from build-budget-actual.py (the POET-only builder) so BOTH the POET
builder and the generalized build-budget-actuals.py driver parse identically.

The parse is the proven FORMULA-GRAPH classifier: two openpyxl loads
(data_only=True for evaluated values + data_only=False for formulas). It
classifies each detail row as:
  - MAJOR CATEGORY : a same-sheet roll-up whose refs are THEMSELVES roll-ups
                     (rollup-of-rollups). Opens a group; its subtotal is NOT
                     added to the leaf sum (it IS the sum of its leaves).
  - SUB-ROLLUP     : =SUM(Cx:Cy) etc. Shown as an indented sub-header. NOT summed.
  - LEAF           : a real cost line (code + description + money). The ONLY rows
                     summed. Budget/Actual read from the evaluated values.

Rules (never fabricate):
  - Variance is RECOMPUTED = Budget - Actual (never trusts the stored E cell).
  - Blanks stay blank (None -> rendered em-dash by the portal).
  - A formula cell with NO cached value (workbook never Excel-saved) reads as
    None -> the caller flags a null grand total rather than inventing a number.

Thread caps are the CALLER's responsibility (set BEFORE importing openpyxl;
solved-box-gotchas ~300 pid limit). This module imports openpyxl lazily inside
parse_budget_actual so a caller that set the caps first is honored.
"""

import io
import re

# Cost-code line rows have a numeric code in col A (e.g. "5051", "5710");
# some cells are typed "=5110" -> normalized.
CODE_RX = re.compile(r"^=?\d{3,4}$")

# A row is a same-sheet ROLL-UP when its budget formula sums/adds OTHER rows of
# THIS sheet (e.g. "=SUM(C56:C61)", "=C63+C68+C71", "=C55"). A LEAF pulls its
# value from another sheet or a literal (e.g. "='OH Calcs'!G31", "=0").
ROLLUP_RX = re.compile(r"^=(SUM\(C\d|C\d+[+\-]|C\d+$)", re.I)
CREF_RX = re.compile(r"C(\d+)")


def _txt(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s.upper() in ("N/A", "NA", "TBD"):
        return ""
    return s


def _num(v):
    """Return a float for a numeric/dollar cell, or None if not numeric.
    Blank cells -> None (rendered blank, never 0)."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return None
    s2 = s.replace("$", "").replace(",", "").strip()
    neg = False
    if s2.startswith("(") and s2.endswith(")"):  # parenthesised negative
        neg = True
        s2 = s2[1:-1]
    try:
        f = float(s2)
        return -f if neg else f
    except ValueError:
        return None


def _round2(x):
    return None if x is None else round(x + 0.0, 2)


def _norm_code(a):
    a = (a or "").strip()
    if a.startswith("="):
        a = a[1:].strip()
    return a


# A leaf Actual cell (col D) is frequently a PLAIN ARITHMETIC sum entered by the
# invoice-coding workflow, e.g. "=1050+3230", "=2466.44+2062.21+4171.29", "=500".
# These are the running total of that line's coded invoices. openpyxl can only read
# the CACHED result, which goes stale after any openpyxl write (same trap as the
# Budget col). But because the formula is pure arithmetic over numeric literals, we
# can EVALUATE it directly from the formula string — no Excel, no stale cache. This
# regex admits ONLY digits, dot, and + - * / ( ) and whitespace (NO cell refs, NO
# names, NO functions) so it is safe to eval. A formula with a cell ref (=D6+D9),
# a SUM(), or anything else returns None here (we then fall back to the cached value,
# or leave it blank — never fabricate).
_ARITH_ACTUAL_RX = re.compile(r"^=[0-9.+\-*/() ]+$")


def _eval_arith_formula(f):
    """Return the float value of a pure-arithmetic Excel formula string (e.g.
    '=1050+3230'), or None if the formula is not a safe literal-only arithmetic
    expression. Used to recover a leaf Actual that Excel has not cached (stale)."""
    if not isinstance(f, str):
        return None
    s = f.strip()
    if not s.startswith("=") or not _ARITH_ACTUAL_RX.match(s.replace(" ", "")):
        return None
    expr = s[1:].strip()
    if expr == "":
        return None
    try:
        # literal-only arithmetic; the regex has already excluded names/refs/calls.
        val = eval(expr, {"__builtins__": {}}, {})  # noqa: S307 - constrained input
        if isinstance(val, (int, float)):
            return float(val)
    except (SyntaxError, ZeroDivisionError, TypeError, ValueError):
        return None
    return None


# ---- BOOKED-ACTUAL (GREEN) fill gate ----------------------------------------
# Brad's rule (feedback-portal-actuals-green-only): a col-D Actual is a REAL cost
# ONLY when its cell carries the GREEN booked-actual fill (a real invoice was
# received + coded). YELLOW (FFFF00) is a placeholder Peter has not yet replaced
# with a real invoice — it must NEVER be pulled as an actual. Any other fill /
# no-fill is likewise NOT a booked actual. Fail-safe: only confirmed green pulls;
# ambiguous -> EXCLUDE. A row can thus have a budget but a blank actual (invoice
# not yet received) — that is correct.
#
# The green on the coded cells (verified on 26-002 5405 D104/D106/D108) is the
# Excel theme color theme=9 tint≈0.8 (light green), which resolves to ~E2EFDA.
# We match BOTH the theme+tint signature AND the resolved-RGB E2EFDA family so the
# gate survives a workbook that stores the same green as an explicit RGB.
_GREEN_THEME = 9
_GREEN_TINT = 0.8
_GREEN_TINT_TOL = 0.08
# E2EFDA and close neighbors (light green booked-actual). Compared loosely so a
# minor palette drift still matches, while YELLOW (FFFF00) and greys never do.
_GREEN_RGB_PREFIXES = ("E2EFDA", "C6E0B4", "D9EAD3", "E2F0D9")


def _is_booked_green(cell):
    """Return True IFF this cell's fill is the GREEN booked-actual fill. Yellow
    (FFFF00), grey/blue theme bands, and no-fill all return False (fail-safe)."""
    try:
        fill = cell.fill
        if fill is None or fill.patternType != "solid":
            return False
        fg = fill.fgColor
        # theme-based green (the observed booked-actual style): theme 9 + tint ~0.8
        if getattr(fg, "type", None) == "theme":
            if getattr(fg, "theme", None) == _GREEN_THEME:
                tint = getattr(fg, "tint", 0.0) or 0.0
                if abs(tint - _GREEN_TINT) <= _GREEN_TINT_TOL:
                    return True
            return False  # any other theme (grey/blue bands) is NOT booked-actual
        # explicit-RGB green (some workbooks store the fill as RGB, not theme)
        rgb = getattr(fg, "rgb", None)
        if isinstance(rgb, str) and len(rgb) >= 6:
            hex6 = rgb[-6:].upper()
            if hex6 in ("FFFF00", "FFFFFF00"[-6:]):  # explicit yellow -> never
                return False
            for pre in _GREEN_RGB_PREFIXES:
                if hex6 == pre:
                    return True
        return False
    except Exception:  # noqa: BLE001 - any fill-read oddity -> fail-safe exclude
        return False


def parse_budget_actual(raw_bytes, sheet_name="Budget vs Actual",
                        default_job="", default_name=""):
    """Parse a Turnover Budget workbook's Budget-vs-Actual sheet from raw bytes.

    Args:
      raw_bytes:   the .xlsm/.xlsx file content (bytes).
      sheet_name:  worksheet name (default "Budget vs Actual").
      default_job / default_name: fallbacks if the sheet header is blank.

    Returns a dict:
      { job, name, location, groups:[{title, rows:[...], subtotal}],
        grand_total:{label,budget,actual,variance},
        source_sheet, has_cached_values }
    Raises ValueError if the sheet is absent.

    NOTE: caller must set thread caps BEFORE this runs (openpyxl imported here).
    """
    import openpyxl  # lazy; caller sets thread caps first
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True, keep_vba=False)
    wbf = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=False, keep_vba=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"sheet '{sheet_name}' not found (have: {wb.sheetnames})")
    ws = wb[sheet_name]
    wsf = wbf[sheet_name]

    def cell(r, c):
        return ws.cell(row=r, column=c).value

    def fmla(r, c):
        v = wsf.cell(row=r, column=c).value
        return str(v) if v is not None else ""

    # ---- header (Name / Job # / Location) ----
    job = _txt(cell(1, 4)) or default_job
    name = _txt(cell(1, 2)) or default_name
    location = _txt(cell(2, 2))
    if location.lower() in ("project location", ""):
        location = ""

    # ---- find the detail header row + the summary band ----
    header_row = None
    for r in range(1, min(ws.max_row, 12) + 1):
        if _txt(cell(r, 1)).lower() == "cost code":
            header_row = r
            break
    if header_row is None:
        header_row = 4

    summary_row = None
    for r in range(header_row + 1, ws.max_row + 1):
        if _txt(cell(r, 3)).lower().startswith("estimating budget"):
            summary_row = r
            break
    detail_end = (summary_row - 1) if summary_row else ws.max_row

    # ---- classify via the FORMULA GRAPH ----
    rows_meta = {}
    for r in range(header_row + 1, detail_end + 1):
        cf = fmla(r, 3).replace(" ", "")
        is_rollup = bool(ROLLUP_RX.match(cf))
        refs = [int(x) for x in CREF_RX.findall(cf)] if cf.startswith("=") else []
        refs = [x for x in refs if header_row < x <= detail_end]
        rows_meta[r] = {"cf": cf, "is_rollup": is_rollup, "refs": refs}

    rollup_rows = {r for r, m in rows_meta.items() if m["is_rollup"]}

    # MAJOR CATEGORY = a rollup row whose refs are ALL themselves rollups.
    major_rows = set()
    for r in rollup_rows:
        child_refs = rows_meta[r]["refs"]
        if child_refs and all(c in rollup_rows for c in child_refs):
            major_rows.add(r)

    groups = []
    cur = None
    had_cached = False  # did we see ANY numeric evaluated value? (Excel-saved check)

    def flush():
        nonlocal cur
        if cur is not None:
            groups.append(cur)
        cur = None

    for r in range(header_row + 1, detail_end + 1):
        a_raw = _txt(cell(r, 1))
        a = _norm_code(a_raw)
        b = _txt(cell(r, 2))
        budget = _num(cell(r, 3))
        actual = _num(cell(r, 4))
        vendor = _txt(cell(r, 6))
        notes = _txt(cell(r, 7))
        is_code = bool(CODE_RX.match(a_raw))
        m = rows_meta[r]
        if budget is not None or actual is not None:
            had_cached = True

        if not a and not b and budget is None and actual is None:
            continue  # spacer

        if r in major_rows:
            flush()
            title = b or a or "(Category)"
            cur = {
                "title": title,
                "rows": [],
                "subtotal": {
                    "budget": _round2(budget),
                    "actual": _round2(actual),
                    "variance": _round2((budget or 0.0) - (actual or 0.0)),
                },
            }
            continue

        if cur is None:
            cur = {"title": "(Uncategorized)", "rows": [],
                   "subtotal": {"budget": None, "actual": None, "variance": None}}

        # SUB-ROLLUP (rollup but not major): sub-header. Not summed.
        if m["is_rollup"]:
            cur["rows"].append({
                "cost_code": a if is_code else "",
                "description": b or a,
                "budget": _round2(budget),
                "actual": _round2(actual),
                "variance": _round2((budget or 0.0) - (actual or 0.0)),
                "vendor": vendor,
                "notes": notes,
                "is_subtotal": True,
            })
            continue

        # LEAF cost line. Recover a STALE Actual (col D) from its arithmetic formula
        # when Excel has not cached the value. The invoice-coding workflow enters
        # per-line actuals as pure sums ("=1050+3230"); these are safe to evaluate
        # from the formula string, so the per-row Actual survives an openpyxl re-stale
        # (same defense the budget col needs but can't have — budget refs other
        # sheets). Only fill from the formula when the cached value is missing; a
        # cached value always wins (it IS Excel's own result).
        actual_leaf = actual
        if actual_leaf is None:
            actual_leaf = _round2(_eval_arith_formula(fmla(r, 4)))

        # GREEN-ONLY GATE (Brad, feedback-portal-actuals-green-only): a col-D value is
        # a REAL booked actual ONLY when the cell has the green booked-actual fill.
        # YELLOW (FFFF00) placeholders and any non-green/no-fill are NOT confirmed
        # costs -> the row's actual is BLANK (None). Fail-safe: ambiguous excludes.
        # The gate applies to BOTH the cached value and the formula-recovered value
        # (a yellow cell can still hold a cached/typed number — never pull it).
        if actual_leaf is not None and not _is_booked_green(wsf.cell(row=r, column=4)):
            actual_leaf = None  # not booked-green -> no confirmed actual for this row
        if actual_leaf is not None:
            had_cached = True   # we have a real, GREEN-confirmed per-row actual
        # Skip pure-noise.
        if (budget in (None, 0)) and (actual_leaf in (None, 0)) and not b:
            continue
        variance = None
        if budget is not None or actual_leaf is not None:
            variance = _round2((budget or 0.0) - (actual_leaf or 0.0))
        cur["rows"].append({
            "cost_code": a if is_code else "",
            "description": b,
            "budget": _round2(budget),
            "actual": _round2(actual_leaf),
            "variance": variance,
            "vendor": vendor,
            "notes": notes,
            "is_subtotal": False,
        })
    flush()

    groups = [g for g in groups
              if g["rows"] or any(v not in (None, 0) for v in g["subtotal"].values())]

    # ---- grand total from the summary band ("Total Construction Contract") ----
    grand = {"budget": None, "actual": None, "variance": None, "label": ""}
    if summary_row:
        for r in range(summary_row, ws.max_row + 1):
            lbl = _txt(cell(r, 2))
            if lbl.lower().startswith("total construction contract"):
                gb = _num(cell(r, 3))
                ga = _num(cell(r, 4))
                grand = {
                    "label": lbl,
                    "budget": _round2(gb),
                    "actual": _round2(ga),
                    "variance": _round2((gb or 0.0) - (ga or 0.0)),
                }
                break

    return {
        "job": job,
        "name": name,
        "location": location,
        "groups": groups,
        "grand_total": grand,
        "source_sheet": sheet_name,
        "has_cached_values": had_cached,
    }


def leaf_actuals_by_code(record):
    """Flatten a parsed record into { normalized_group_title: { cost_code: actual } }
    for the LEAF rows only (is_subtotal False, has a cost_code, actual not None).
    Used by the portal overlay to place each workbook Actual onto the matching
    template row (match on group-title + cost_code).

    CRITICAL: a cost code can REPEAT within a group (e.g. 5405 appears 3x under
    Equipment: VSC Rig / Predrill / Fall-Off). We SUM all rows sharing a
    (group,code) so the overlaid per-code total equals the workbook — an earlier
    overwrite dropped all but the last row and undercounted the grand Actual.
    Returns (by_group, flat) where flat is { cost_code: summed_actual } across all
    groups. The portal places this per-(group,code) SUM on the first matching
    template row for that (group,code) so the grand total reconciles exactly."""
    by_group = {}
    flat = {}
    for g in record.get("groups", []):
        gt = _norm_title(g.get("title", ""))
        for row in g.get("rows", []):
            if row.get("is_subtotal"):
                continue
            code = (row.get("cost_code") or "").strip()
            actual = row.get("actual")
            if not code or actual is None:
                continue
            grp = by_group.setdefault(gt, {})
            grp[code] = round(grp.get(code, 0.0) + actual, 2)   # SUM repeats, don't overwrite
            flat[code] = round(flat.get(code, 0.0) + actual, 2)
    return by_group, flat


def _norm_title(s):
    """Normalize a group title for matching (lowercase, collapse ws, strip punct)."""
    s = (s or "").lower().strip()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()
