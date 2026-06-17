#!/usr/bin/env python3
"""
build-awarded-index.py — Awarded Projects index for the PF Platform.

Pulls every "Awarded" row from the Project Bid Log (BOTH the "Agg Pier Bid Log"
and "Helical Pier Bid Log" sheets), and writes data/awarded-projects.js
(window.PF_AWARDED = [...]) consumed by the Awarded Projects panel in index.html.

Source of truth: SharePoint `01 - Admin / 13 - Master Spreadsheets / Project Bid Log.xlsx`.
  - Headers live on a header row (row 6 for Agg Pier, row 3 for Helical). We locate
    columns BY NAME so the two different layouts (Agg Pier has an extra leading
    "Number" column) are both handled correctly. Awarded = Bid Status == "Awarded".

Fields pulled per row: project number, name, city/state, GC, bid total value, discipline.

Only projects that already have a built record link into it (the
RECORD_LINKS map). Everything else shows a muted "record being built".

Never fabricate — only real Awarded rows are emitted.

Run with thread caps BEFORE importing openpyxl (Box pid/thread guard):
  OMP_NUM_THREADS=1 OPENBLAS_NUM_THREADS=1 MKL_NUM_THREADS=1 python3 build-awarded-index.py
"""

import os
# Cap BLAS / OpenMP threads before any heavy import (openpyxl/numpy) — Box thread guard.
for _v in ("OMP_NUM_THREADS", "OPENBLAS_NUM_THREADS", "MKL_NUM_THREADS",
           "NUMEXPR_NUM_THREADS", "VECLIB_MAXIMUM_THREADS"):
    os.environ.setdefault(_v, "1")

import sys
import json
import urllib.request
from datetime import datetime, timezone

sys.path.insert(0, "/home/aiciv/tools")
import pf_email  # noqa: E402  (token + .env reader)

import openpyxl  # noqa: E402

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.normpath(os.path.join(HERE, "..", "data"))
DL_DIR = os.path.join(HERE, "downloads")

BID_LOG_ITEM = "016ISVH6Y7M7KQIB5C5FDLNKI5H3IZFXRK"
BID_LOG_WEBURL = (
    "https://pierfoundations.sharepoint.com/sites/pf/_layouts/15/Doc.aspx?"
    "sourcedoc=%7B04D5671F-A207-46E9-B6A9-1D3ED192DE2A%7D&file=Project%20Bid%20Log.xlsx"
    "&action=default&mobileredirect=true"
)

# Sheets to scan -> discipline label for the index.
SHEETS = [
    ("Agg Pier Bid Log", "Aggregate Piers"),
    ("Helical Pier Bid Log", "Helical Pilings"),
]

# project_number -> showModule id of a BUILT project record.
# Only POET (26-002) has a built record today. Add ids here as records ship.
RECORD_LINKS = {
    "26-002": "project-poet",
}


def _token():
    return pf_email._token()


def _drive():
    return pf_email._env()["SP_DRIVE_ID"]


def download_bid_log():
    os.makedirs(DL_DIR, exist_ok=True)
    fp = os.path.join(DL_DIR, "Project_Bid_Log.xlsx")
    url = "https://graph.microsoft.com/v1.0/drives/%s/items/%s/content" % (_drive(), BID_LOG_ITEM)
    req = urllib.request.Request(url, headers={"Authorization": "Bearer " + _token()})
    with urllib.request.urlopen(req, timeout=120) as r, open(fp, "wb") as f:
        f.write(r.read())
    print("  downloaded Project_Bid_Log.xlsx (%d bytes)" % os.path.getsize(fp))
    return fp


def find_header_row(ws, scan=14):
    """Return the 1-indexed row that contains both 'Bid Status' and 'Project Name'."""
    for r in range(1, scan + 1):
        joined = " ".join(
            str(ws.cell(row=r, column=c).value or "") for c in range(1, 36)
        )
        if "Bid Status" in joined and "Project Name" in joined:
            return r
    return None


def header_map(ws, hr):
    """{header_text: 1-indexed col} for the header row."""
    m = {}
    for c in range(1, 36):
        v = ws.cell(row=hr, column=c).value
        if v not in (None, ""):
            m[str(v).strip()] = c
    return m


def clean(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def num_or_none(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def extract_awarded(ws, discipline):
    hr = find_header_row(ws)
    if not hr:
        print("  WARN: no header row found in '%s' — skipping" % ws.title)
        return []
    h = header_map(ws, hr)
    c_num = h.get("Project Number")
    c_name = h.get("Project Name")
    c_city = h.get("City / State")
    c_gc = h.get("General Contractor")
    c_stat = h.get("Bid Status")
    c_val = h.get("Bid Total Value")
    if not (c_name and c_stat):
        print("  WARN: '%s' missing key columns — skipping" % ws.title)
        return []

    out = []
    for r in range(hr + 1, ws.max_row + 1):
        status = ws.cell(row=r, column=c_stat).value
        if not status or str(status).strip().lower() != "awarded":
            continue
        name = clean(ws.cell(row=r, column=c_name).value) if c_name else ""
        if not name:
            continue  # don't emit blank phantom rows
        number = clean(ws.cell(row=r, column=c_num).value) if c_num else ""
        out.append({
            "number": number,
            "name": name,
            "city_state": clean(ws.cell(row=r, column=c_city).value) if c_city else "",
            "gc": clean(ws.cell(row=r, column=c_gc).value) if c_gc else "",
            "value": num_or_none(ws.cell(row=r, column=c_val).value) if c_val else None,
            "discipline": discipline,
            "record": RECORD_LINKS.get(number),  # showModule id or None
        })
    print("  '%s' (hdr row %d): %d awarded" % (ws.title, hr, len(out)))
    return out


def main():
    print("PF Awarded Index build — %s UTC" % datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"))
    print("Authenticating + downloading bid log...")
    fp = download_bid_log()
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)

    awarded = []
    for sheet_name, discipline in SHEETS:
        if sheet_name not in wb.sheetnames:
            print("  WARN: sheet '%s' not present — skipping" % sheet_name)
            continue
        awarded.extend(extract_awarded(wb[sheet_name], discipline))

    payload = {
        "generated": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M") + " UTC",
        "source": "SharePoint/01 - Admin/13 - Master Spreadsheets/Project Bid Log.xlsx",
        "source_url": BID_LOG_WEBURL,
        "projects": awarded,
    }

    os.makedirs(DATA_DIR, exist_ok=True)
    out_path = os.path.join(DATA_DIR, "awarded-projects.js")
    with open(out_path, "w") as f:
        f.write("// AUTO-GENERATED by sync/build-awarded-index.py — do not edit by hand.\n")
        f.write("// Awarded rows from Project Bid Log (Agg Pier + Helical Pier sheets).\n")
        f.write("window.PF_AWARDED = ")
        json.dump(payload, f, indent=2)
        f.write(";\n")

    print("\nWrote %s (%d awarded projects)" % (out_path, len(awarded)))


if __name__ == "__main__":
    main()
