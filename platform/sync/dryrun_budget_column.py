#!/usr/bin/env python3
"""
DRY-RUN: Populate the portal "Budget" column from each project's Turnover Budget
=================================================================================
DISCOVERY + DRY-RUN ONLY. Writes NOTHING to KV, deploys NOTHING. Produces a
per-project report of exactly what WOULD be written to KV
`project_budget_v1:<num>` (rows keyed by the portal's stable row_key
`g<gi>_r<ri>`, value {budget, cost_code}) so Peter can review before greenlight.

Reuses the PROVEN pipeline already on this branch:
  - resolve_workbook()  (sync/build-budget-actuals.py) -> find each Turnover Budget
  - parse_budget_actual()  (sync/budget_actual_parser.py) -> read col C "Budget"
The ONLY new logic here is MAPPING each workbook leaf-budget onto the portal's
cost-code TEMPLATE row_key by (normalized group title + cost_code + description).

The portal render (index.html baRenderTable) keys the Budget column by
row_key = 'g'+groupIndex+'_r'+rowIndex over window.PF_COST_CODE_TEMPLATE.groups.
A saved budget (KV) WINS over the template zero. So the write target is the KV
rows map keyed by that same row_key.
"""
import os
for _v in ("OPENBLAS_NUM_THREADS", "OMP_NUM_THREADS", "MKL_NUM_THREADS",
           "NUMEXPR_NUM_THREADS", "VECLIB_MAXIMUM_THREADS"):
    os.environ.setdefault(_v, "1")

import re
import sys
import json
from datetime import datetime, timezone

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
sys.path.insert(0, "/home/aiciv/tools")

from pf_email import _token  # noqa: E402
import importlib.util as _ilu  # noqa: E402
_spec = _ilu.spec_from_file_location(
    "build_budget_actuals", os.path.join(HERE, "build-budget-actuals.py"))
bba = _ilu.module_from_spec(_spec)
_spec.loader.exec_module(bba)  # reuse resolver + graph helpers
from budget_actual_parser import parse_budget_actual, _norm_title  # noqa: E402

PLATFORM = os.path.dirname(HERE)
DATA_DIR = os.path.join(PLATFORM, "data")
TEMPLATE_JS = os.path.join(DATA_DIR, "cost-code-template.js")
BUDGET_SHEET = "Budget vs Actual"


def _norm_desc(s):
    s = (s or "").lower().strip()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def load_template():
    """Parse window.PF_COST_CODE_TEMPLATE out of the JS file -> the group/row tree
    with the SAME indices the portal uses to build row_key 'g<gi>_r<ri>'."""
    txt = open(TEMPLATE_JS).read()
    txt = txt.split("window.PF_COST_CODE_TEMPLATE = ", 1)[1]
    # strip trailing ';' + comments after the object
    depth = 0
    end = None
    for i, ch in enumerate(txt):
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                end = i + 1
                break
    obj = json.loads(txt[:end])
    return obj


def build_template_index(tpl):
    """Return a list of leaf template rows with their row_key and match keys.
    Each: {row_key, gi, ri, group_title, gnorm, code, desc, dnorm}.
    Only NON-subtotal rows that carry a cost_code are budget-writable targets
    (matches baRenderTable: 'if (!isSub && code)')."""
    leaves = []
    for gi, g in enumerate(tpl.get("groups", [])):
        gtitle = g.get("title", "")
        gnorm = _norm_title(gtitle)
        for ri, r in enumerate(g.get("rows", [])):
            if r.get("is_subtotal"):
                continue
            code = (r.get("cost_code") or "").strip()
            if not code:
                continue
            leaves.append({
                "row_key": f"g{gi}_r{ri}",
                "gi": gi, "ri": ri,
                "group_title": gtitle, "gnorm": gnorm,
                "code": code,
                "desc": r.get("description", ""),
                "dnorm": _norm_desc(r.get("description", "")),
            })
    return leaves


def probe_budget_cache(raw_bytes, sheet_name=BUDGET_SHEET):
    """Diagnose WHY a workbook's Budget column may be empty. Returns
    {formula_budget_cells, cached, literal_budget_cells, stale}.
    stale=True means col C is formula-driven but Excel never cached results
    (openpyxl can't evaluate) -> budgets read as None through NO fault of the
    parser. The fix is a one-time Excel open+save (or a formula evaluator), NOT
    a different file. Distinguishes this from a genuinely-blank budget column."""
    import io as _io
    import openpyxl  # thread caps already set at module import
    wbv = openpyxl.load_workbook(_io.BytesIO(raw_bytes), data_only=True, keep_vba=False)
    wbf = openpyxl.load_workbook(_io.BytesIO(raw_bytes), data_only=False, keep_vba=False)
    if sheet_name not in wbv.sheetnames:
        return {"formula_budget_cells": 0, "cached": 0, "literal_budget_cells": 0, "stale": False}
    wsv = wbv[sheet_name]; wsf = wbf[sheet_name]
    fcells = cached = literal = 0
    for r in range(1, min(wsv.max_row, 170) + 1):
        cf = wsf.cell(row=r, column=3).value
        cv = wsv.cell(row=r, column=3).value
        if isinstance(cf, str) and cf.startswith("="):
            fcells += 1
            if cv is not None:
                cached += 1
        elif isinstance(cv, (int, float)) and cv not in (0,):
            literal += 1
    # STALE = the Budget column is predominantly formula-driven but Excel cached
    # (almost) NONE of it. A stray literal 0 or a single cached cell must NOT mask a
    # workbook whose real budget (dozens of formula cells) never evaluated. Trigger
    # when there are many formula budget cells and effectively nothing cached.
    stale = (fcells >= 5 and cached == 0)
    return {"formula_budget_cells": fcells, "cached": cached,
            "literal_budget_cells": literal, "stale": stale}


def workbook_leaf_budgets(rec):
    """From a parsed Turnover-Budget record, yield every LEAF row that carries a
    cost_code AND a non-null budget (col C). Returns list of dicts:
    {gnorm, group_title, code, desc, dnorm, budget, vendor, notes}.

    vendor (col F) and notes (col G) are PLAIN TEXT (never formula-driven) so they
    read reliably even when col C's formula cache is stale — the parser already
    populated row['vendor']/row['notes']. We carry them here so the write step can
    place them on the matching template row_key alongside the budget."""
    out = []
    for g in rec.get("groups", []):
        gnorm = _norm_title(g.get("title", ""))
        for row in g.get("rows", []):
            if row.get("is_subtotal"):
                continue
            code = (row.get("cost_code") or "").strip()
            b = row.get("budget")
            if not code or b is None:
                continue
            out.append({
                "gnorm": gnorm, "group_title": g.get("title", ""),
                "code": code, "desc": row.get("description", ""),
                "dnorm": _norm_desc(row.get("description", "")),
                "budget": b,
                "vendor": (row.get("vendor") or ""),
                "notes": (row.get("notes") or ""),
                "actual": row.get("actual"),
            })
    return out


def workbook_leaf_meta(rec):
    """Like workbook_leaf_budgets, but returns EVERY code-bearing leaf that carries
    a vendor (col F), notes (col G), OR a per-row ACTUAL (col D) — REGARDLESS of
    whether col C (budget) has a cached value. This is the key to vendor/notes AND
    per-row actuals being INDEPENDENT of budget staleness: a job whose Budget column
    is a stale formula cache (e.g. 26-002 POET) still has plain-text F/G and
    arithmetic-recoverable col D that read fine, and we must place those onto the KV
    rows without touching the (already-correct) existing budget.

    per-row actual: the parser recovers col D from its arithmetic formula when the
    cache is stale (invoice-coding sums like '=1050+3230'), so row['actual'] is
    trustworthy per row. Storing it per row_key breaks out multi-row cost codes
    (e.g. 5405 VSC Rig / Predrill / Fall Off) so each sub-line's actual aligns with
    its own budget — instead of the daemon feed's (group,code) LUMP.
    Returns list of {gnorm, group_title, code, desc, dnorm, vendor, notes, actual}."""
    out = []
    for g in rec.get("groups", []):
        gnorm = _norm_title(g.get("title", ""))
        for row in g.get("rows", []):
            if row.get("is_subtotal"):
                continue
            code = (row.get("cost_code") or "").strip()
            if not code:
                continue
            vendor = (row.get("vendor") or "").strip()
            notes = (row.get("notes") or "").strip()
            actual = row.get("actual")
            if not vendor and not notes and actual is None:
                continue  # nothing to place for this leaf
            out.append({
                "gnorm": gnorm, "group_title": g.get("title", ""),
                "code": code, "desc": row.get("description", ""),
                "dnorm": _norm_desc(row.get("description", "")),
                "vendor": vendor, "notes": notes, "actual": actual,
            })
    return out


def map_meta(leaves, wb_meta):
    """Match each workbook vendor/notes/actual row to ONE template row_key, using the
    SAME tiered matching as map_budgets (each target row_key consumed once). Returns
    {row_key: {vendor, notes, actual?, cost_code}} — the vendor/notes/actual write
    payload. Budget is intentionally ABSENT so the writer never touches an existing
    budget. 'actual' is included only when the workbook row had a (recovered) value."""
    by_gcd = {}
    by_cd = {}
    by_gc = {}
    by_c = {}
    code_counts = {}
    for lf in leaves:
        by_gcd.setdefault((lf["gnorm"], lf["code"], lf["dnorm"]), []).append(lf)
        by_cd.setdefault((lf["code"], lf["dnorm"]), []).append(lf)
        by_gc.setdefault((lf["gnorm"], lf["code"]), []).append(lf)
        by_c.setdefault(lf["code"], []).append(lf)
        code_counts[lf["code"]] = code_counts.get(lf["code"], 0) + 1

    used = set()
    matched = {}

    def take(cands):
        for c in cands:
            if c["row_key"] not in used:
                return c
        return None

    for wb in wb_meta:
        cand = take(by_gcd.get((wb["gnorm"], wb["code"], wb["dnorm"]), []))
        if not cand:
            cand = take(by_cd.get((wb["code"], wb["dnorm"]), []))
        if not cand:
            cand = take(by_gc.get((wb["gnorm"], wb["code"]), []))
        if not cand and code_counts.get(wb["code"], 0) == 1:
            cand = take(by_c.get(wb["code"], []))
        if cand:
            used.add(cand["row_key"])
            entry = {"vendor": wb.get("vendor", ""),
                     "notes": wb.get("notes", ""),
                     "cost_code": wb["code"]}
            if wb.get("actual") is not None:
                entry["actual"] = round(float(wb["actual"]), 2)
            matched[cand["row_key"]] = entry
    return matched


def map_budgets(leaves, wb_budgets):
    """Match each workbook budget row to ONE template row_key.
    Match tiers (strongest first), each target row_key consumed once:
      T1: (gnorm, code, dnorm) exact
      T2: (code, dnorm) exact  (group title drift tolerated)
      T3: (gnorm, code) exact  (description drift tolerated)
      T4: (code) exact         (last resort; only if code is UNIQUE in template)
    Returns (matched_rows, unmatched_wb, matched_keys)
      matched_rows: {row_key: {budget, cost_code, vendor, notes}}  <- KV write payload
      unmatched_wb: [ {code, desc, group_title, budget, reason} ]
    """
    # Index template leaves by the tier keys.
    by_gcd = {}
    by_cd = {}
    by_gc = {}
    by_c = {}
    code_counts = {}
    for lf in leaves:
        by_gcd.setdefault((lf["gnorm"], lf["code"], lf["dnorm"]), []).append(lf)
        by_cd.setdefault((lf["code"], lf["dnorm"]), []).append(lf)
        by_gc.setdefault((lf["gnorm"], lf["code"]), []).append(lf)
        by_c.setdefault(lf["code"], []).append(lf)
        code_counts[lf["code"]] = code_counts.get(lf["code"], 0) + 1

    used = set()
    matched = {}
    unmatched = []
    audit = []  # (code, desc, tier, row_key)

    def take(cands):
        for c in cands:
            if c["row_key"] not in used:
                return c
        return None

    for wb in wb_budgets:
        cand = None
        tier = None
        c = take(by_gcd.get((wb["gnorm"], wb["code"], wb["dnorm"]), []))
        if c:
            cand, tier = c, "T1 g+code+desc"
        if not cand:
            c = take(by_cd.get((wb["code"], wb["dnorm"]), []))
            if c:
                cand, tier = c, "T2 code+desc"
        if not cand:
            c = take(by_gc.get((wb["gnorm"], wb["code"]), []))
            if c:
                cand, tier = c, "T3 g+code"
        if not cand and code_counts.get(wb["code"], 0) == 1:
            c = take(by_c.get(wb["code"], []))
            if c:
                cand, tier = c, "T4 code-unique"
        if cand:
            used.add(cand["row_key"])
            entry = {"budget": round(wb["budget"], 2),
                     "cost_code": wb["code"],
                     "vendor": wb.get("vendor", ""),
                     "notes": wb.get("notes", "")}
            if wb.get("actual") is not None:
                entry["actual"] = round(float(wb["actual"]), 2)
            matched[cand["row_key"]] = entry
            audit.append((wb["code"], wb["desc"], tier, cand["row_key"]))
        else:
            reason = ("code-not-in-template" if wb["code"] not in by_c
                      else "no-uniquely-matchable-template-row")
            unmatched.append({"code": wb["code"], "desc": wb["desc"],
                              "group_title": wb["group_title"],
                              "budget": round(wb["budget"], 2), "reason": reason})
    return matched, unmatched, audit


def _m(x):
    return "—" if x is None else f"${x:,.2f}"


def main():
    args = sys.argv[1:]
    only = None
    if "--only" in args:
        only = set()
        i = 0
        while i < len(args):
            if args[i] == "--only" and i + 1 < len(args):
                only.add(args[i + 1]); i += 2
            else:
                i += 1
    active_only = "--active-only" in args
    active_set = None
    if active_only:
        # Optional: pass an explicit active list via --active 26-013,26-017,...
        for i, a in enumerate(args):
            if a == "--active" and i + 1 < len(args):
                active_set = set(x.strip() for x in args[i + 1].split(","))

    if not bba.DRIVE_ID:
        print("ERROR: SP_DRIVE_ID not set", file=sys.stderr); sys.exit(1)

    token = _token()
    tpl = load_template()
    leaves = build_template_index(tpl)
    print(f"# Template: {len(tpl.get('groups', []))} groups, "
          f"{len(leaves)} budget-writable leaf rows (code-bearing).", file=sys.stderr)

    manifest = bba.load_manifest()
    jobs = bba.load_jobs(token)
    if active_set:
        jobs = [(j, n) for (j, n) in jobs if j in active_set]
    if only:
        jobs = [(j, n) for (j, n) in jobs if j in only]

    report = []
    clean = 0
    flagged = 0
    for job, name in jobs:
        print(f"  resolving {job} {name}...", file=sys.stderr)
        folder, wbpath, src, status = bba.resolve_workbook(token, job, manifest)
        entry = {"job": job, "name": name, "status": status,
                 "folder": folder, "file": wbpath, "source_kind": src}
        if status != "ok" or not wbpath:
            entry["flags"] = [f"workbook-{status}"]
            report.append(entry)
            flagged += 1
            continue
        try:
            raw = bba.download_path(token, wbpath)
            rec = parse_budget_actual(raw, sheet_name=BUDGET_SHEET,
                                      default_job=job, default_name=name)
        except Exception as e:  # noqa: BLE001
            entry["flags"] = [f"parse-error: {str(e)[:120]}"]
            report.append(entry)
            flagged += 1
            continue

        wb_budgets = workbook_leaf_budgets(rec)
        matched, unmatched, audit = map_budgets(leaves, wb_budgets)
        # Vendor/notes payload — mapped INDEPENDENT of budget staleness (plain-text
        # F/G read fine even when col C's formula cache is stale). Produced for every
        # parsed job, including budget-stale/non-writable ones (e.g. 26-002 POET).
        wb_meta = workbook_leaf_meta(rec)
        meta_rows = map_meta(leaves, wb_meta)
        gt = rec.get("grand_total", {})
        cache = probe_budget_cache(raw)
        entry.update({
            "budget_column_header": "col C — row3 'Precon' / row4 'Budget' (leaf rows)",
            "has_cached_values": rec.get("has_cached_values"),
            "budget_cache": cache,
            "location": ("active" if src == "active" else "completed"),
            "wb_budget_rows": len(wb_budgets),
            "matched": len(matched),
            "unmatched": len(unmatched),
            "unmatched_rows": unmatched,
            "grand_total_budget": gt.get("budget"),
            "sum_matched_budget": round(sum(v["budget"] for v in matched.values()), 2),
            "sample": audit[:6],
            "kv_key": f"project_budget_v1:{job}",
            "kv_payload_rows": matched,   # budget+vendor+notes (only for cached-budget rows)
            "kv_meta_rows": meta_rows,    # vendor/notes ONLY (budget-independent; all F/G leaves)
            "meta_rows_count": len(meta_rows),
        })
        # BLOCKING flags = would prevent/ruin a write. Completed-location is NO
        # LONGER a flag (completed jobs are IN scope per Brad's 90-day rule).
        flags = []
        writable = True
        if cache.get("stale"):
            flags.append("STALE-CACHE: Budget col is formula-driven but workbook has "
                         "0 cached values (openpyxl can't evaluate). Needs one-time "
                         "Excel open+save (or a formula evaluator) before it can be "
                         "read. NOT a wrong-file issue — only one workbook in folder.")
            writable = False
        elif not rec.get("has_cached_values"):
            flags.append("no-cached-values (workbook has no evaluated budget/actual cells)")
            writable = False
        if len(matched) == 0:
            flags.append("ZERO budget rows to write")
            writable = False
        elif entry["sum_matched_budget"] == 0:
            flags.append("ALL matched budgets are $0 (nothing meaningful to write) — "
                         "likely an unfilled/uncached Budget column")
            writable = False
        if unmatched:
            flags.append(f"{len(unmatched)} unmatched workbook budget rows (surfaced, not dropped)")
        entry["flags"] = flags
        entry["writable"] = writable
        report.append(entry)
        if entry.get("writable"):
            clean += 1
        else:
            flagged += 1

    # ---- buckets ----
    writable = [e for e in report if e.get("writable")]
    stale = [e for e in report if e["status"] == "ok" and e.get("budget_cache", {}).get("stale")]
    noworkbook = [e for e in report if e["status"] != "ok"]

    # ---- human summary to stderr ----
    print("\n" + "=" * 82, file=sys.stderr)
    print("DRY-RUN SUMMARY — Budget column population (ALL jobs w/ workbook)", file=sys.stderr)
    print("=" * 82, file=sys.stderr)
    print(f"\n  WRITABLE ({len(writable)}) — would populate KV project_budget_v1:<num>:", file=sys.stderr)
    for e in sorted(writable, key=lambda x: x["job"]):
        loc = e.get("location", "")
        print(f"    {e['job']:<9} {e['name'][:24]:<24} [{loc:<9}] "
              f"matched={e['matched']:>2}/{e['wb_budget_rows']:>2} "
              f"unmatched={e['unmatched']:>2}  GTbudget={_m(e['grand_total_budget'])} "
              f"sumMatched={_m(e['sum_matched_budget'])}", file=sys.stderr)
    if stale:
        print(f"\n  STALE-CACHE ({len(stale)}) — real file, but Budget formulas uncached "
              f"(need Excel open+save):", file=sys.stderr)
        for e in sorted(stale, key=lambda x: x["job"]):
            bc = e["budget_cache"]
            print(f"    {e['job']:<9} {e['name'][:24]:<24} "
                  f"[{e.get('file','').split('/')[-1]}] "
                  f"formula-cells={bc['formula_budget_cells']} cached={bc['cached']}",
                  file=sys.stderr)
    if noworkbook:
        print(f"\n  NO-WORKBOOK ({len(noworkbook)}) — no Turnover Budget file exists yet:", file=sys.stderr)
        for e in sorted(noworkbook, key=lambda x: x["job"]):
            print(f"    {e['job']:<9} {e['name'][:24]:<24} {e['status']}", file=sys.stderr)
    print(f"\n  {clean} writable / {flagged} not-writable  (of {len(report)} jobs)", file=sys.stderr)

    out = {
        "generated": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "mode": "DRY-RUN (no KV write, no deploy)",
        "template_leaf_rows": len(leaves),
        "writable": clean, "not_writable": flagged, "jobs": len(report),
        "buckets": {
            "writable": [e["job"] for e in sorted(writable, key=lambda x: x["job"])],
            "stale_cache": [e["job"] for e in sorted(stale, key=lambda x: x["job"])],
            "no_workbook": [e["job"] for e in sorted(noworkbook, key=lambda x: x["job"])],
        },
        "projects": report,
    }
    outpath = os.path.join(HERE, "dryrun-budget-column-report.json")
    with open(outpath, "w") as f:
        json.dump(out, f, indent=2)
    print(f"\n  Full JSON report: {outpath}", file=sys.stderr)


if __name__ == "__main__":
    main()
