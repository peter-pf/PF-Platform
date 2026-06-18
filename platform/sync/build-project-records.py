#!/usr/bin/env python3
"""
Build Project Records — ALL awarded projects  [Phase 2, generalized, READ-ONLY]
================================================================================
Generalizes the POET-only build-project-record.py to EVERY awarded project
(Aggregate Piers + Helical). For each awarded project it produces a record with
the SAME 11-section schema POET uses, populated RESILIENTLY from the data we
already pull. Missing pieces stay blank; a project missing a folder / contacts
file / bid-log row still produces a valid record with what's known.

Output: a SINGLE keyed file  data/project-records.js
        window.PF_PROJECT_RECORDS = { "26-013": {...}, "26-007": {...}, ... }
keyed by project number. (Chosen over per-file: one <script> tag, one fetch,
the renderer just indexes by number — cleaner than 20 separate files/tags.)

POET (26-002) is LEFT UNTOUCHED — its richer record (subcontract extraction +
live contract embed) stays in data/project-record-poet.js (window.PF_PROJECT_POET).
This builder SKIPS POET by default so its deep record is never overwritten. The
renderer checks PF_PROJECT_RECORDS first, then falls back to PF_PROJECT_POET.

PER PROJECT, this populates:
  - General Info + Contract Info basics  <- Bid Log row (matched by project number
    across BOTH the Agg Pier and Helical Pier sheets).
  - PF Team + contacts                   <- that project's "...Project Info.xlsx"
    (sheet "...Project Contacts") IF present in the folder. Many won't have it
    fully filled — that's fine, left minimal.
  - Folder doc links                     <- Graph webUrl for the standard subfolders
    that ACTUALLY EXIST (Engineering & Design, Subcontract Agreement, Invoicing,
    Field, GC Drawings, QAQC). Only existing folders are linked.
  - QA/QC progress                       <- data/progress-data.js if the project has
    a key there (installed cols/LF, %).

NOT done here (POET-only / per-job on-demand): per-project subcontract PDF
extraction. We only LINK the Subcontract Agreement folder. As Brad forwards each
contract, the POET-style extractor is run per job.

Reuses the Graph auth + drive + contact-parse patterns from build-project-record.py.
Thread caps set BEFORE openpyxl import (Box ~300 pid guard; solved-box-gotchas).

Usage:
  python3 build-project-records.py --all            # all awarded (skips POET) -> data/project-records.js
  python3 build-project-records.py --project 26-013  # one project (merges into the keyed file)
  python3 build-project-records.py --all --dump      # print summary table only, no write
"""

# --- thread caps BEFORE any heavy import (solved-box-gotchas) ---
import os
for _v in ("OPENBLAS_NUM_THREADS", "OMP_NUM_THREADS", "MKL_NUM_THREADS",
           "NUMEXPR_NUM_THREADS", "VECLIB_MAXIMUM_THREADS"):
    os.environ.setdefault(_v, "1")

import io
import re
import sys
import json
import argparse
import urllib.request
import urllib.parse
import urllib.error
from datetime import datetime, timezone

# --- reuse the Graph token helper + env loader from pf_email.py ---
sys.path.insert(0, "/home/aiciv/tools")
from pf_email import _token, _env  # noqa: E402

GRAPH = "https://graph.microsoft.com/v1.0"

HERE = os.path.dirname(os.path.abspath(__file__))
PLATFORM = os.path.dirname(HERE)
DATA_DIR = os.path.join(PLATFORM, "data")
OUT_JS = os.path.join(DATA_DIR, "project-records.js")
PROGRESS_JS = os.path.join(DATA_DIR, "progress-data.js")
AWARDED_JS = os.path.join(DATA_DIR, "awarded-projects.js")
PRECON_JS = os.path.join(DATA_DIR, "precon-pipeline.js")

PROJECTS_ROOT = "04 - Project Management/02 - Projects"

# POET is built by the dedicated build-project-record.py (deep record). Never overwrite it here.
POET_NUMBER = "26-002"

# Bid Log workbook (same source build-awarded-index.py / build-project-record.py use).
BID_LOG_FILE = "01 - Admin/13 - Master Spreadsheets/Project Bid Log.xlsx"
BID_SHEETS = ["Agg Pier Bid Log", "Helical Pier Bid Log"]

_env_cache = _env()
DRIVE_ID = _env_cache.get("SP_DRIVE_ID", "")


# ---------------- Graph helpers (pattern from build-project-record.py) ----------------
def gget(token, url):
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    return json.loads(urllib.request.urlopen(req).read())


def list_children_by_path(token, path):
    p = urllib.parse.quote(path)
    url = f"{GRAPH}/drives/{DRIVE_ID}/root:/{p}:/children"
    items = []
    try:
        while url:
            data = gget(token, url)
            items.extend(data.get("value", []))
            url = data.get("@odata.nextLink")
    except urllib.error.HTTPError:
        return []
    return items


def get_item_by_path(token, path):
    """driveItem metadata (incl webUrl) for a folder/file path, or None."""
    p = urllib.parse.quote(path)
    url = f"{GRAPH}/drives/{DRIVE_ID}/root:/{p}"
    try:
        return gget(token, url)
    except urllib.error.HTTPError:
        return None


def download_path(token, path):
    p = urllib.parse.quote(path)
    url = f"{GRAPH}/drives/{DRIVE_ID}/root:/{p}:/content"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    return urllib.request.urlopen(req).read()


# ---------------- project folder discovery (match by NN-NNN prefix) ----------------
def list_project_folders(token):
    """Return {project_number: folder_name} for every '04.../02 - Projects/NN-NNN - ...' folder."""
    out = {}
    for item in list_children_by_path(token, PROJECTS_ROOT):
        if not item.get("folder"):
            continue
        name = item["name"]
        m = re.match(r"^(\d{2}-\d{3,4})\b", name.strip())
        if m:
            out[m.group(1)] = name
    return out


# ---------------- cleaning helpers (from build-project-record.py) ----------------
def _clean(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s.upper() in ("N/A", "NA", "TBD"):
        return ""
    return s


def _xlclean(v):
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        try:
            return v.strftime("%Y-%m-%d")
        except Exception:
            return str(v)
    s = str(v).strip()
    if s.upper() in ("N/A", "NA", "TBD"):
        return ""
    if s.endswith(" 00:00:00"):
        s = s[:-9]
    return s


# ---------------- contacts parse (same shape as POET builder) ----------------
def _contact(ws, r):
    return {
        "scope": _clean(ws.cell(row=r, column=2).value),
        "company": _clean(ws.cell(row=r, column=3).value),
        "name": _clean(ws.cell(row=r, column=4).value),
        "address": _clean(ws.cell(row=r, column=5).value),
        "phone": _clean(ws.cell(row=r, column=6).value),
        "cell": _clean(ws.cell(row=r, column=7).value),
        "email": _clean(ws.cell(row=r, column=8).value),
        "website": _clean(ws.cell(row=r, column=9).value),
        "notes": _clean(ws.cell(row=r, column=12).value),
    }


def _has_contact(c):
    return bool(c["company"] or c["name"] or c["email"])


SECTION_MAP = [
    (re.compile(r"\bowner\b", re.I), "owner"),
    (re.compile(r"general contractor", re.I), "gc"),
    (re.compile(r"engineering", re.I), "engineering"),
    (re.compile(r"pier foundations", re.I), "pf_team"),
    (re.compile(r"vendor", re.I), "vendors"),
]


def find_contacts_file(token, folder_name):
    """Locate the project's '...Project Info.xlsx' in the project folder root.
    Returns (path, filename) or (None, None)."""
    base = f"{PROJECTS_ROOT}/{folder_name}"
    for item in list_children_by_path(token, base):
        n = item.get("name", "")
        if not item.get("folder") and n.lower().endswith(".xlsx") and "project info" in n.lower():
            return f"{base}/{n}", n
    return None, None


def parse_contacts(token, folder_name):
    """Read the project contact log if present. Returns grouped dict or None."""
    import openpyxl
    path, fname = find_contacts_file(token, folder_name)
    if not path:
        return None
    try:
        raw = download_path(token, path)
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception:
        return None

    # Prefer a sheet whose name mentions "Contacts"; else first sheet.
    ws = None
    for sn in wb.sheetnames:
        if "contact" in sn.lower():
            ws = wb[sn]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    groups = {"owner": [], "gc": [], "engineering": [], "pf_team": [], "vendors": []}
    current = None
    title = _clean(ws.cell(row=1, column=1).value)
    for r in range(2, ws.max_row + 1):
        a = _clean(ws.cell(row=r, column=1).value)
        if a:
            matched = None
            for rx, key in SECTION_MAP:
                if rx.search(a):
                    matched = key
                    break
            if matched:
                current = matched
            continue
        if current is None:
            continue
        c = _contact(ws, r)
        if _has_contact(c):
            groups[current].append(c)

    return {"title": title, "groups": groups, "source_file": fname}


# ---------------- bid log (General Info + Contract Info + award) ----------------
def _find_header_row(ws, scan=14):
    for r in range(1, scan + 1):
        joined = " ".join(str(ws.cell(row=r, column=c).value or "") for c in range(1, 40))
        if "Bid Status" in joined and "Project Name" in joined:
            return r
    return None


def _header_map(ws, hr):
    m = {}
    for c in range(1, 40):
        v = ws.cell(row=hr, column=c).value
        if v not in (None, ""):
            m[str(v).strip()] = c
    return m


def load_bid_index(token):
    """Build {project_number: bidrow_dict} from BOTH bid-log sheets, located by header name.
    Resilient to the two layouts (Agg Pier vs Helical) by mapping columns BY HEADER TEXT."""
    import openpyxl
    raw = download_path(token, BID_LOG_FILE)
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    index = {}
    for sheet in BID_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hr = _find_header_row(ws)
        if not hr:
            continue
        h = _header_map(ws, hr)

        def col(*names):
            for n in names:
                if n in h:
                    return h[n]
            return None

        c = {
            "number": col("Project Number"),
            "name": col("Project Name"),
            "city_state": col("City / State"),
            "address": col("Address"),
            "site_acres": col("Site Size (Acres)"),
            "total_sf": col("Total SF (Bldg Pad)", "Total SF"),
            "total_lf": col("Total LF"),
            "total_columns": col("Total Columns", "Total Piles", "Total Piers"),
            "total_stone_tn": col("Total Stone (TN)", "Total Stone"),
            "duration_days": col("Project Duration (Days)", "Duration (Days)"),
            "feed_type": col("Top vs Bottom Feed"),
            "tax": col("Tax"),
            "min_insurance_req": col("Min Insurance Req", "Minimum Insurance Req"),
            "wage_req": col("Wage Req.", "Wage Req"),
            "invite_date": col("Invite Date"),
            "due_date": col("Due Date"),
            "date_submitted": col("Date Submitted"),
            "projected_start": col("Projected Start Date", "Projected Start"),
            "follow_up_date": col("Follow Up Date"),
            "bid_status": col("Bid Status"),
            "bid_total_value": col("Bid Total Value"),
            "gc_name": col("General Contractor"),
            "gc_contact": col("Contact Name"),
            "gc_email": col("GC Email"),
            "gc_phone": col("GC Phone"),
            "engineer_firm": col("Engineer / Design Firm", "Engineer/Design Firm"),
            "prelim_design_fee": col("Prelim Design Fee"),
            "design_completed_date": col("Design Completed Date"),
            "design_paid_date": col("Date Paid"),
        }
        if not (c["number"] and c["name"]):
            continue

        for r in range(hr + 1, ws.max_row + 1):
            pn = ws.cell(row=r, column=c["number"]).value
            pn = str(pn).strip() if pn is not None else ""
            if not pn:
                continue
            row = {"row": r, "source_file": BID_LOG_FILE, "source_sheet": sheet}
            for key, cc in c.items():
                if cc:
                    row[key] = _xlclean(ws.cell(row=r, column=cc).value)
            # first sheet wins if a number appears twice (shouldn't)
            index.setdefault(pn, row)
    return index


# ---------------- section document links (standard subfolders) ----------------
# Standard per-project folder layout (confirmed across 26-007/009/013):
#   01 - Preconstruction / Vendor Quotes
#   02 - Project Management / {Invoicing, Subcontract Agreement, Vendors}
#   03 - Engineering & Design / {Approved Shop Dwgs, Geotech Report, Modulus Test, QAQC, Stamped Drawings, ...}
#   04 - GC Drawings & Specs
#   05 - Field / {02 - Project Photos, 03 - Drawings, 04 - Testing, 05 - Approved Materials, 06 - Safety}
# Only folders that ACTUALLY EXIST become links (found:false otherwise).
def section_link_spec(folder_name):
    base = f"{PROJECTS_ROOT}/{folder_name}"
    return {
        "general": [
            ("GC Drawings & Specs", f"{base}/04 - GC Drawings & Specs"),
            ("Project Folder (root)", base),
        ],
        "contract": [
            ("Subcontract Agreement", f"{base}/02 - Project Management/Subcontract Agreement"),
        ],
        "engineering": [
            ("Approved Shop Dwgs", f"{base}/03 - Engineering & Design/Approved Shop Dwgs"),
            ("Bid Prelim", f"{base}/03 - Engineering & Design/Bid Prelim"),
            ("Geotech Report", f"{base}/03 - Engineering & Design/Geotech Report"),
            ("Modulus Test", f"{base}/03 - Engineering & Design/Modulus Test"),
            ("Stamped Drawings", f"{base}/03 - Engineering & Design/Stamped Drawings"),
            ("QAQC", f"{base}/03 - Engineering & Design/QAQC"),
        ],
        "safety": [
            ("Field — Safety", f"{base}/05 - Field/06 - Safety"),
        ],
        "site": [
            ("Field — Drawings", f"{base}/05 - Field/03 - Drawings"),
            ("Field — Testing", f"{base}/05 - Field/04 - Testing"),
            ("Field — Approved Materials", f"{base}/05 - Field/05 - Approved Materials"),
            ("Field — Project Photos", f"{base}/05 - Field/02 - Project Photos"),
        ],
        "financials": [
            ("Invoicing", f"{base}/02 - Project Management/Invoicing"),
        ],
        "qaqc": [
            ("QAQC Logs (GUHMA)", f"{base}/03 - Engineering & Design/QAQC"),
            ("Modulus Test", f"{base}/03 - Engineering & Design/Modulus Test"),
        ],
        "material": [
            ("Field — Approved Materials", f"{base}/05 - Field/05 - Approved Materials"),
            ("Vendor Quotes", f"{base}/01 - Preconstruction/Vendor Quotes"),
        ],
    }


def resolve_links(token, folder_name):
    """Resolve each section link to a Graph webUrl, only for folders that exist."""
    out = {}
    spec = section_link_spec(folder_name)
    for section, entries in spec.items():
        resolved = []
        for label, path in entries:
            item = get_item_by_path(token, path)
            if item and item.get("webUrl"):
                resolved.append({"label": label, "url": item["webUrl"], "found": True})
            else:
                resolved.append({"label": label, "url": "", "found": False})
        out[section] = resolved
    return out


# ---------------- QA/QC from progress-data.js ----------------
def load_progress_index():
    if not os.path.exists(PROGRESS_JS):
        return {}
    with open(PROGRESS_JS) as f:
        txt = f.read()
    m = re.search(r"window\.PF_PROGRESS\s*=\s*(\{.*\});", txt, re.S)
    if not m:
        return {}
    try:
        data = json.loads(m.group(1))
    except json.JSONDecodeError:
        return {}
    return data.get("projects") or {}


# ---------------- awarded list (the projects to build) ----------------
def load_awarded_numbers():
    """Return ordered list of (number, name, discipline) for awarded projects.
    Source of truth: data/awarded-projects.js (built by build-awarded-index.py from
    the Bid Log awarded sections). Falls back to precon-pipeline if awarded file
    is missing."""
    out = []
    if os.path.exists(AWARDED_JS):
        with open(AWARDED_JS) as f:
            txt = f.read()
        m = re.search(r"window\.PF_AWARDED\s*=\s*(\{.*\});", txt, re.S)
        if m:
            try:
                data = json.loads(m.group(1))
                for p in data.get("projects", []):
                    num = str(p.get("number") or "").strip()
                    if num:
                        out.append((num, p.get("name", ""), p.get("discipline", "")))
                return out
            except json.JSONDecodeError:
                pass
    # fallback: precon-pipeline awarded buckets
    if os.path.exists(PRECON_JS):
        with open(PRECON_JS) as f:
            txt = f.read()
        m = re.search(r"window\.PF_PRECON\s*=\s*(\{.*\});", txt, re.S)
        if m:
            d = json.loads(m.group(1))
            disc_label = {"ap": "Aggregate Piers", "hp": "Helical Pilings"}
            for disc in ("ap", "hp"):
                for p in d.get(disc, {}).get("awarded", []):
                    num = str(p.get("number") or "").strip()
                    if num:
                        out.append((num, p.get("name", ""), disc_label[disc]))
    return out


# ---------------- per-project assembly ----------------
def assemble_project(token, number, disc_label, folders, bid_index, progress_index):
    """Build one project record dict resiliently. Returns (record, populate_flags)."""
    folder_name = folders.get(number)
    bid = bid_index.get(number)
    progress = progress_index.get(number)

    contacts = None
    links = {}
    contacts_file = None
    folder_link_count = 0
    if folder_name:
        contacts = parse_contacts(token, folder_name)
        if contacts:
            contacts_file = contacts.get("source_file")
        links = resolve_links(token, folder_name)
        folder_link_count = sum(
            1 for lst in links.values() for x in lst if x.get("found")
        )

    # name: prefer Bid Log name, then awarded-list name
    name = (bid or {}).get("name") or ""
    location = (bid or {}).get("city_state") or ""

    record = {
        "project_number": number,
        "project_name": name,
        "location": location,
        "discipline": disc_label,
        "sp_folder": (f"{PROJECTS_ROOT}/{folder_name}" if folder_name else ""),
        "folder_found": bool(folder_name),
        "generated": datetime.now(timezone.utc).isoformat() + "Z",
        "data_note": ("READ-ONLY. Bulk-populated from data we already hold: the Project "
                      "Bid Log (General Info metrics, Contract Info dates, award value, "
                      "GC/engineer), the project's Project Info contacts file (if present), "
                      "folder doc links (folders that exist), and the auto-progress engine "
                      "(QA/QC installed qty). Per-job subcontract extraction is run separately "
                      "as each contract is forwarded (POET is the reference deep record). "
                      "Fields with no confirmed source render blank — never fabricated."),
        "contacts": contacts or {"title": "", "groups": {"owner": [], "gc": [], "engineering": [], "pf_team": [], "vendors": []}, "source_file": None},
        "links": links,
        "qaqc": progress,
        "bid_log": bid,
        "subcontract": None,   # POET-only / per-job on-demand; folder is linked under contract
        "discrepancies": [],
    }

    flags = {
        "number": number,
        "name": name or (folders.get(number) or "(unknown)"),
        "folder": "Y" if folder_name else "N",
        "bidlog": "Y" if bid else "N",
        "contacts": "Y" if (contacts and any(contacts["groups"].values())) else "N",
        "contacts_file": "Y" if contacts_file else "N",
        "links": folder_link_count,
        "progress": "Y" if progress else "N",
    }
    return record, flags


# ---------------- write / merge keyed file ----------------
def read_existing_records():
    """Load the existing window.PF_PROJECT_RECORDS map for merge (single-project mode)."""
    if not os.path.exists(OUT_JS):
        return {}
    with open(OUT_JS) as f:
        txt = f.read()
    m = re.search(r"window\.PF_PROJECT_RECORDS\s*=\s*(\{.*\});", txt, re.S)
    if not m:
        return {}
    try:
        data = json.loads(m.group(1))
    except json.JSONDecodeError:
        return {}
    return data.get("records") or {}


def write_records(records_map):
    os.makedirs(DATA_DIR, exist_ok=True)
    payload = {
        "generated": datetime.now(timezone.utc).isoformat() + "Z",
        "note": ("All awarded project records EXCEPT POET (26-002), which keeps its deeper "
                 "dedicated record in project-record-poet.js. Keyed by project number. "
                 "The project-record renderer checks this map first, then PF_PROJECT_POET."),
        "records": records_map,
    }
    with open(OUT_JS, "w") as f:
        f.write("// AUTO-GENERATED by sync/build-project-records.py — do not edit by hand.\n")
        f.write("// All awarded project records (except POET, which is project-record-poet.js).\n")
        f.write("// window.PF_PROJECT_RECORDS keyed by project number.\n")
        f.write("window.PF_PROJECT_RECORDS = ")
        json.dump(payload, f, indent=2)
        f.write(";\n")
    return OUT_JS


# ---------------- main ----------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--all", action="store_true", help="build all awarded projects (skips POET)")
    ap.add_argument("--project", help="build a single project by number, e.g. 26-013")
    ap.add_argument("--include-poet", action="store_true",
                    help="(advanced) also build a basic POET record into the keyed file; "
                         "off by default so the deep POET record is never shadowed")
    ap.add_argument("--dump", action="store_true", help="print summary only, do not write")
    args = ap.parse_args()

    if not DRIVE_ID:
        print("ERROR: SP_DRIVE_ID not set in /home/aiciv/.env", file=sys.stderr)
        sys.exit(1)
    if not (args.all or args.project):
        ap.error("specify --all or --project NN-NNN")

    print(f"build-project-records — {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC",
          file=sys.stderr)
    token = _token()

    print("Loading folder list, bid log, progress, awarded list...", file=sys.stderr)
    folders = list_project_folders(token)
    bid_index = load_bid_index(token)
    progress_index = load_progress_index()
    awarded = load_awarded_numbers()
    awarded_disc = {num: disc for num, _name, disc in awarded}

    # Decide target set
    if args.project:
        targets = [args.project]
    else:
        targets = [num for num, _n, _d in awarded]

    # de-dup, preserve order; drop POET unless explicitly included
    seen = set()
    ordered = []
    for num in targets:
        if num in seen:
            continue
        seen.add(num)
        if num == POET_NUMBER and not args.include_poet:
            continue
        ordered.append(num)

    records_map = {} if args.all else read_existing_records()
    flags_rows = []
    for num in ordered:
        disc = awarded_disc.get(num, "")
        rec, flags = assemble_project(token, num, disc, folders, bid_index, progress_index)
        records_map[num] = rec
        flags_rows.append(flags)

    # ---- summary table ----
    print("\nPer-project populate:")
    print(f"  {'Number':<8} {'Fld':<3} {'Bid':<3} {'Con':<3} {'CF':<3} {'Lnk':<4} {'Prg':<3}  Name")
    for f in flags_rows:
        print(f"  {f['number']:<8} {f['folder']:<3} {f['bidlog']:<3} {f['contacts']:<3} "
              f"{f['contacts_file']:<3} {str(f['links']):<4} {f['progress']:<3}  {f['name']}")
    print(f"\n  Legend: Fld=folder Bid=bidlog row Con=contacts present CF=contacts file "
          f"Lnk=folder links resolved Prg=QA/QC progress")
    print(f"  Built {len(flags_rows)} record(s). POET {'INCLUDED' if args.include_poet else 'skipped (deep record preserved)'}.")

    if args.dump:
        print("\n[--dump] no file written.", file=sys.stderr)
        return

    out = write_records(records_map)
    print(f"\nWrote: {out}  ({len(records_map)} records total in file)", file=sys.stderr)


if __name__ == "__main__":
    main()
