#!/usr/bin/env python3
"""Unit harness for build-turnover-stone.extract_stone() against SYNTHETIC in-memory
Turnover Budget workbooks (no Graph / no network). Proves the label-anchored read, the
case/space-insensitive 'Detailed Budget' tab match, and the fail-closed formula-no-cache
guard (never fabricate a stale number). Run: python3 test-turnover-stone-extract.py"""
import os
for _v in ("OPENBLAS_NUM_THREADS", "OMP_NUM_THREADS"):
    os.environ.setdefault(_v, "1")
import io
import importlib.util
import openpyxl

_here = os.path.dirname(os.path.abspath(__file__))
_spec = importlib.util.spec_from_file_location("bts", os.path.join(_here, "build-turnover-stone.py"))
bts = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(bts)


def mk(tabname, label_cell, label, val_cell, val, unit_cell=None, unit=None, formula=False):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tabname
    ws[label_cell] = label
    ws[val_cell] = '=A1*2' if formula else val
    if unit_cell:
        ws[unit_cell] = unit
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def main():
    P = F = 0

    def ck(n, c):
        nonlocal P, F
        print(("  PASS " if c else "  FAIL ") + n)
        P += bool(c)
        F += (not c)

    r = bts.extract_stone(mk("Detailed Budget ", "G15", "Tons of Stone", "H15", 4195.01, "I15", "TONS"))
    ck("ok: 4195.01 @ H15 unit TONS (POET layout)",
       r["status"] == "ok" and r["stone_tn"] == 4195.01 and r["cell"] == "H15"
       and r["unit"] == "TONS" and r["tab"] == "Detailed Budget ")

    r = bts.extract_stone(mk("DETAILED  budget", "B3", "Tons of Stone", "C3", 100, "D3", "TN"))
    ck("case/space-insensitive tab match", r["status"] == "ok" and r["stone_tn"] == 100)

    r = bts.extract_stone(mk("Budget vs Actual", "G15", "Tons of Stone", "H15", 5))
    ck("no-tab when no detailed-budget sheet", r["status"] == "no-tab" and r["stone_tn"] is None)

    r = bts.extract_stone(mk("Detailed Budget", "G15", "Tons of Rock", "H15", 5))
    ck("no-line when label absent", r["status"] == "no-line" and r["stone_tn"] is None)

    r = bts.extract_stone(mk("Detailed Budget", "G15", "Tons of Stone", "H15", None, formula=True))
    ck("formula-no-cache -> null (never fabricated)",
       r["status"] == "formula-no-cache" and r["stone_tn"] is None)

    r = bts.extract_stone(mk("Detailed Budget", "A1", "Ton of Stone", "B1", 42))
    ck("singular 'Ton of Stone' matches", r["status"] == "ok" and r["stone_tn"] == 42)

    print(f"\nRESULT: {P} passed, {F} failed")
    return 1 if F else 0


if __name__ == "__main__":
    raise SystemExit(main())
