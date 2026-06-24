#!/usr/bin/env python3
"""
PF Dashboard ingest (build-pf-dashboard.py)
===========================================
Reads the "PF Dashboard" tab of the "PF Project Master" workbook in SharePoint
and emits a structured data feed for the platform's PF Dashboard view.

The PF Dashboard tab is a company-wide MONTHLY KPI tracker (year in C3). Each
month occupies a 5-column block: Quantity[Actual,Goal], $Amount[Actual,Budget],
Delta. Metric names live in column B (rows 7-22). It mixes:
  - count metrics            (Bid Invites Received, Garbin Prelims Completed)
  - count + $ metrics        (Bids Sent, Awarded/LOI, Contracts, Completed, Billed)
  - $-only metrics           (Monthly Income, Expenses, Net Income, cash balances)
  - percent metrics          (Avg Project Profit Margin, Net Monthly Margin)

Output: platform/data/pf-dashboard.js  (window.PF_DASHBOARD = {...})

ACCESS: company-wide financials -> financials_global (admin/partner ONLY).
auth.js DATA_FILE_AREAS maps /data/pf-dashboard.js -> 'financials_global', and
index.html only injects this feed for privileged roles.

Thread caps are set BEFORE importing openpyxl (solved-box-gotchas).

Usage:
  python3 build-pf-dashboard.py            # download from SharePoint + build
  python3 build-pf-dashboard.py --local /tmp/pf-project-master.xlsx  # use a local file
"""

# --- thread caps BEFORE any heavy import ---
import os
for _v in ("OPENBLAS_NUM_THREADS", "OMP_NUM_THREADS", "MKL_NUM_THREADS",
           "NUMEXPR_NUM_THREADS", "VECLIB_MAXIMUM_THREADS"):
    os.environ.setdefault(_v, "1")

import sys
import json
import datetime
import urllib.request
import urllib.parse

HERE = os.path.dirname(os.path.abspath(__file__))
PLATFORM = os.path.dirname(HERE)
DATA_DIR = os.path.join(PLATFORM, "data")
OUT_JS = os.path.join(DATA_DIR, "pf-dashboard.js")

SOURCE_FILE = "PF Project Master.xlsx"
SOURCE_TAB = "PF Dashboard"

# Month -> starting column (1-indexed). Each month spans 5 cols:
#   +0 qtyActual, +1 qtyGoal, +2 $actual, +3 $budget, +4 delta
MONTH_START = {
    'January': 3, 'February': 8, 'March': 13, 'April': 18, 'May': 23, 'June': 28,
    'July': 33, 'August': 38, 'September': 43, 'October': 48, 'November': 53, 'December': 58,
}
MONTH_ORDER = list(MONTH_START.keys())
QUARTERS = {
    'Q1': ['January', 'February', 'March'],
    'Q2': ['April', 'May', 'June'],
    'Q3': ['July', 'August', 'September'],
    'Q4': ['October', 'November', 'December'],
}

# (rowNumber, metricName, type)
#   count       -> Quantity only (Actual vs Goal)
#   count_money -> Quantity (Actual vs Goal) AND $Amount (Actual vs Budget)
#   money_only  -> $Amount only (Actual vs Budget)
#   pct         -> $Amount column holds a percentage (Actual vs Budget)
METRICS = [
    (7,  'Bid Invites Received',      'count'),
    (8,  'Garbin Prelims Completed',  'count'),
    (9,  'Bids Sent',                 'count_money'),
    (10, 'Projects Awarded (LOI)',    'count_money'),
    (11, 'Contracts Fully Executed',  'count_money'),
    (12, 'Projects Completed',        'count_money'),
    (13, 'Projects Billed',           'count_money'),
    (14, 'Monthly Income',            'money_only'),
    (15, 'Project Expenses',          'money_only'),
    (16, 'Avg Project Profit Margin', 'pct'),
    (17, 'Non Project Expenses',      'money_only'),
    (18, 'Net Monthly Income',        'money_only'),
    (19, 'Net Monthly Margin',        'pct'),
    (20, 'Month End Chase Balance',   'money_only'),
    (21, 'Month End State Bank Bal',  'money_only'),
    (22, 'Total PF Cash @ Month End', 'money_only'),
]


def _token():
    sys.path.insert(0, "/home/aiciv/tools")
    from pf_email import _token as t
    return t()


def download_workbook():
    """Find + download PF Project Master.xlsx from SharePoint -> bytes."""
    tok = _token()
    drive = os.environ["SP_DRIVE_ID"]
    q = urllib.parse.quote(SOURCE_FILE.replace(".xlsx", ""))
    url = "https://graph.microsoft.com/v1.0/drives/%s/root/search(q='%s')" % (drive, q)
    req = urllib.request.Request(url, headers={"Authorization": "Bearer " + tok})
    res = json.load(urllib.request.urlopen(req, timeout=120))
    fid = None
    for it in res.get("value", []):
        if it.get("name") == SOURCE_FILE:
            fid = it.get("id")
            break
    if not fid:
        raise SystemExit("Could not find %r in SharePoint drive" % SOURCE_FILE)
    curl = "https://graph.microsoft.com/v1.0/drives/%s/items/%s/content" % (drive, fid)
    creq = urllib.request.Request(curl, headers={"Authorization": "Bearer " + tok})
    return urllib.request.urlopen(creq, timeout=180).read()


def num_or_none(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s == "" or s.upper() == "N/A":
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return v


def build(xlsx_bytes):
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb[SOURCE_TAB]
    year = ws.cell(row=3, column=3).value  # C3

    def cell(r, c):
        return ws.cell(row=r, column=c).value

    months = {}
    month_has_data = {}
    for m in MONTH_ORDER:
        start = MONTH_START[m]
        rows = []
        any_data = False
        for (rn, name, typ) in METRICS:
            qa = num_or_none(cell(rn, start))
            qg = num_or_none(cell(rn, start + 1))
            da = num_or_none(cell(rn, start + 2))
            db = num_or_none(cell(rn, start + 3))
            dl = num_or_none(cell(rn, start + 4))
            rows.append({
                'metric': name, 'type': typ,
                'qtyActual': qa, 'qtyGoal': qg,
                'amtActual': da, 'amtBudget': db, 'delta': dl,
            })
            if any(x is not None for x in (qa, qg, da, db)):
                any_data = True
        months[m] = {'rows': rows}
        month_has_data[m] = any_data

    # Detect whether the month columns are all identical (template / not-yet-filled).
    # If every populated month equals January, the sheet still holds seed values.
    def sig(m):
        return json.dumps(months[m]['rows'], sort_keys=True)
    populated = [m for m in MONTH_ORDER if month_has_data[m]]
    is_template = len(populated) > 1 and len({sig(m) for m in populated}) == 1

    metrics_meta = [{'metric': n, 'type': t} for (_, n, t) in METRICS]

    # The populated year (the tab we just parsed).
    populated_year = {
        'year': year,
        'isTemplateData': is_template,   # true => months are identical seed values
        'monthOrder': MONTH_ORDER,
        'quarters': QUARTERS,
        'monthHasData': month_has_data,
        'metricsMeta': metrics_meta,
        'months': months,
    }

    # MULTI-YEAR MAP (PF started 2024). Cross-year compare needs a years map; the
    # ONLY monthly-KPI source today is the single "PF Dashboard" tab (its year).
    # Prior years with no source are EMPTY placeholders (monthHasData all False);
    # they light up automatically once a prior-year KPI source is added. We do NOT
    # fabricate prior-year numbers.
    def empty_year(yy):
        emonths = {m: {'rows': [{'metric': n, 'type': t, 'qtyActual': None,
                                 'qtyGoal': None, 'amtActual': None, 'amtBudget': None,
                                 'delta': None} for (_, n, t) in METRICS]}
                   for m in MONTH_ORDER}
        return {
            'year': yy,
            'isTemplateData': False,
            'monthOrder': MONTH_ORDER,
            'quarters': QUARTERS,
            'monthHasData': {m: False for m in MONTH_ORDER},
            'metricsMeta': metrics_meta,
            'months': emonths,
        }

    years = {}
    for yy in ('2024', '2025', '2026'):
        years[yy] = populated_year if str(yy) == str(year) else empty_year(int(yy))
    # If the parsed year is outside 2024-2026 (future), still include it.
    if str(year) not in years:
        years[str(year)] = populated_year

    out = {
        'year': year,
        'sourceFile': SOURCE_FILE,
        'sourceTab': SOURCE_TAB,
        'syncedAt': datetime.datetime.now(datetime.timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
        'isTemplateData': is_template,   # true => months are identical seed values
        # Top-level fields point at the CURRENT (parsed) year so the existing
        # Week/Month/Quarter/YTD/Annual views work unchanged.
        'monthOrder': MONTH_ORDER,
        'quarters': QUARTERS,
        'monthHasData': month_has_data,
        'metricsMeta': metrics_meta,
        'months': months,
        # Multi-year map for the cross-year compare view.
        'years': years,
    }
    return out


def write_js(data):
    note = ("// AUTO-GENERATED by sync/build-pf-dashboard.py — DO NOT EDIT BY HAND.\n"
            "// Source: %s :: '%s' tab  (synced %s)\n"
            "// Company-wide MONTHLY KPI tracker. ACCESS: financials_global "
            "(admin/partner ONLY).\n"
            % (data['sourceFile'], data['sourceTab'], data['syncedAt']))
    body = "window.PF_DASHBOARD = " + json.dumps(data, indent=2) + ";\n"
    with open(OUT_JS, "w") as f:
        f.write(note + body)
    print("wrote", OUT_JS,
          "(year %s, %d metrics, template=%s)"
          % (data['year'], len(data['metricsMeta']), data['isTemplateData']))


def main():
    if "--local" in sys.argv:
        path = sys.argv[sys.argv.index("--local") + 1]
        with open(path, "rb") as f:
            xlsx = f.read()
    else:
        xlsx = download_workbook()
    data = build(xlsx)
    write_js(data)


if __name__ == "__main__":
    main()
